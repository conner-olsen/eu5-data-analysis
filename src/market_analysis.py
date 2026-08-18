"""EU5 goods market price model and building break-even workbook.

Implements the verified target price formula
    R = (1 + effective supply) / (1 + effective demand)
    target = base * min((1 + s*R) / (R + s), K)
where s = age price_stability and K = age max_price, plus the pop demand price
elasticity fixed point, and uses them to answer two questions: when is
constructing a building level net-negative for total income, and when is
turning off an existing building net-positive, counting the price impact on
every other holding the owner taxes in the same market.

Self-tests run first: the formula must reproduce the 9 in-game tooltip samples
and the closed-form equilibrium must match a fixed-point iteration across a
parameter sweep, otherwise no workbook is written.

Run: python src/market_analysis.py  ->  data/eu5_market_analysis.xlsx
"""

import math
import os
import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from parser import parse_file, parse_directory
from scraper import scrape_production_recipes, scrape_pop_demands

GAME_DIR = Path("C:/Steam/steamapps/common/Europa Universalis V/game")
COMMON_DIR = GAME_DIR / "in_game" / "common"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"

# ---- Game defines (loading_screen/common/defines/00_defines.txt) ----
MONTHLY_PRICE_CHANGE = 0.05        # L1755 NMarket.MONTHLY_PRICE_CHANGE
SD_STABILITY_OFFSET = 1.0          # L1907 NEconomy.SUPPLY_AND_DEMAND_STABILITY_OFFSET_CONSTANT
ELASTICITY_COEFF = 0.40            # L1769 NMarket.DEMAND_ELASTICITY_COEFFICIENT
ELASTICITY_FLOOR = 0.35            # L1770 NMarket.DEMAND_ELASTICITY_FLOOR
BURGHER_SUPPLY_SCALE = 0.1         # L1760 NMarket.BURGHER_TRADE_IMPACT_ON_SUPPLY_SCALE
TRADE_SUPPLY_SCALE = 0.75          # L1761 NMarket.TRADE_IMPACT_ON_SUPPLY_SCALE
BURGHER_DEMAND_SCALE = 0.25        # L1763 NMarket.BURGHER_TRADE_IMPACT_ON_DEMAND_SCALE
TRADE_DEMAND_SCALE = 0.75          # L1764 NMarket.TRADE_IMPACT_ON_DEMAND_SCALE
STOCKPILE_SUPPLY_SCALE = 0.0       # L1834 NMarket.STOCKPILE_TRADE_IMPACT_ON_SUPPLY_SCALE
CONSTRUCTION_BLOCK_FACTOR = 2.0    # L1810 NMarket.MARKET_CONSTRUCTION_NEEDS_BLOCK_FACTOR
PE_PER_LEVEL = 0.01                # L1919 NEconomy.PRODUCTION_EFFIENCY_BONUS_PER_LEVEL (sic)
RAW_MATERIAL_IN_PROVINCE_PE = 0.1  # in_game/common/auto_modifiers/country.txt:100
REFUND_FACTOR = 0.8                # L1931 NEconomy.REFUND_FACTOR (consumer unverified)
UNPROFITABLE_LAYOFF_PCT = 10       # L1909 NEconomy.UNPROFITABLE_BUILDING_WORKERS_LAID_OFF_PERCENTAGE
PROFITABLE_REHIRE_PCT = 10         # L1910 NEconomy.PROFITABLE_BUILDING_WORKERS_REHIRED_PERCENTAGE

# In-game price tooltips (London market, 2026-07-15, age 1/2: s=0.10, K=3).
# (label, base price, effective supply, effective demand, displayed target)
TOOLTIP_SAMPLES = [
    ("Alum", 3.00, 1.08, 0.00, 1.66),
    ("Livestock", 1.50, 28.90, 1.87, 0.29),
    ("Amber", 4.00, 0.00, 0.00, 4.00),
    ("Beer", 2.00, 9.75, 3.45, 0.98),
    ("Beeswax", 2.00, 7.66, 2.53, 0.97),
    ("Fruit", 1.00, 10.90, 19.52, 1.55),
    ("Wild Game", 1.00, 15.10, 22.55, 1.36),
    ("Glass", 3.00, 0.00, 3.35, 9.00),
    ("Sturdy Grains (millet)", 1.00, 5.40, 7.99, 1.32),
]
SAMPLE_STABILITY = 0.10
SAMPLE_MAX_PRICE = 3.0
# Displayed S/D and target are both rounded to 2 decimals, so worst-case
# disagreement between a recomputed target and the displayed one is ~0.011.
SAMPLE_TOLERANCE = 0.011

# ---- Styling (mirrors analyze.py / religion_analysis.py) ----
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(size=10, italic=True, color="555555")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALC_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NOTE_FONT = Font(size=9, italic=True, color="555555")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

LOC_DIRS = [
    GAME_DIR / "main_menu" / "localization" / "english",
    GAME_DIR / "in_game" / "localization" / "english",
]


def pretty(ident: str) -> str:
    return str(ident).replace("_", " ").title()


def _parse_loc(relpath: str) -> dict:
    out = {}
    for d in LOC_DIRS:
        path = d / relpath
        if not path.exists():
            continue
        for m in re.finditer(r'^\s*([\w.]+):\d*\s*"(.*)"\s*$',
                             path.read_text(encoding="utf-8-sig"), re.MULTILINE):
            out.setdefault(m.group(1), m.group(2))
    return out


GOODS_LOC = _parse_loc("goods_l_english.yml")
AGE_LOC = {k: v for k, v in _parse_loc("advances_l_english.yml").items() if k.startswith("age_")}
BUILDINGS_LOC = _parse_loc("buildings_l_english.yml")
TOWN_RIGHTS_LOC = _parse_loc("town_rights_l_english.yml")
ADVANCES_LOC = _parse_loc("advances_l_english.yml")
ALL_LOC = {**GOODS_LOC, **BUILDINGS_LOC, **TOWN_RIGHTS_LOC, **ADVANCES_LOC}


def _label(loc: dict, ident: str) -> str:
    """Display name, following a value that is only a $key$ pointing at another."""
    name = loc.get(ident, "")
    for _ in range(3):
        ref = re.fullmatch(r"\$(\w+)\$", name.strip())
        if not ref:
            break
        name = ALL_LOC.get(ref.group(1), "")
    if not name or "[" in name or "$" in name:
        return pretty(ident)
    return name


def good_label(good_id: str) -> str:
    return _label(GOODS_LOC, good_id)


# =====================================================================
# Scrapes
# =====================================================================

def scrape_ages() -> list:
    """Age price/economy parameters from in_game/common/age/00_default.txt.

    Returns ordered [{id, num, name, stability, max_price, efficiency}].
    """
    raw = parse_file(COMMON_DIR / "age" / "00_default.txt")
    ages = []
    for age_id, data in raw.items():
        if not isinstance(data, dict) or "price_stability" not in data:
            continue
        ages.append({
            "id": age_id,
            "year": data.get("year", 0),
            "stability": float(data["price_stability"]),
            "max_price": float(data["max_price"]),
            "efficiency": float(data.get("efficiency", 1.0)),
            "name": AGE_LOC.get(age_id, pretty(age_id)),
        })
    ages.sort(key=lambda a: a["year"])
    for i, a in enumerate(ages):
        a["num"] = i + 1
    return ages


def scrape_goods() -> dict:
    """All goods definitions: base price, category, method, food value, transport cost."""
    raw = parse_directory(COMMON_DIR / "goods")
    goods = {}
    for name, data in raw.items():
        if not isinstance(data, dict) or "default_market_price" not in data:
            continue
        goods[name] = {
            "base_price": float(data["default_market_price"]),
            "category": data.get("category", ""),
            "method": data.get("method", ""),
            "food": float(data["food"]) if isinstance(data.get("food"), (int, float)) else 0.0,
            "transport_cost": data.get("transport_cost", 1),
            "base_production": data.get("base_production", 0),
        }
    return goods


def scrape_shared_methods() -> dict:
    """Production methods defined in production_methods/ (village and shared input methods).

    Same recipe shape as scraper.scrape_production_recipes: {produced: [recipe, ...]}.
    """
    recipes_by_good = {}
    skip = {"produced", "output", "category", "debug_max_profit"}
    for filepath in sorted((COMMON_DIR / "production_methods").glob("*.txt")):
        if filepath.name.startswith("__"):
            continue
        raw = parse_file(filepath)
        for method_name, data in raw.items():
            if not isinstance(data, dict):
                continue
            produced, output_amt = data.get("produced"), data.get("output")
            if not produced or not isinstance(output_amt, (int, float)):
                continue
            inputs = {k: v for k, v in data.items()
                      if k not in skip and isinstance(v, (int, float)) and not isinstance(v, bool)}
            recipes_by_good.setdefault(produced, []).append({
                "method": method_name,
                "building": "(shared)",
                "tier": data.get("category", "shared"),
                "inputs": inputs,
                "output": output_amt,
            })
    return recipes_by_good


# ---- Urban rights: goods output bonuses on production buildings ----

# in_game/common/script_values/default_values.txt:1261
TOWN_RIGHT_EFFICIENCY_PENALTY = -0.05

# Production chain eras, earliest first. A location holds one member of a chain
# at a time: each tier's building lists the previous one as obsolete.
BUILDING_ERAS = ["Guild", "Workshop", "Manufactory", "Mill"]

# max_levels token -> era. Anything else is a one-off basic building (tar_kiln
# on rural_building_cap, hand_cannon_guild capped at 1) and sits in the first.
ERA_BY_MAX_LEVELS = [
    ("mills_max_level", "Mill"),
    ("manufactory_max_level", "Manufactory"),
    ("workshop_max_level", "Workshop"),
    ("guild_max_level", "Guild"),
]
# Buildings an event or a unique start hands out, so no country can simply
# build one. Dropped for the same reason a culture-restricted method is.
UNBUILDABLE_FILES = {"event_only_buildings.txt", "unique_buildings.txt"}
METHOD_SKIP_KEYS = {"produced", "output", "category", "debug_max_profit"}


def _resolve_modifier(value):
    """A modifier written as a named script value, resolved to its number."""
    if value == "town_right_efficiency_penalty":
        return TOWN_RIGHT_EFFICIENCY_PENALTY
    return value


def scrape_town_right_output_bonuses(goods_data: dict, unlocks: dict) -> list:
    """Urban rights whose location_modifier carries a goods output bonus, less
    the ones a culture or country restriction puts out of most countries' reach.

    A right is restricted when it carries a potential of its own (the Byzantine
    set) or when the advance unlocking it does (the Scandinavian set). The nine
    that survive are the production specializations off Borough Privileges.

    Returns [{id, name, file, bonuses: {good: pct}, other: {mod: value}}].
    """
    rights = []
    for filepath in sorted((COMMON_DIR / "town_rights").glob("*.txt")):
        if filepath.name == "readme.txt":
            continue
        for right_id, data in parse_file(filepath).items():
            if not isinstance(data, dict):
                continue
            loc_mod = data.get("location_modifier")
            if not isinstance(loc_mod, dict):
                continue
            bonuses, other = {}, {}
            for key, value in loc_mod.items():
                good = None
                if key.startswith("local_") and key.endswith("_output_modifier"):
                    good = key[len("local_"):-len("_output_modifier")]
                if good in goods_data and isinstance(value, (int, float)):
                    bonuses[good] = float(value)
                else:
                    other[key] = value
            if not bonuses or "potential" in data:
                continue
            if unlocks.get(right_id, {}).get("restricted", True):
                continue
            rights.append({
                "id": right_id,
                "name": _label(TOWN_RIGHTS_LOC, right_id),
                "file": filepath.name,
                "bonuses": bonuses,
                "efficiency": float(_resolve_modifier(
                    other.get("local_production_efficiency", 0.0))),
                "other": other,
            })
    return rights


def scrape_advance_unlocks(field: str) -> dict:
    """What each advance unlocks through `field` (unlock_production_method,
    unlock_town_rights).

    Returns {unlocked_id: {advance, age, restricted}}; restricted marks an
    advance whose potential limits it to certain cultures or countries.
    """
    unlocks = {}
    for filepath in sorted((COMMON_DIR / "advances").glob("*.txt")):
        if field not in filepath.read_text(encoding="utf-8-sig"):
            continue
        for advance_id, data in parse_file(filepath).items():
            if not isinstance(data, dict):
                continue
            unlocked = data.get(field)
            if not unlocked:
                continue
            for unlocked_id in (unlocked if isinstance(unlocked, list) else [unlocked]):
                unlocks[unlocked_id] = {
                    "advance": _label(ADVANCES_LOC, advance_id),
                    "age": _label(AGE_LOC, str(data.get("age", ""))),
                    "restricted": "potential" in data,
                }
    return unlocks


def scrape_producing_buildings() -> dict:
    """Every building type with production methods, slots kept separate.

    scrape_production_recipes flattens the repeated unique_production_methods
    blocks into one list; the split matters here because a second block is an
    enhancement slot that runs alongside the first and adds its own output.

    Returns {building_id: {file, era, klass, max_levels, obsolete, slots}}.
    """
    buildings = {}
    for filepath in sorted((COMMON_DIR / "building_types").glob("*.txt")):
        for building_id, data in parse_file(filepath).items():
            if not isinstance(data, dict):
                continue
            raw_slots = data.get("unique_production_methods")
            if isinstance(raw_slots, dict):
                raw_slots = [raw_slots]
            elif not isinstance(raw_slots, list):
                continue
            slots = [b for b in raw_slots if isinstance(b, dict)]
            if not slots:
                continue
            max_levels = str(data.get("max_levels", ""))
            era = next((e for token, e in ERA_BY_MAX_LEVELS if token in max_levels), None)
            buildings[building_id] = {
                "file": filepath.name,
                "era": era or BUILDING_ERAS[0],
                "klass": era or "Basic",
                "max_levels": max_levels,
                "obsolete": data.get("obsolete"),
                "slots": slots,
            }
    return buildings


def best_production(building: dict, good: str, goods_data: dict, unlocks: dict):
    """Highest-output method in every slot of one building that makes `good`.

    Output is what a goods output bonus scales, so the most valuable setup is
    the highest total output; ties break on the better margin. A second slot is
    the enhancement slot, whose output stacks on the base slot's. Methods behind
    a culture or country restricted advance are skipped, so what is left is what
    any country can run.

    Returns {output, methods, base_value, margin} or None.
    """
    base_price = goods_data[good]["base_price"]
    output, cost, methods = 0.0, 0.0, []
    for slot in building["slots"]:
        candidates = []
        for method_id, method in slot.items():
            if not isinstance(method, dict) or method.get("produced") != good:
                continue
            if unlocks.get(method_id, {}).get("restricted"):
                continue
            qty = method.get("output")
            if not isinstance(qty, (int, float)) or isinstance(qty, bool):
                continue
            inputs = {k: v for k, v in method.items()
                      if k not in METHOD_SKIP_KEYS
                      and isinstance(v, (int, float)) and not isinstance(v, bool)}
            spend = sum(goods_data.get(g, {}).get("base_price", 0) * q
                        for g, q in inputs.items())
            candidates.append((float(qty), float(qty) * base_price - spend, method_id, spend))
        if not candidates:
            continue
        qty, _margin, method_id, spend = max(candidates, key=lambda c: (c[0], c[1]))
        output += qty
        cost += spend
        methods.append(method_id)
    if not methods:
        return None
    return {
        "output": round(output, 6),
        "methods": methods,
        "base_value": round(output * base_price, 6),
        "margin": round(output * base_price - cost, 6),
    }


def urban_right_rows(rights: list, buildings: dict, goods_data: dict, unlocks: dict) -> list:
    """One row per (right, era, boosted good): the building the location holds
    in that era and the value the output bonus adds to a level of it.

    A good with no building of its own in an era keeps the newest one it does
    have, since nothing has obsoleted it (tar never advances past the kiln,
    wine stops at the manufactory).

    The right's own local_production_efficiency lands in the same bucket as the
    output bonus and the assumption is that only the boosted buildings are
    built, so the two net against each other into one effective bonus.
    """
    boosted = {good for right in rights for good in right["bonuses"]}
    best_by_good = {}
    for building_id, building in buildings.items():
        if building["file"] in UNBUILDABLE_FILES:
            continue
        for good in boosted:
            found = best_production(building, good, goods_data, unlocks)
            if found:
                best_by_good.setdefault(good, []).append((building_id, building, found))
    missing = sorted(boosted - set(best_by_good))
    if missing:
        raise SystemExit(f"No buildable producer for boosted goods: {missing}")

    rows = []
    for right in rights:
        for era_index, era in enumerate(BUILDING_ERAS):
            for good in sorted(right["bonuses"], key=good_label):
                available = [
                    (bid, b, f) for bid, b, f in best_by_good.get(good, [])
                    if BUILDING_ERAS.index(b["era"]) <= era_index
                ]
                if not available:
                    continue
                building_id, building, found = max(
                    available, key=lambda c: (c[2]["output"], c[2]["margin"]))
                bonus = right["bonuses"][good]
                gates = [unlocks[m] for m in found["methods"] if m in unlocks]
                rows.append({
                    "right_id": right["id"],
                    "right": right["name"],
                    "era": era,
                    "good": good,
                    "bonus": bonus,
                    "efficiency": right["efficiency"],
                    "net_bonus": round(bonus + right["efficiency"], 6),
                    "building_id": building_id,
                    "building": _label(BUILDINGS_LOC, building_id),
                    "klass": building["klass"],
                    "methods": ", ".join(_label(GOODS_LOC, m) for m in found["methods"]),
                    "output": found["output"],
                    "base_price": goods_data[good]["base_price"],
                    "base_value": found["base_value"],
                    "value_added": round(found["base_value"] * (bonus + right["efficiency"]), 6),
                    "margin": found["margin"],
                    "unlock": ", ".join(
                        f"{g['advance']} ({g['age']})" for g in gates),
                })
    return rows


# =====================================================================
# Price model core
# =====================================================================

def target_multiplier(R: float, s: float, cap: float) -> float:
    return min((1.0 + s * R) / (R + s), cap)


def target_price(S: float, D: float, base: float, s: float, cap: float) -> float:
    R = (SD_STABILITY_OFFSET + S) / (SD_STABILITY_OFFSET + D)
    return base * target_multiplier(R, s, cap)


def elasticity_factor(price: float, base: float) -> float:
    """Pop demand multiplier at a given current price (1.0 at or below base)."""
    if price <= base:
        return 1.0
    return max(1.0 - ELASTICITY_COEFF * (price / base - 1.0), ELASTICITY_FLOOR)


@dataclass
class EqResult:
    price: float
    demand_eff: float      # effective demand at the equilibrium price
    branch: str            # inelastic / interior / floor
    capped: bool


def solve_equilibrium(S: float, d_exo: float, pop_base: float,
                      base: float, s: float, cap: float) -> EqResult:
    """Long-run price where target(demand(P)) = P.

    d_exo = price-inelastic demand (building inputs, trade-scaled, etc.);
    pop_base = pop demand at or below base price. Closed form: on the interior
    branch pop demand linearizes to D(P) = D0 - k*P, and substituting into
    P*(R+s) = base*(1+s*R) with R = (1+S)/(1+D(P)) yields a quadratic in P.
    """
    p_inelastic = target_price(S, d_exo + pop_base, base, s, cap)
    if pop_base <= 0 or p_inelastic <= base:
        e = elasticity_factor(p_inelastic, base)
        return EqResult(p_inelastic, d_exo + pop_base * e, "inelastic", p_inelastic >= base * cap - 1e-9)

    k = ELASTICITY_COEFF * pop_base / base
    d0 = d_exo + (1.0 + ELASTICITY_COEFF) * pop_base
    qa = -s * k
    qb = (1.0 + S) + s * (1.0 + d0) + base * k
    qc = -base * ((1.0 + d0) + s * (1.0 + S))
    disc = max(qb * qb - 4.0 * qa * qc, 0.0)
    p_quad = (-qb + math.sqrt(disc)) / (2.0 * qa)

    if 1.0 - ELASTICITY_COEFF * (p_quad / base - 1.0) < ELASTICITY_FLOOR:
        candidate = target_price(S, d_exo + ELASTICITY_FLOOR * pop_base, base, s, cap)
        branch = "floor"
    else:
        candidate = p_quad
        branch = "interior"

    capped = candidate >= base * cap - 1e-9
    price = min(candidate, base * cap)
    demand = d_exo + pop_base * elasticity_factor(price, base)
    return EqResult(price, demand, branch, capped)


def solve_equilibrium_iterative(S: float, d_exo: float, pop_base: float,
                                base: float, s: float, cap: float) -> float:
    """Damped fixed-point iteration; cross-check for the closed form."""
    p = base
    for _ in range(500):
        d = d_exo + pop_base * elasticity_factor(p, base)
        nxt = target_price(S, d, base, s, cap)
        if abs(nxt - p) < 1e-10:
            return nxt
        p = 0.5 * p + 0.5 * nxt
    return p


def price_partials(S: float, d_eff: float, base: float, s: float, cap: float,
                   pop_base: float, price: float):
    """(dP/dS, dP/dD, damping lambda) at an equilibrium point; zeros when capped."""
    R = (SD_STABILITY_OFFSET + S) / (SD_STABILITY_OFFSET + d_eff)
    if (1.0 + s * R) / (R + s) >= cap - 1e-9:
        return 0.0, 0.0, 1.0
    dp_ds = base * (s * s - 1.0) / ((R + s) ** 2 * (SD_STABILITY_OFFSET + d_eff))
    dp_dd = -R * dp_ds
    k = ELASTICITY_COEFF * pop_base / base if pop_base > 0 else 0.0
    floor_binding = price > base and elasticity_factor(price, base) <= ELASTICITY_FLOOR + 1e-12
    lam = 1.0 if (k == 0.0 or price <= base or floor_binding) else 1.0 / (1.0 + k * dp_dd)
    return dp_ds, dp_dd, lam


def unpin_supply(d_eff: float, s: float, cap: float) -> float:
    """Supply level at which a price pinned at the cap starts to move."""
    r_k = (1.0 - cap * s) / (cap - s)
    return r_k * (SD_STABILITY_OFFSET + d_eff) - SD_STABILITY_OFFSET


# =====================================================================
# Self-tests
# =====================================================================

def verify_tooltip_samples() -> list:
    """Recompute the 9 in-game samples; abort-worthy if any misses."""
    rows = []
    for label, base, S, D, expected in TOOLTIP_SAMPLES:
        computed = target_price(S, D, base, SAMPLE_STABILITY, SAMPLE_MAX_PRICE)
        rows.append((label, base, S, D, expected, computed, abs(computed - expected)))
    failures = [r for r in rows if r[6] > SAMPLE_TOLERANCE]
    if failures:
        for r in failures:
            print(f"  SAMPLE FAIL: {r[0]} expected {r[4]} computed {r[5]:.4f}")
        raise SystemExit("Tooltip sample verification failed - formula or constants are wrong.")
    return rows


def verify_closed_form() -> int:
    """Closed-form equilibrium must match damped iteration across a sweep."""
    count = 0
    for S in (0.0, 0.5, 2.0, 10.0, 50.0):
        for d_exo in (0.0, 1.0, 5.0, 30.0):
            for pop_base in (0.0, 1.0, 8.0):
                for base in (1.0, 3.0):
                    for s, cap in ((0.10, 3.0), (0.06, 5.0), (0.02, 5.0)):
                        closed = solve_equilibrium(S, d_exo, pop_base, base, s, cap).price
                        iterative = solve_equilibrium_iterative(S, d_exo, pop_base, base, s, cap)
                        if abs(closed - iterative) > 1e-6:
                            raise SystemExit(
                                f"Equilibrium mismatch at S={S} Dexo={d_exo} pop={pop_base} "
                                f"base={base} s={s} K={cap}: closed {closed:.8f} vs iter {iterative:.8f}")
                        count += 1
    return count


# =====================================================================
# Decision engine
# =====================================================================

@dataclass
class GoodState:
    """One good's market state as read from its price tooltip."""
    good: str
    base: float
    eff_supply: float
    eff_demand: float           # displayed effective demand (post-elasticity)
    pop_demand: float = 0.0     # pop portion of eff_demand (displayed)
    current_price: float = None

    def decompose(self, s: float, cap: float):
        """(d_exo, pop_base): back out the price-inelastic part and the at-base pop demand."""
        cur = self.current_price
        if cur is None:
            cur = target_price(self.eff_supply, self.eff_demand, self.base, s, cap)
        e = elasticity_factor(cur, self.base)
        pop_base = self.pop_demand / e if self.pop_demand > 0 else 0.0
        return self.eff_demand - self.pop_demand, pop_base

    def equilibrium(self, s: float, cap: float, d_supply: float = 0.0,
                    d_demand: float = 0.0) -> EqResult:
        d_exo, pop_base = self.decompose(s, cap)
        return solve_equilibrium(self.eff_supply + d_supply, d_exo + d_demand,
                                 pop_base, self.base, s, cap)


@dataclass
class GoodOutcome:
    good: str
    role: str
    delta_supply: float
    delta_demand: float
    price_before: float
    price_after: float
    exposure: float
    capped_before: bool
    capped_after: bool

    @property
    def delta_price(self):
        return self.price_after - self.price_before

    @property
    def gain(self):
        return self.delta_price * self.exposure


@dataclass
class Decision:
    kind: str                   # build / close
    own_effect: float           # taxed profit change from the building itself
    outcomes: list = field(default_factory=list)
    detail: dict = field(default_factory=dict)

    @property
    def revaluation(self):
        return sum(o.gain for o in self.outcomes)

    @property
    def delta_income(self):
        return self.own_effect + self.revaluation


def marginal_output(q_out: float, level: int, pe_other: float, access: float,
                    employment: float, age_efficiency: float) -> tuple:
    """(base part, PE bump on existing levels) of the next level's output.

    Out(L) = L*q*(1 + pe_other + (L-1)*c)*A*E with c = PE_PER_LEVEL*age_efficiency,
    so Out(L+1) - Out(L) = q*A*E*((1 + pe_other) + 2*c*L).
    """
    c = PE_PER_LEVEL * age_efficiency
    base_part = q_out * access * employment * (1.0 + pe_other)
    pe_bump = q_out * access * employment * 2.0 * c * level
    return base_part, pe_bump


def evaluate_build(goods: dict, s: float, cap: float, age_efficiency: float,
                   out_good: str, q_out: float, inputs: dict, level: int,
                   pe_other: float, access: float, employment: float,
                   tax_share: float, exposures: dict,
                   construction_cost: float = 0.0) -> Decision:
    """Income change from adding one level. exposures must INCLUDE the building's existing levels."""
    base_part, pe_bump = marginal_output(q_out, level, pe_other, access, employment, age_efficiency)
    d_out = base_part + pe_bump
    d_in = {g: q * access * employment for g, q in inputs.items()}

    outcomes, price_after = [], {}
    for g, state in goods.items():
        ds = d_out if g == out_good else 0.0
        dd = d_in.get(g, 0.0)
        before = state.equilibrium(s, cap)
        after = state.equilibrium(s, cap, ds, dd)
        price_after[g] = after.price
        outcomes.append(GoodOutcome(g, "output" if g == out_good else "input",
                                    ds, dd, before.price, after.price,
                                    exposures.get(g, 0.0), before.capped, after.capped))

    unit_profit = tax_share * (d_out * price_after[out_good]
                               - sum(q * price_after[g] for g, q in d_in.items()))
    dec = Decision("build", unit_profit, outcomes)
    dec.detail = {
        "d_out": d_out, "d_out_base": base_part, "d_out_pe_bump": pe_bump,
        "d_in": d_in, "construction_cost": construction_cost,
        "payback_months": (construction_cost / dec.delta_income
                           if construction_cost > 0 and dec.delta_income > 0 else None),
    }
    return dec


def evaluate_close(goods: dict, s: float, cap: float, out_good: str,
                   out_monthly: float, in_monthly: dict, taxed_profit: float,
                   exposures: dict) -> Decision:
    """Income change from closing a building. exposures must EXCLUDE this building."""
    outcomes = []
    for g, state in goods.items():
        ds = -out_monthly if g == out_good else 0.0
        dd = -in_monthly.get(g, 0.0)
        before = state.equilibrium(s, cap)
        after = state.equilibrium(s, cap, ds, dd)
        outcomes.append(GoodOutcome(g, "output" if g == out_good else "input",
                                    ds, dd, before.price, after.price,
                                    exposures.get(g, 0.0), before.capped, after.capped))
    return Decision("close", -taxed_profit, outcomes)


# =====================================================================
# Excel helpers
# =====================================================================

def _put(ws, row, col, value, fill=None, font=None, fmt=None, align=None, border=False):
    cell = ws.cell(row=row, column=col, value=value)
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if fmt is not None:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if border:
        cell.border = THIN_BORDER
    return cell


def _title(ws, text, subtitle=None):
    _put(ws, 1, 1, text, font=TITLE_FONT)
    if subtitle:
        _put(ws, 2, 1, subtitle, font=SUBTITLE_FONT)


def _header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        _put(ws, row, start_col + i, h, fill=HEADER_FILL, font=HEADER_FONT_WHITE,
             align=CENTER, border=True)


def _section(ws, row, text, width=8):
    for c in range(1, width + 1):
        _put(ws, row, c, "", fill=SECTION_FILL)
    _put(ws, row, 1, text, fill=SECTION_FILL, font=SECTION_FONT)


def _autosize(ws, min_width=8, max_width=44):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        best = min_width
        for cell in column_cells:
            if cell.value is not None and not str(cell.value).startswith("="):
                best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = best


# The equilibrium helper chain, emitted as live formulas. Column keys in order.
CHAIN_BEFORE = [
    ("s", "s (stability)"), ("cap", "K (max mult)"), ("rnow", "R now"),
    ("tgt", "Target now"), ("move", "Price move /mo"), ("pref", "Price for elasticity"),
    ("ecur", "Pop demand factor"), ("popb", "Pop demand at base"), ("dexo", "Inelastic demand"),
    ("kq", "k"), ("d0", "D0"), ("qa", "quad a"), ("qb", "quad b"), ("qc", "quad c"),
    ("rin", "R inelastic"), ("pinel", "P inelastic"), ("pquad", "P quadratic"),
    ("equad", "e at P quad"), ("rfl", "R at floor"), ("pfloor", "P floor"),
    ("peq", "P equilibrium"), ("deff", "Demand at eq"), ("req", "R at eq"), ("capf", "Capped?"),
]
CHAIN_AFTER = [
    ("dexo2", "Inelastic demand'"), ("d02", "D0'"), ("qb2", "quad b'"), ("qc2", "quad c'"),
    ("rin2", "R inelastic'"), ("pinel2", "P inelastic'"), ("pquad2", "P quadratic'"),
    ("equad2", "e at P quad'"), ("rfl2", "R at floor'"), ("pfloor2", "P floor'"),
    ("peq2", "P equilibrium'"), ("deff2", "Demand at eq'"), ("req2", "R at eq'"), ("capf2", "Capped after?"),
]
CHAIN_PARTIALS = [
    ("dpds", "dP/dS"), ("dpdd", "dP/dD"), ("lam", "Damping"),
    ("dp1s", "dP per +1 supply"), ("dp1d", "dP per +1 demand"),
    ("unpin", "Supply headroom while capped"), ("flag", "Input check"),
]

EC = ELASTICITY_COEFF
EF = ELASTICITY_FLOOR
E1 = 1.0 + ELASTICITY_COEFF


def _emit_chain_before(ws, r, start_col, inp, age_table_ref, guard):
    """Write the before-state equilibrium chain; returns {key: cell ref}.

    inp: dict of Excel expressions for BASE, S, D, POP, CUR, AGE.
    Composite expressions must arrive parenthesized.
    """
    c = {}
    for i, (key, _label) in enumerate(CHAIN_BEFORE):
        c[key] = f"{get_column_letter(start_col + i)}{r}"
    B, S, D, POP, CUR, AGE = inp["BASE"], inp["S"], inp["D"], inp["POP"], inp["CUR"], inp["AGE"]

    f = {
        "s": f'VLOOKUP(IF({AGE}="",1,{AGE}),{age_table_ref},2,FALSE)',
        "cap": f'VLOOKUP(IF({AGE}="",1,{AGE}),{age_table_ref},3,FALSE)',
        "rnow": f'(1+{S})/(1+{D})',
        "tgt": f'{B}*MIN((1+{c["s"]}*{c["rnow"]})/({c["rnow"]}+{c["s"]}),{c["cap"]})',
        "move": f'IF({CUR}="","",{MONTHLY_PRICE_CHANGE}*({c["tgt"]}-{CUR}))',
        "pref": f'IF({CUR}="",{c["tgt"]},{CUR})',
        "ecur": f'IF({c["pref"]}<={B},1,MAX(1-{EC}*({c["pref"]}/{B}-1),{EF}))',
        "popb": f'IF({POP}<=0,0,{POP}/{c["ecur"]})',
        "dexo": f'{D}-{POP}',
        "kq": f'IF({c["popb"]}=0,0,{EC}*{c["popb"]}/{B})',
        "d0": f'{c["dexo"]}+{E1}*{c["popb"]}',
        "qa": f'-{c["s"]}*{c["kq"]}',
        "qb": f'(1+{S})+{c["s"]}*(1+{c["d0"]})+{B}*{c["kq"]}',
        "qc": f'-{B}*((1+{c["d0"]})+{c["s"]}*(1+{S}))',
        "rin": f'(1+{S})/(1+{c["dexo"]}+{c["popb"]})',
        "pinel": f'{B}*MIN((1+{c["s"]}*{c["rin"]})/({c["rin"]}+{c["s"]}),{c["cap"]})',
        "pquad": (f'IF({c["qa"]}=0,{c["pinel"]},'
                  f'(-{c["qb"]}+SQRT(MAX({c["qb"]}^2-4*{c["qa"]}*{c["qc"]},0)))/(2*{c["qa"]}))'),
        "equad": f'1-{EC}*({c["pquad"]}/{B}-1)',
        "rfl": f'(1+{S})/(1+{c["dexo"]}+{EF}*{c["popb"]})',
        "pfloor": f'{B}*MIN((1+{c["s"]}*{c["rfl"]})/({c["rfl"]}+{c["s"]}),{c["cap"]})',
        "peq": (f'IF({c["pinel"]}<={B},{c["pinel"]},'
                f'IF({c["equad"]}<{EF},{c["pfloor"]},MIN({c["pquad"]},{B}*{c["cap"]})))'),
        "deff": f'{c["dexo"]}+{c["popb"]}*IF({c["peq"]}<={B},1,MAX(1-{EC}*({c["peq"]}/{B}-1),{EF}))',
        "req": f'(1+{S})/(1+{c["deff"]})',
        "capf": f'IF((1+{c["s"]}*{c["req"]})/({c["req"]}+{c["s"]})>={c["cap"]}-0.000001,"CAP","")',
    }
    for i, (key, _label) in enumerate(CHAIN_BEFORE):
        _put(ws, r, start_col + i, f'=IF({guard},"",{f[key]})', fmt="0.0000")
    return c


def _emit_chain_after(ws, r, start_col, inp, c, guard):
    """After-state chain reusing s/cap/popb/kq/qa from the before chain.

    inp adds DS and DD (delta supply / delta inelastic demand expressions).
    """
    c2 = {}
    for i, (key, _label) in enumerate(CHAIN_AFTER):
        c2[key] = f"{get_column_letter(start_col + i)}{r}"
    B, S = inp["BASE"], inp["S"]
    S2 = f'({S}+{inp["DS"]})'

    f = {
        "dexo2": f'{c["dexo"]}+{inp["DD"]}',
        "d02": f'{c2["dexo2"]}+{E1}*{c["popb"]}',
        "qb2": f'(1+{S2})+{c["s"]}*(1+{c2["d02"]})+{B}*{c["kq"]}',
        "qc2": f'-{B}*((1+{c2["d02"]})+{c["s"]}*(1+{S2}))',
        "rin2": f'(1+{S2})/(1+{c2["dexo2"]}+{c["popb"]})',
        "pinel2": f'{B}*MIN((1+{c["s"]}*{c2["rin2"]})/({c2["rin2"]}+{c["s"]}),{c["cap"]})',
        "pquad2": (f'IF({c["qa"]}=0,{c2["pinel2"]},'
                   f'(-{c2["qb2"]}+SQRT(MAX({c2["qb2"]}^2-4*{c["qa"]}*{c2["qc2"]},0)))/(2*{c["qa"]}))'),
        "equad2": f'1-{EC}*({c2["pquad2"]}/{B}-1)',
        "rfl2": f'(1+{S2})/(1+{c2["dexo2"]}+{EF}*{c["popb"]})',
        "pfloor2": f'{B}*MIN((1+{c["s"]}*{c2["rfl2"]})/({c2["rfl2"]}+{c["s"]}),{c["cap"]})',
        "peq2": (f'IF({c2["pinel2"]}<={B},{c2["pinel2"]},'
                 f'IF({c2["equad2"]}<{EF},{c2["pfloor2"]},MIN({c2["pquad2"]},{B}*{c["cap"]})))'),
        "deff2": f'{c2["dexo2"]}+{c["popb"]}*IF({c2["peq2"]}<={B},1,MAX(1-{EC}*({c2["peq2"]}/{B}-1),{EF}))',
        "req2": f'(1+{S2})/(1+{c2["deff2"]})',
        "capf2": f'IF((1+{c["s"]}*{c2["req2"]})/({c2["req2"]}+{c["s"]})>={c["cap"]}-0.000001,"CAP","")',
    }
    for i, (key, _label) in enumerate(CHAIN_AFTER):
        _put(ws, r, start_col + i, f'=IF({guard},"",{f[key]})', fmt="0.0000")
    return c2


def _emit_partials(ws, r, start_col, inp, c, guard):
    """Marginal price sensitivity columns for the Price Calculator."""
    cp = {}
    for i, (key, _label) in enumerate(CHAIN_PARTIALS):
        cp[key] = f"{get_column_letter(start_col + i)}{r}"
    B, S = inp["BASE"], inp["S"]

    f = {
        "dpds": (f'IF({c["capf"]}="CAP",0,'
                 f'{B}*({c["s"]}^2-1)/(({c["req"]}+{c["s"]})^2*(1+{c["deff"]})))'),
        "dpdd": f'-{c["req"]}*{cp["dpds"]}',
        "lam": (f'IF(OR({c["peq"]}<={B},{c["kq"]}=0,'
                f'1-{EC}*({c["peq"]}/{B}-1)<={EF}),1,1/(1+{c["kq"]}*{cp["dpdd"]}))'),
        "dp1s": f'{cp["dpds"]}*{cp["lam"]}',
        "dp1d": f'{cp["dpdd"]}*{cp["lam"]}',
        "unpin": (f'IF({c["capf"]}="CAP",'
                  f'((1-{c["cap"]}*{c["s"]})/({c["cap"]}-{c["s"]}))*(1+{c["deff"]})-1-{S},"")'),
        "flag": f'IF({c["dexo"]}<0,"pop demand exceeds total demand","")',
    }
    for i, (key, _label) in enumerate(CHAIN_PARTIALS):
        _put(ws, r, start_col + i, f'=IF({guard},"",{f[key]})', fmt="0.0000")
    return cp


# =====================================================================
# Sheets
# =====================================================================

def build_readme(wb):
    ws = wb.active
    ws.title = "Read Me"
    _title(ws, "EU5 Market Price and Building Break-Even Analysis",
           "Answers: what will a good's price settle at, is another building level worth it, should an existing building be turned off, and what a goods output Urban Right is worth.")
    lines = [
        "",
        "THE FORMULA (verified against 9 in-game tooltips, see Price Model sheet)",
        "    R = (1 + Effective Supply) / (1 + Effective Demand)",
        "    Target Price = Base Price * min( (1 + s*R) / (R + s), K )",
        "    s = age Price Stability, K = age max price multiplier. Current price moves 5% of the gap to target per month.",
        "",
        "WHERE EACH INPUT COMES FROM IN-GAME",
        "    Base price, Effective Supply, Effective Demand, Price Stability: hover a good's price in the market panel (the Target Price tooltip).",
        "    'Of which pop demand': the same tooltip's demand breakdown hover lists pop demand rows; their sum. Optional - leave 0 to treat all demand as price-insensitive.",
        "    Current price: the market panel price. Used to back out how much pops have already cut their demand ('Due to price, Pop demands -X%').",
        "    Building recipe quantities: the building panel's production method line (per-level input/output).",
        "    Building monthly flows and taxed profit: the building tooltip - 'contributing to the tax base with X, of which we tax Y'. Y is the taxed profit; Y/X is your tax share t.",
        "    Your exposure per good: you own the whole market (the standing assumption), so: OUTPUT good X = t * (Eff Supply - your buildings' consumption of it),",
        "    the slice sold to pops plus any overproduction. INPUT goods you produce yourself: X = t * (production - building consumption), usually near 0.",
        "",
        "THE QUICK RULE (full market ownership)",
        "    Margin must beat 0.8 x outward share.",
        "    Margin = building profit / (output * price). Outward share = fraction of the output NOT consumed by your own buildings.",
        "    Intermediates your industry eats: threshold ~0 - keep or build anything profitable. Input price swings cancel across your book.",
        "    Pop end-goods: threshold ~80% margin - add supply only while margins are fat (scarce, above-base goods); below it, close the worst levels first.",
        "    Below base price pops do NOT buy more when it gets cheaper, so oversupplying a pop good is pure self-harm.",
        "    Within about 2x of the threshold, or for big changes: use the calculators.",
        "",
        "URBAN RIGHTS (Urban Rights Ranking / Urban Rights Detail sheets)",
        "    Every Urban Right that grants a goods output bonus, ranked by the gold per month it adds per building level.",
        "    Culture and country restricted rights are left out, so what is ranked is the production specialization set off the Borough Privileges advance.",
        "    An output bonus lifts output only, never throughput, so the recipe's inputs never enter it.",
        "    Production efficiency is the same bucket, so each right's own -5% nets against its bonus: a 20% output bonus is a 15% bonus. Value added = output x price x net bonus, assuming only the boosted buildings are built.",
        "    Ranked once per production era (Guild, Workshop, Manufactory, Mill), because the boosted building grows through the chain while the bonus stays fixed.",
        "    Each building runs its highest-output production method plus its enhancement slot where it has one, skipping methods behind a culture or country restricted advance.",
        "    Scored: the output bonus net of the efficiency penalty. Everything else a right does (Jewelry Rights' +10 guild levels) sits unvalued in its own column.",
        "    Dyes and wine are raw materials, so RGOs produce them too and the same modifier lifts that output; RGO size is location-specific, so only the buildings are counted here.",
        "",
        "THE THREE CALCULATORS",
        "    Price Calculator: one good per row. Gives target now, the long-run equilibrium price (pop demand elasticity included), and marginal price sensitivities.",
        "    Build Calculator: one scenario per block. Verdict on adding one building level, including the price displacement on your own holdings and payback time.",
        "    Turn-Off Calculator: one scenario per block. Verdict on closing a building: close when your other holdings gain more from better prices than the building earns.",
        "",
        "READING THE VERDICT MATH",
        "    Build delta income = t * (new level's output * new price - new level's inputs * new prices)  +  sum over goods of (price change * your exposure).",
        "    Close delta income = - (building's taxed profit)  +  sum over goods of (price change * your exposure excluding it).",
        "    Exposure sign: positive = net producer of the good (an output price rise helps you), negative = net consumer.",
        "",
        "CAVEATS",
        "    All effects are per-market: only holdings in the same market as the building count, and prices in other markets are untouched.",
        "    Full ownership is assumed (you own every location, RGO, and building in the market); with foreign owners present, shrink exposures to just your own holdings.",
        "    Gold income only: cheap goods still feed pop food and satisfaction, which can be a deliberate reason to oversupply anyway.",
        "    Steady-state: prices take months to converge (5% of the gap per month, half-life ~13.5 months). Verdicts compare long-run monthly income.",
        "    Trade is held fixed: burgher/manual trade contributions to supply and demand are treated as constant while your change settles.",
        "    Employment ramps: closing lays off pops gradually (10%/month) and unprofitable buildings shed workers at the same rate on their own.",
        "    Food goods: the food system perturbs food-category goods; the formula matched all food-good samples, but treat food verdicts as slightly softer.",
        "",
        "VALIDATE IN-GAME (optional)",
        "    1. Pick a small market. Note a good's tooltip (S, D, target). Close one building producing it.",
        "    2. Re-read the tooltip: Effective Supply should fall by the building's displayed output, and the new target should match the Price Calculator.",
        "    3. Next month's price move should be 5% of (target - current).",
        "",
        "Regenerate with: python src/market_analysis.py (self-tests re-verify the formula against the 9 samples first).",
    ]
    for i, line in enumerate(lines):
        font = None
        if line and not line.startswith("    "):
            font = Font(bold=True, size=11)
        _put(ws, 3 + i, 1, line, font=font)
    ws.column_dimensions["A"].width = 160


def build_price_model(wb, ages):
    """Formula reference, defines, age table (the VLOOKUP source), verification table.

    Returns the age-table range reference used by every calculator sheet.
    """
    ws = wb.create_sheet("Price Model")
    _title(ws, "Price Model",
           "The formula, every constant it uses (with game file sources), and the live verification against the 9 in-game tooltips.")

    r = 4
    _section(ws, r, "FORMULA", 10)
    formula_lines = [
        "R = (1 + Effective Supply) / (1 + Effective Demand)          the +1s are SUPPLY_AND_DEMAND_STABILITY_OFFSET_CONSTANT",
        "Target Price = Base Price * min( (1 + s*R) / (R + s), K )    s = age price stability, K = age max price",
        "Supply glut: R -> infinity pushes the multiplier to s (10% of base early game) - an asymptote, never quite reached.",
        "Demand glut: R -> 0 pushes it to 1/s, but the age max price K clamps first (300% in ages 1-2, up to 500% later).",
        "S = D gives exactly base price. Zero supply AND zero demand also gives base price (Amber sample).",
        "Pop demand shrinks above base price: multiplier max(1 - 0.40*(price/base - 1), 0.35), evaluated at the CURRENT price.",
        "Effective values fold trade in at reduced weight: supply + 0.1*burgher imports + 0.75*manual imports; demand + 0.25*burgher exports + 0.75*manual exports.",
        "Current price moves 5% of (target - current) each month.",
    ]
    for line in formula_lines:
        r += 1
        _put(ws, r, 1, line)

    r += 2
    _section(ws, r, "AGE TABLE (calculators look s, K, and the level-PE efficiency up here)", 10)
    r += 1
    _header_row(ws, r, ["Age #", "Price stability s", "Max price K",
                        "Efficiency (scales +1%/level PE)", "Age", "R where cap binds"])
    age_first_row = r + 1
    for a in ages:
        r += 1
        _put(ws, r, 1, a["num"], border=True, align=CENTER)
        _put(ws, r, 2, a["stability"], fmt="0%", border=True, align=CENTER)
        _put(ws, r, 3, a["max_price"], fmt="0", border=True, align=CENTER)
        _put(ws, r, 4, a["efficiency"], fmt="0.0", border=True, align=CENTER)
        _put(ws, r, 5, a["name"], border=True)
        _put(ws, r, 6, f"=(1-C{r}*B{r})/(C{r}-B{r})", fmt="0.0000", border=True, align=CENTER)
    age_table_ref = f"'Price Model'!$A${age_first_row}:$D${r}"

    r += 2
    _section(ws, r, "VERIFICATION - the 9 in-game tooltip samples (Excel column is a live formula)", 10)
    r += 1
    _put(ws, r, 1, "Stability s:", font=Font(bold=True))
    s_cell = f"$B${r}"
    _put(ws, r, 2, SAMPLE_STABILITY, fmt="0%", fill=CALC_FILL)
    _put(ws, r, 3, "Max price K:", font=Font(bold=True))
    k_cell = f"$D${r}"
    _put(ws, r, 4, SAMPLE_MAX_PRICE, fmt="0", fill=CALC_FILL)
    _put(ws, r, 5, "(the samples are all from ages 1-2)", font=NOTE_FONT)
    r += 1
    _header_row(ws, r, ["Good", "Base", "Eff Supply", "Eff Demand", "In-game target",
                        "Python target", "Excel target (live)", "Diff", "Check"])
    sample_rows = verify_tooltip_samples()
    for label, base, S, D, expected, computed, _diff in sample_rows:
        r += 1
        _put(ws, r, 1, label, border=True)
        _put(ws, r, 2, base, fmt="0.00", border=True, align=CENTER)
        _put(ws, r, 3, S, fmt="0.00", border=True, align=CENTER)
        _put(ws, r, 4, D, fmt="0.00", border=True, align=CENTER)
        _put(ws, r, 5, expected, fmt="0.00", border=True, align=CENTER)
        _put(ws, r, 6, round(computed, 4), fmt="0.0000", border=True, align=CENTER)
        _put(ws, r, 7,
             f"=B{r}*MIN((1+{s_cell}*((1+C{r})/(1+D{r})))/(((1+C{r})/(1+D{r}))+{s_cell}),{k_cell})",
             fmt="0.0000", border=True, align=CENTER)
        _put(ws, r, 8, f"=G{r}-E{r}", fmt="0.0000", border=True, align=CENTER)
        _put(ws, r, 9, f'=IF(ABS(G{r}-E{r})<={SAMPLE_TOLERANCE},"PASS","FAIL")',
             border=True, align=CENTER)

    r += 2
    _section(ws, r, "DEFINES (loading_screen/common/defines/00_defines.txt unless noted)", 10)
    r += 1
    _header_row(ws, r, ["Define", "Value", "Source line", "Meaning"])
    defines = [
        ("MONTHLY_PRICE_CHANGE", MONTHLY_PRICE_CHANGE, "L1755", "Fraction of (target - current) closed per month"),
        ("SUPPLY_AND_DEMAND_STABILITY_OFFSET_CONSTANT", SD_STABILITY_OFFSET, "L1907", "The +1 added to supply and demand in R"),
        ("DEMAND_ELASTICITY_COEFFICIENT", ELASTICITY_COEFF, "L1769", "Pop demand drops 40% per 100% above base price"),
        ("DEMAND_ELASTICITY_FLOOR", ELASTICITY_FLOOR, "L1770", "Pop demand never drops below 35% of its base"),
        ("BURGHER_TRADE_IMPACT_ON_SUPPLY_SCALE", BURGHER_SUPPLY_SCALE, "L1760", "Burgher imports count 10% toward effective supply"),
        ("TRADE_IMPACT_ON_SUPPLY_SCALE", TRADE_SUPPLY_SCALE, "L1761", "Manual trade imports count 75%"),
        ("BURGHER_TRADE_IMPACT_ON_DEMAND_SCALE", BURGHER_DEMAND_SCALE, "L1763", "Burgher exports count 25% toward effective demand"),
        ("TRADE_IMPACT_ON_DEMAND_SCALE", TRADE_DEMAND_SCALE, "L1764", "Manual trade exports count 75%"),
        ("STOCKPILE_TRADE_IMPACT_ON_SUPPLY_SCALE", STOCKPILE_SUPPLY_SCALE, "L1834", "Stockpile bleed never suppresses price"),
        ("MARKET_CONSTRUCTION_NEEDS_BLOCK_FACTOR", CONSTRUCTION_BLOCK_FACTOR, "L1810", "Construction blocked when demand >= 2x supply of a needed good"),
        ("PRODUCTION_EFFIENCY_BONUS_PER_LEVEL", PE_PER_LEVEL, "L1919", "+1% PE per building level beyond the first, scaled by age efficiency"),
        ("raw_material_in_province_impact", RAW_MATERIAL_IN_PROVINCE_PE, "auto_modifiers/country.txt:100", "+10% PE when the required raw material is made in the same province"),
        ("UNPROFITABLE_BUILDING_WORKERS_LAID_OFF_PERCENTAGE", UNPROFITABLE_LAYOFF_PCT, "L1909", "Unprofitable unsubsidized buildings shed 10% workers/month"),
        ("PROFITABLE_BUILDING_WORKERS_REHIRED_PERCENTAGE", PROFITABLE_REHIRE_PCT, "L1910", "Profitable buildings rehire 10%/month"),
        ("REFUND_FACTOR", REFUND_FACTOR, "L1931", "Refund factor (consumer unverified; likely destroy/cancel)"),
        ("MIN_PRICE_IMPACT / MAX_PRICE_IMPACT", "-0.33 / 3.0", "L1766-1767", "Clamp on price-scaled costs (construction), NOT on the target price"),
    ]
    for name, val, src, meaning in defines:
        r += 1
        _put(ws, r, 1, name, border=True)
        _put(ws, r, 2, val, border=True, align=CENTER)
        _put(ws, r, 3, src, border=True, align=CENTER)
        _put(ws, r, 4, meaning, border=True)

    _autosize(ws)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["D"].width = 70
    return age_table_ref


def build_price_calculator(wb, age_table_ref):
    ws = wb.create_sheet("Price Calculator")
    _title(ws, "Price Calculator",
           "One good per row. Yellow = type it in from the price tooltip. Everything else computes live. Rows 6-14 are prefilled with the London samples.")
    _put(ws, 3, 1, "Long-run equilibrium includes pop demand elasticity: it is where the price would settle if supply and non-pop demand stayed fixed. "
                   "'Of which pop demand' is optional (0 = treat all demand as fixed; overstates price swings for pop goods above base price).",
         font=NOTE_FONT)

    inputs = ["Good", "Age (1-6)", "Base price", "Eff Supply", "Eff Demand",
              "of which pop demand", "Current price"]
    outputs = ["Target now", "Move /mo", "Long-run price", "vs base", "Capped?",
               "dP per +1 supply", "dP per +1 demand", "Supply headroom while capped", "Check"]
    n_in = len(inputs)
    header = inputs + outputs + ["|"] + [lbl for _k, lbl in CHAIN_BEFORE] + [lbl for _k, lbl in CHAIN_PARTIALS]
    _header_row(ws, 5, header)
    _put(ws, 4, n_in + len(outputs) + 2, "internal machinery - ignore", font=NOTE_FONT)

    first, last = 6, 6 + 24
    chain_start = n_in + len(outputs) + 2
    for r in range(first, last + 1):
        for col in range(1, n_in + 1):
            fmt = "0.00" if col >= 3 else "0"
            _put(ws, r, col, None, fill=INPUT_FILL, fmt=fmt, border=True)
        inp = {"BASE": f"$C{r}", "S": f"$D{r}", "D": f"$E{r}", "POP": f"$F{r}",
               "CUR": f"$G{r}", "AGE": f"$B{r}"}
        guard = f'$C{r}=""'
        c = _emit_chain_before(ws, r, chain_start, inp, age_table_ref, guard)
        cp = _emit_partials(ws, r, chain_start + len(CHAIN_BEFORE), inp, c, guard)
        out_cells = [
            (f'=IF({guard},"",{c["tgt"]})', "0.000"),
            (f'=IF({guard},"",{c["move"]})', "+0.000;-0.000;0"),
            (f'=IF({guard},"",{c["peq"]})', "0.000"),
            (f'=IF({guard},"",{c["peq"]}/$C{r}-1)', "+0%;-0%;0%"),
            (f'=IF({guard},"",{c["capf"]})', "@"),
            (f'=IF({guard},"",{cp["dp1s"]})', "+0.0000;-0.0000;0"),
            (f'=IF({guard},"",{cp["dp1d"]})', "+0.0000;-0.0000;0"),
            (f'=IF({guard},"",{cp["unpin"]})', "0.00"),
            (f'=IF({guard},"",{cp["flag"]})', "@"),
        ]
        for i, (formula, fmt) in enumerate(out_cells):
            _put(ws, r, n_in + 1 + i, formula, fill=CALC_FILL, fmt=fmt, border=True, align=CENTER)

    prefills = [(t[0], 1, t[1], t[2], t[3], None, None) for t in TOOLTIP_SAMPLES]
    prefills.append(("Fruit (with pop demand)", 1, 1.00, 10.90, 19.52, 15.00, 1.34))
    for i, (label, age, base, S, D, pop, cur) in enumerate(prefills):
        r = first + i
        _put(ws, r, 1, label, fill=INPUT_FILL, border=True)
        _put(ws, r, 2, age, fill=INPUT_FILL, border=True, fmt="0")
        _put(ws, r, 3, base, fill=INPUT_FILL, border=True, fmt="0.00")
        _put(ws, r, 4, S, fill=INPUT_FILL, border=True, fmt="0.00")
        _put(ws, r, 5, D, fill=INPUT_FILL, border=True, fmt="0.00")
        if pop is not None:
            _put(ws, r, 6, pop, fill=INPUT_FILL, border=True, fmt="0.00")
        if cur is not None:
            _put(ws, r, 7, cur, fill=INPUT_FILL, border=True, fmt="0.00")

    ws.freeze_panes = "A6"
    for col in range(1, n_in + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["A"].width = 20
    for col in range(n_in + 1, n_in + len(outputs) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for col in range(chain_start, chain_start + len(CHAIN_BEFORE) + len(CHAIN_PARTIALS)):
        ws.column_dimensions[get_column_letter(col)].width = 10


GOODS_TABLE_HEADERS = ["Role", "Good", "Base price", "Eff Supply", "Eff Demand",
                       "of which pop demand", "Current price"]


def _emit_goods_row(ws, r, age_cell, ds_expr, dd_expr, age_table_ref, chain_start):
    """One goods row of a Build/Turn-Off block: inputs, deltas, before/after chains, dP, gain."""
    for col in range(2, 10):
        fmt = "0.00" if col >= 3 else "@"
        _put(ws, r, col, None, fill=INPUT_FILL, fmt=fmt, border=True)
    inp = {"BASE": f"$C{r}", "S": f"$D{r}", "D": f"$E{r}", "POP": f"$F{r}",
           "CUR": f"$G{r}", "AGE": age_cell, "DS": ds_expr, "DD": dd_expr}
    guard = f'$C{r}=""'
    _put(ws, r, 10, f'=IF({guard},"",{ds_expr})', fill=CALC_FILL, fmt="+0.000;-0.000;0", border=True, align=CENTER)
    _put(ws, r, 11, f'=IF({guard},"",{dd_expr})', fill=CALC_FILL, fmt="+0.000;-0.000;0", border=True, align=CENTER)
    c = _emit_chain_before(ws, r, chain_start, inp, age_table_ref, guard)
    c2 = _emit_chain_after(ws, r, chain_start + len(CHAIN_BEFORE), inp, c, guard)
    _put(ws, r, 12, f'=IF({guard},"",{c["peq"]})', fill=CALC_FILL, fmt="0.000", border=True, align=CENTER)
    _put(ws, r, 13, f'=IF({guard},"",{c2["peq2"]})', fill=CALC_FILL, fmt="0.000", border=True, align=CENTER)
    _put(ws, r, 14, f'=IF({guard},"",{c2["peq2"]}-{c["peq"]})', fill=CALC_FILL,
         fmt="+0.0000;-0.0000;0", border=True, align=CENTER)
    _put(ws, r, 15, f'=IF({guard},"",({c2["peq2"]}-{c["peq"]})*$I{r})', fill=CALC_FILL,
         fmt="+0.0000;-0.0000;0", border=True, align=CENTER)
    _put(ws, r, 16, f'=IF({guard},"",IF(OR({c["capf"]}="CAP",{c2["capf2"]}="CAP"),"price at cap",""))',
         fill=CALC_FILL, border=True, align=CENTER)
    return c, c2


def _goods_table_header(ws, r, qty_label, exposure_label, chain_start):
    headers = GOODS_TABLE_HEADERS + [qty_label, exposure_label, "dSupply", "dDemand",
                                     "Price before", "Price after", "dPrice", "Your gain /mo", "Note"]
    _header_row(ws, r, headers)
    chain_headers = [lbl for _k, lbl in CHAIN_BEFORE] + [lbl for _k, lbl in CHAIN_AFTER]
    for i, h in enumerate(chain_headers):
        _put(ws, r, chain_start + i, h, fill=CALC_FILL, font=Font(size=8, color="888888"),
             align=CENTER)
    _put(ws, r - 1, chain_start, "internal machinery - ignore", font=NOTE_FONT)


def _label_value(ws, r, col, label, value, fmt="0.00", input_cell=True, note=None):
    _put(ws, r, col, label, font=Font(bold=True))
    cell = _put(ws, r, col + 1, value, fill=INPUT_FILL if input_cell else CALC_FILL,
                fmt=fmt, border=True, align=CENTER)
    if note:
        _put(ws, r, col + 2, note, font=NOTE_FONT)
    return cell.coordinate


def build_build_calculator(wb, age_table_ref):
    ws = wb.create_sheet("Build Calculator")
    _title(ws, "Build Calculator",
           "Should you add one more level? Yellow = your inputs. Exposure must INCLUDE this building's existing levels (they suffer the price move too).")
    chain_start = 18

    def block(r0, title, prefill):
        _section(ws, r0, title, 16)
        _put(ws, r0 + 1, 1, "Market state per good (price tooltip) + your taxed net exposure. "
             "Output row: the good this building makes. Input rows: leave blank if unused.",
             font=NOTE_FONT)

        params_r = r0 + 9
        age_cell = _label_value(ws, params_r, 1, "Age (1-6)", prefill["age"], "0",
                                note="drives price stability, max price, and the per-level PE efficiency")
        lvl_cell = _label_value(ws, params_r + 1, 1, "Current level L", prefill["L"], "0",
                                note="0 = building does not exist yet")
        pe_cell = _label_value(ws, params_r + 2, 1, "PE from other sources", prefill["pe"], "0.0%",
                               note="everything except the per-level bonus: province raw material, modifiers, market-supplied inputs")
        ma_cell = _label_value(ws, params_r + 3, 1, "Market access", prefill["ma"], "0.0%")
        emp_cell = _label_value(ws, params_r + 4, 1, "Employment fill", prefill["emp"], "0.0%")
        tax_cell = _label_value(ws, params_r + 5, 1, "Your tax share t", prefill["tax"], "0.0%",
                                note="building tooltip: taxed income / taxable profit")
        cost_cell = _label_value(ws, params_r + 6, 1, "Construction gold cost", prefill["cost"], "0",
                                 note="optional, for payback")

        eff_r = params_r
        _put(ws, eff_r, 5, "Age efficiency", font=Font(bold=True))
        eff_cell = f"$F${eff_r}"
        _put(ws, eff_r, 6, f'=VLOOKUP(IF({age_cell}="",1,{age_cell}),{age_table_ref},4,FALSE)',
             fill=CALC_FILL, fmt="0.0", border=True, align=CENTER)
        _put(ws, eff_r + 1, 5, "c = PE per level", font=Font(bold=True))
        c_cell = f"$F${eff_r + 1}"
        _put(ws, eff_r + 1, 6, f"={PE_PER_LEVEL}*{eff_cell}", fill=CALC_FILL, fmt="0.000%",
             border=True, align=CENTER)

        table_r = r0 + 3
        _goods_table_header(ws, table_r, "Recipe qty /level", "Your taxed net exposure /mo (incl. this bld)", chain_start)
        out_r = table_r + 1
        in_rows = [table_r + 2, table_r + 3, table_r + 4]

        qout_ref = f'IF($H${out_r}="",0,$H${out_r})'
        dout_base_expr = f"({qout_ref}*{ma_cell}*{emp_cell}*(1+{pe_cell}))"
        dout_bump_expr = f"({qout_ref}*{ma_cell}*{emp_cell}*2*{c_cell}*{lvl_cell})"
        dout_expr = f"({dout_base_expr}+{dout_bump_expr})"

        _put(ws, out_r, 1, "OUTPUT", font=Font(bold=True), border=True)
        _emit_goods_row(ws, out_r, age_cell, dout_expr, "0", age_table_ref, chain_start)
        for i, rr in enumerate(in_rows):
            _put(ws, rr, 1, f"INPUT {i + 1}", font=Font(bold=True), border=True)
            din_expr = f'(IF($H{rr}="",0,$H{rr})*{ma_cell}*{emp_cell})'
            _emit_goods_row(ws, rr, age_cell, "0", din_expr, age_table_ref, chain_start)

        _put(ws, eff_r + 2, 5, "New level output: base part", font=Font(bold=True))
        _put(ws, eff_r + 2, 6, f'={dout_base_expr}', fill=CALC_FILL, fmt="0.000", border=True, align=CENTER)
        _put(ws, eff_r + 3, 5, "New level output: PE bump on existing levels", font=Font(bold=True))
        _put(ws, eff_r + 3, 6, f'={dout_bump_expr}', fill=CALC_FILL, fmt="0.000", border=True, align=CENTER)
        _put(ws, eff_r + 4, 5, "New level output total (dSupply)", font=Font(bold=True))
        _put(ws, eff_r + 4, 6, f'={dout_expr}', fill=CALC_FILL, fmt="0.000", border=True, align=CENTER)

        res_r = params_r + 8
        _section(ws, res_r, "RESULT", 16)
        in_terms = "+".join(f'IF($K{rr}="",0,IF($K{rr}=0,0,$K{rr}*$M{rr}))' for rr in in_rows)
        _put(ws, res_r + 1, 1, "New level taxed profit /mo", font=Font(bold=True))
        marg = _put(ws, res_r + 1, 2,
                    f'=IF($C{out_r}="","",{tax_cell}*($J{out_r}*$M{out_r}-({in_terms})))',
                    fill=CALC_FILL, fmt="+0.000;-0.000;0", border=True, align=CENTER).coordinate
        _put(ws, res_r + 1, 3, "at post-build prices", font=NOTE_FONT)
        _put(ws, res_r + 2, 1, "Price displacement on your exposure /mo", font=Font(bold=True))
        reval = _put(ws, res_r + 2, 2, f'=SUM($O{out_r}:$O{in_rows[-1]})',
                     fill=CALC_FILL, fmt="+0.000;-0.000;0", border=True, align=CENTER).coordinate
        _put(ws, res_r + 3, 1, "DELTA INCOME /mo", font=Font(bold=True, size=12))
        dinc = _put(ws, res_r + 3, 2, f'=IF({marg}="","",{marg}+{reval})', fill=CALC_FILL,
                    fmt="+0.000;-0.000;0", border=True, align=CENTER).coordinate
        _put(ws, res_r + 4, 1, "VERDICT", font=Font(bold=True, size=12))
        _put(ws, res_r + 4, 2, f'=IF({dinc}="","",IF({dinc}>0,"BUILD","DO NOT BUILD"))',
             fill=CALC_FILL, border=True, align=CENTER)
        _put(ws, res_r + 5, 1, "Payback (months)", font=Font(bold=True))
        _put(ws, res_r + 5, 2,
             f'=IF(AND(ISNUMBER({dinc}),{dinc}>0,{cost_cell}>0),{cost_cell}/{dinc},"n/a")',
             fill=CALC_FILL, fmt="0.0", border=True, align=CENTER)
        _put(ws, res_r + 6, 1, "Notes", font=Font(bold=True))
        _put(ws, res_r + 6, 2,
             '"price at cap" in a goods row = that price is pinned at the age max and will not move until supply grows enough (see Price Calculator headroom).',
             font=NOTE_FONT)

        con_r = res_r + 8
        _put(ws, con_r, 1, "Construction goods check (optional): a needed good with demand >= 2x supply blocks the build.",
             font=Font(bold=True))
        _header_row(ws, con_r + 1, ["Construction good", "Eff Supply", "Eff Demand", "Status"])
        for rr in (con_r + 2, con_r + 3):
            for col in (1, 2, 3):
                _put(ws, rr, col, None, fill=INPUT_FILL, fmt="0.00" if col > 1 else "@", border=True)
            _put(ws, rr, 4,
                 f'=IF($B{rr}="","",IF($C{rr}>={CONSTRUCTION_BLOCK_FACTOR}*$B{rr},"BLOCKED",'
                 f'IF($C{rr}>=1.6*$B{rr},"NEAR BLOCK","ok")))',
                 fill=CALC_FILL, border=True, align=CENTER)

        for rr, vals in prefill["goods"].items():
            row = {"out": out_r, "in1": in_rows[0], "in2": in_rows[1], "in3": in_rows[2]}[rr]
            for col, v in vals.items():
                _put(ws, row, col, v, fill=INPUT_FILL, border=True,
                     fmt="0.00" if col >= 3 else "@")
        return con_r + 5

    end1 = block(5, "SCENARIO 1 - dominant producer adds a level (prefilled: worked example B)", {
        "age": 2, "L": 12, "pe": 0.1, "ma": 1.0, "emp": 1.0, "tax": 0.55, "cost": 300,
        "goods": {
            "out": {2: "cloth", 3: 3.0, 4: 16.0, 5: 13.95, 6: 11.0, 7: 2.7, 8: 1.0, 9: 7.18},
            "in1": {2: "cotton", 3: 3.0, 4: 12.19, 5: 13.0, 6: 2.0, 7: 3.15, 8: 0.8333, 9: 0.65},
        },
    })
    block(end1 + 2, "SCENARIO 2 - blank", {
        "age": 1, "L": 0, "pe": 0.0, "ma": 1.0, "emp": 1.0, "tax": 0.5, "cost": 0,
        "goods": {},
    })

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    for col in range(3, 17):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for col in range(chain_start, chain_start + len(CHAIN_BEFORE) + len(CHAIN_AFTER)):
        ws.column_dimensions[get_column_letter(col)].width = 10


def build_turnoff_calculator(wb, age_table_ref):
    ws = wb.create_sheet("Turn-Off Calculator")
    _title(ws, "Turn-Off Calculator",
           "Should you close a building? Close when your OTHER holdings gain more from the price shift than this building earns you. Exposure must EXCLUDE this building.")
    chain_start = 18

    def block(r0, title, prefill):
        _section(ws, r0, title, 16)
        _put(ws, r0 + 1, 1, "Qty /mo = this building's actual monthly output/consumption from the building panel "
             "(already includes level, PE, market access, employment).", font=NOTE_FONT)

        params_r = r0 + 9
        age_cell = _label_value(ws, params_r, 1, "Age (1-6)", prefill["age"], "0")
        profit_cell = _label_value(ws, params_r + 1, 1, "This building's taxed profit /mo",
                                   prefill["profit"], "0.000",
                                   note="building tooltip: 'of which we tax X' - negative if losing money")

        table_r = r0 + 3
        _goods_table_header(ws, table_r, "Qty /mo (this bld)", "Your taxed net exposure /mo (excl. this bld)", chain_start)
        out_r = table_r + 1
        in_rows = [table_r + 2, table_r + 3, table_r + 4]

        _put(ws, out_r, 1, "OUTPUT", font=Font(bold=True), border=True)
        _emit_goods_row(ws, out_r, age_cell, f'(-IF($H{out_r}="",0,$H{out_r}))', "0",
                        age_table_ref, chain_start)
        for i, rr in enumerate(in_rows):
            _put(ws, rr, 1, f"INPUT {i + 1}", font=Font(bold=True), border=True)
            _emit_goods_row(ws, rr, age_cell, "0", f'(-IF($H{rr}="",0,$H{rr}))', age_table_ref, chain_start)

        res_r = params_r + 3
        _section(ws, res_r, "RESULT", 16)
        _put(ws, res_r + 1, 1, "Price displacement gain on your other holdings /mo", font=Font(bold=True))
        reval = _put(ws, res_r + 1, 2, f'=SUM($O{out_r}:$O{in_rows[-1]})', fill=CALC_FILL,
                     fmt="+0.000;-0.000;0", border=True, align=CENTER).coordinate
        _put(ws, res_r + 2, 1, "Lost taxed profit /mo", font=Font(bold=True))
        _put(ws, res_r + 2, 2, f'=-{profit_cell}', fill=CALC_FILL, fmt="+0.000;-0.000;0",
             border=True, align=CENTER)
        _put(ws, res_r + 3, 1, "DELTA INCOME if closed /mo", font=Font(bold=True, size=12))
        dinc = _put(ws, res_r + 3, 2, f'=IF($C{out_r}="","",{reval}-{profit_cell})', fill=CALC_FILL,
                    fmt="+0.000;-0.000;0", border=True, align=CENTER).coordinate
        _put(ws, res_r + 4, 1, "VERDICT", font=Font(bold=True, size=12))
        _put(ws, res_r + 4, 2, f'=IF({dinc}="","",IF({dinc}>0,"CLOSE","KEEP OPEN"))',
             fill=CALC_FILL, border=True, align=CENTER)
        _put(ws, res_r + 5, 1, "Destroying instead", font=Font(bold=True))
        _put(ws, res_r + 5, 2,
             "Same market effect as closing. Additionally frees the building slot and cannot be reopened; "
             "closing is reversible and sheds workers at 10%/month rather than instantly.",
             font=NOTE_FONT)

        for rr, vals in prefill["goods"].items():
            row = {"out": out_r, "in1": in_rows[0], "in2": in_rows[1], "in3": in_rows[2]}[rr]
            for col, v in vals.items():
                _put(ws, row, col, v, fill=INPUT_FILL, border=True,
                     fmt="0.00" if col >= 3 else "@")
        return res_r + 7

    end1 = block(5, "SCENARIO 1 - close the marginal level-1 guild? (prefilled: worked example A)", {
        "age": 2, "profit": 0.022,
        "goods": {
            "out": {2: "cloth", 3: 3.0, 4: 16.0, 5: 13.95, 6: 11.0, 7: 2.7, 8: 0.7, 9: 6.79},
            "in1": {2: "cotton", 3: 3.0, 4: 12.19, 5: 13.0, 6: 2.0, 7: 3.15, 8: 0.583, 9: 0.98},
        },
    })
    block(end1 + 2, "SCENARIO 2 - blank", {"age": 1, "profit": 0, "goods": {}})

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    for col in range(3, 17):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for col in range(chain_start, chain_start + len(CHAIN_BEFORE) + len(CHAIN_AFTER)):
        ws.column_dimensions[get_column_letter(col)].width = 10


def _example_goods_table(ws, r, goods, s, cap, outcomes):
    _header_row(ws, r, ["Good", "Base", "Eff S", "Eff D (pop part)", "dSupply", "dDemand",
                        "Price before", "Price after", "dPrice", "Exposure", "Gain /mo"])
    by_good = {o.good: o for o in outcomes}
    for g, state in goods.items():
        o = by_good[g]
        r += 1
        vals = [good_label(g), state.base, state.eff_supply,
                f"{state.eff_demand:.2f} ({state.pop_demand:.2f})",
                o.delta_supply, o.delta_demand, o.price_before, o.price_after,
                o.delta_price, o.exposure, o.gain]
        for i, v in enumerate(vals):
            fmt = "0.000" if isinstance(v, float) else "@"
            _put(ws, r, 1 + i, round(v, 4) if isinstance(v, float) else v,
                 fmt=fmt, border=True, align=CENTER)
    return r


def build_worked_examples(wb, goods_data):
    ws = wb.create_sheet("Worked Examples")
    _title(ws, "Worked Examples",
           "Computed with the same engine the calculator sheets mirror. Age 2 throughout: stability 10%, max price 3x, age efficiency 0.8.")
    s, cap, eff = 0.10, 3.0, 0.8

    cloth_base = goods_data.get("cloth", {}).get("base_price", 3.0)
    cotton_base = goods_data.get("cotton", {}).get("base_price", 1.5)
    glass_base = goods_data.get("glass", {}).get("base_price", 3.0)

    r = 4

    # --- Example A: the level-1 vs level-12 question ---
    _section(ws, r, "A. Close the marginal level-1 cloth guild? (the level-1 vs level-12 question)", 11)
    r += 1
    setup = [
        "You own the whole market: the level-12 cloth guild (market access 100%, +10% PE from cotton RGO in province, tax share 55%),",
        "a level-1 cloth guild in a worse spot (market access 70%, tax share 40%), and all the cotton RGOs. Recipe: 0.8333 cotton -> 1 cloth per level.",
        "Level-1 monthly flows: 0.70 cloth out, 0.58 cotton in.",
        f"Market (converged): cloth base {cloth_base:.2f}, S 16.0, D 13.95 (11.0 pop + 2.95 your buildings), price 2.70; "
        f"cotton base {cotton_base:.2f}, S 12.19, D 13.0 (2.0 pop + 11.0 your guilds), price 3.15.",
        "Exposure excluding the level-1: cloth 0.55 * (16.0 - 0.7 - 2.95) = +6.79 (pops and surplus buy the rest);",
        "cotton 0.55 * (12.19 - 10.42) = +0.98 (only your net cotton surplus counts - the internal 10.42 is a wash).",
    ]
    for line in setup:
        _put(ws, r, 1, line)
        r += 1

    goods_a = {
        "cloth": GoodState("cloth", cloth_base, 16.0, 13.95, 11.0, 2.70),
        "cotton": GoodState("cotton", cotton_base, 12.19, 13.0, 2.0, 3.15),
    }
    exposures_a = {"cloth": 6.79, "cotton": 0.98}
    st = goods_a["cloth"].equilibrium(s, cap)
    st_cot = goods_a["cotton"].equilibrium(s, cap)
    l1_gross = 0.70 * st.price - 0.583 * st_cot.price
    l1_taxed = 0.40 * l1_gross
    dec_a = evaluate_close(goods_a, s, cap, "cloth", 0.70, {"cotton": 0.583},
                           l1_taxed, exposures_a)
    r += 1
    r = _example_goods_table(ws, r, goods_a, s, cap, dec_a.outcomes)
    r += 2
    for text, val in [
        ("Level-1 building's taxed profit (what you give up)", l1_taxed),
        ("Price displacement gain on the level-12 (cloth up, cotton down)", dec_a.revaluation),
        ("DELTA INCOME from closing", dec_a.delta_income),
    ]:
        _put(ws, r, 1, text, font=Font(bold=True))
        _put(ws, r, 2, round(val, 4), fmt="+0.000;-0.000;0", align=CENTER)
        r += 1
    verdict_a = "CLOSE" if dec_a.delta_income > 0 else "KEEP OPEN"
    _put(ws, r, 1, f"VERDICT: {verdict_a}", font=Font(bold=True, size=12),
         fill=GOOD_FILL if dec_a.delta_income > 0 else BAD_FILL)
    r += 2
    _put(ws, r, 1, f"Why: the guild earns you {l1_taxed:.3f}/mo, but pulling its 0.70 cloth off the market reprices your entire "
                   "pop-facing cloth position. The cotton dip barely matters: you produce the cotton yourself, so its price "
                   "nets out across your book except for the small surplus.",
         font=NOTE_FONT)
    r += 1
    _put(ws, r, 1, "Quick rule check: the guild's margin is ~1% of its output value; the outward share of cloth is ~80% "
                   "-> margin far below 0.8 x share -> close.", font=NOTE_FONT)
    r += 3

    # --- Example B: dominant producer builds a level ---
    _section(ws, r, "B. Should the dominant producer build level 13? (profitable building, negative build)", 11)
    r += 1
    dec_b = evaluate_build(goods_a, s, cap, eff, "cloth", 1.0, {"cotton": 0.8333},
                           level=12, pe_other=0.10, access=1.0, employment=1.0,
                           tax_share=0.55, exposures={"cloth": 7.18, "cotton": 0.65},
                           construction_cost=300.0)
    setup_b = [
        "Same market. The level-12 guild adds level 13. New level output = 1 * (1 + 0.10) + 2*0.8%*12 = "
        f"{dec_b.detail['d_out_base']:.3f} + {dec_b.detail['d_out_pe_bump']:.3f} = {dec_b.detail['d_out']:.3f} cloth "
        "(the second term is the +1%/level PE bump the new level gives all 12 existing levels).",
        "Exposure includes your whole book: cloth 0.55 * (16.0 - 2.95) = +7.18; cotton surplus 0.55 * (12.19 - 11.0) = +0.65.",
    ]
    for line in setup_b:
        _put(ws, r, 1, line)
        r += 1
    r += 1
    r = _example_goods_table(ws, r, goods_a, s, cap, dec_b.outcomes)
    r += 2
    for text, val in [
        ("New level's taxed profit at post-build prices", dec_b.own_effect),
        ("Price displacement on your existing position", dec_b.revaluation),
        ("DELTA INCOME from building", dec_b.delta_income),
    ]:
        _put(ws, r, 1, text, font=Font(bold=True))
        _put(ws, r, 2, round(val, 4), fmt="+0.000;-0.000;0", align=CENTER)
        r += 1
    verdict_b = "BUILD" if dec_b.delta_income > 0 else "DO NOT BUILD"
    _put(ws, r, 1, f"VERDICT: {verdict_b}", font=Font(bold=True, size=12),
         fill=GOOD_FILL if dec_b.delta_income > 0 else BAD_FILL)
    r += 2
    goods_s = {
        "cloth": GoodState("cloth", cloth_base, 4.0, 8.0, 7.0),
        "cotton": GoodState("cotton", cotton_base, 4.17, 4.17, 0.0),
    }
    dec_b2 = evaluate_build(goods_s, s, cap, eff, "cloth", 1.0, {"cotton": 0.8333},
                            level=0, pe_other=0.0, access=1.0, employment=1.0,
                            tax_share=0.55, exposures={"cloth": 0.55 * 3.0, "cotton": 0.0},
                            construction_cost=300.0)
    _put(ws, r, 1, "Contrast: the same guild in a market where cloth is SCARCE (S 4.0 vs D 8.0, price ~1.6x base, cotton internal and balanced): "
                   f"new level taxed profit {dec_b2.own_effect:+.3f}/mo, displacement {dec_b2.revaluation:+.3f} on your existing position "
                   f"-> delta income {dec_b2.delta_income:+.3f}/mo = {'BUILD' if dec_b2.delta_income > 0 else 'DO NOT BUILD'}.",
         font=NOTE_FONT)
    r += 1
    _put(ws, r, 1, "Same brick, opposite verdicts, and the quick rule calls both: saturated cloth market = thin margin far below "
                   "0.8 x outward share (cut); scarce cloth = fat margin above it (build until the price falls enough that it is not).",
         font=NOTE_FONT)
    r += 3

    # --- Example C: capped good ---
    _section(ws, r, "C. A good pinned at the price cap (Glass in the London samples)", 11)
    r += 1
    glass = GoodState("glass", glass_base, 0.0, 3.35, 3.0, 6.71)
    eq_now = glass.equilibrium(s, cap)
    d_exo_g, pop_base_g = glass.decompose(s, cap)
    setup_c = [
        f"Tooltip state: base {glass_base:.2f}, supply 0, displayed demand 3.35 (3.00 from pops), current price 6.71, tooltip target 9.00 (the 3x cap).",
        f"Pops at 6.71 have already cut demand to {elasticity_factor(6.71, glass_base) * 100:.1f}% of base; at-base pop demand is {pop_base_g:.2f}.",
        f"As the price keeps climbing toward the cap, pop demand keeps shrinking - the LONG-RUN price settles at {eq_now.price:.2f}, "
        f"below the 9.00 the tooltip shows today (branch: {eq_now.branch}).",
    ]
    for line in setup_c:
        _put(ws, r, 1, line)
        r += 1
    r += 1
    glass_recipe_note = "one glassworks level (1.0 output, inputs ignored here - they sit in healthy markets)"
    eq_after = glass.equilibrium(s, cap, d_supply=1.0)
    _put(ws, r, 1, f"Build {glass_recipe_note}: supply 0 -> 1 moves the long-run price {eq_now.price:.2f} -> {eq_after.price:.2f}. "
                   f"Gross revenue of that single level: {eq_after.price:.2f}/mo - the first supplier of a starved good earns multiples of base price.",
         font=Font(bold=True))
    r += 2
    for line in [
        "The 'dead zone' intuition (capped price = adding supply is free) is real but NARROW here: with pop demand already collapsing,",
        f"the cap stops binding almost immediately (headroom {max(0.0, unpin_supply(eq_now.demand_eff, s, cap)):.2f} supply). On goods with large "
        "price-insensitive demand (building inputs), the cap dead zone is wide and early supply truly does not move the price.",
        "Rule: check the Price Calculator's 'Supply headroom while capped' column before assuming a capped price will hold.",
    ]:
        _put(ws, r, 1, line, font=NOTE_FONT)
        r += 1

    ws.column_dimensions["A"].width = 60
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 13


def build_goods_reference(wb, goods_data, pop_demands):
    ws = wb.create_sheet("Goods Reference")
    _title(ws, "Goods Reference", "Every good: base price, category, and per-pop-type base demand (before price elasticity and age modifiers).")
    pop_types = list(pop_demands["pop_types"].keys())
    demand_by_good = {g["name"]: g["demands"] for g in pop_demands["goods"]}

    _header_row(ws, 4, ["Good", "Id", "Category", "Method", "Base price", "Food value",
                        "Transport cost"] + [pretty(p) for p in pop_types])
    r = 4
    for gid in sorted(goods_data, key=lambda g: (goods_data[g]["category"], g)):
        d = goods_data[gid]
        r += 1
        _put(ws, r, 1, good_label(gid), border=True)
        _put(ws, r, 2, gid, border=True)
        _put(ws, r, 3, d["category"], border=True, align=CENTER)
        _put(ws, r, 4, d["method"], border=True, align=CENTER)
        _put(ws, r, 5, d["base_price"], fmt="0.00", border=True, align=CENTER)
        _put(ws, r, 6, d["food"] if d["food"] else None, fmt="0.0", border=True, align=CENTER)
        _put(ws, r, 7, d["transport_cost"], fmt="0.0", border=True, align=CENTER)
        dem = demand_by_good.get(gid, {})
        for i, p in enumerate(pop_types):
            v = dem.get(p, 0)
            _put(ws, r, 8 + i, v if v else None, fmt="0.0000", border=True, align=CENTER)
    ws.freeze_panes = "A5"
    _autosize(ws)


def build_recipes_reference(wb, recipes, goods_data, age_table_ref):
    ws = wb.create_sheet("Recipes Reference")
    _title(ws, "Recipes Reference",
           "Every production method: per-level quantities and margins at base prices. Margins ignore PE and market access (both scale them up/down).")
    _put(ws, 3, 1, "Age for the stress columns:", font=Font(bold=True))
    age_cell = "$B$3"
    _put(ws, 3, 2, 2, fill=INPUT_FILL, fmt="0", border=True, align=CENTER)
    s_ref = f"VLOOKUP({age_cell},{age_table_ref},2,FALSE)"
    k_ref = f"VLOOKUP({age_cell},{age_table_ref},3,FALSE)"
    _put(ws, 3, 3, "Floor margin = output at its price floor (s*base), inputs at base. Worst case = output at floor, inputs at their cap (K*base).",
         font=NOTE_FONT)

    _header_row(ws, 5, ["Output good", "Building", "Method", "Tier", "Output /level",
                        "Output base price", "Input 1", "Qty", "Base", "Input 2", "Qty", "Base",
                        "Input 3", "Qty", "Base", "Margin at base prices",
                        "Margin, output at floor", "Worst-case margin"])
    r = 5
    for produced in sorted(recipes):
        for recipe in recipes[produced]:
            r += 1
            out_base = goods_data.get(produced, {}).get("base_price", 0)
            _put(ws, r, 1, good_label(produced), border=True)
            _put(ws, r, 2, recipe["building"], border=True)
            _put(ws, r, 3, recipe["method"], border=True)
            _put(ws, r, 4, recipe["tier"], border=True, align=CENTER)
            _put(ws, r, 5, recipe["output"], fmt="0.000", border=True, align=CENTER)
            _put(ws, r, 6, out_base, fmt="0.00", border=True, align=CENTER)
            ins = sorted(recipe["inputs"].items())[:3]
            for i in range(3):
                col = 7 + i * 3
                if i < len(ins):
                    gname, qty = ins[i]
                    _put(ws, r, col, good_label(gname), border=True)
                    _put(ws, r, col + 1, qty, fmt="0.000", border=True, align=CENTER)
                    _put(ws, r, col + 2, goods_data.get(gname, {}).get("base_price", 0),
                         fmt="0.00", border=True, align=CENTER)
                else:
                    for cc in (col, col + 1, col + 2):
                        _put(ws, r, cc, None, border=True)
            terms = []
            for i in range(3):
                qcol, bcol = get_column_letter(8 + i * 3), get_column_letter(9 + i * 3)
                terms.append(f'IF(${qcol}{r}="",0,${qcol}{r}*${bcol}{r})')
            in_cost = "+".join(terms)
            _put(ws, r, 16, f'=$E{r}*$F{r}-({in_cost})', fmt="+0.000;-0.000;0", border=True, align=CENTER)
            _put(ws, r, 17, f'=$E{r}*$F{r}*{s_ref}-({in_cost})', fmt="+0.000;-0.000;0", border=True, align=CENTER)
            _put(ws, r, 18, f'=$E{r}*$F{r}*{s_ref}-({in_cost})*{k_ref}', fmt="+0.000;-0.000;0", border=True, align=CENTER)
    ws.freeze_panes = "A6"
    _autosize(ws)


URBAN_RIGHTS_FIRST_ROW = 7
# Urban Rights Detail columns the sheet's own formulas and the ranking read back.
COL_BONUS, COL_EFFICIENCY, COL_NET_BONUS = 5, 6, 7
COL_OUTPUT, COL_PRICE, COL_BASE_VALUE, COL_VALUE_ADDED = 11, 12, 13, 14


def urban_rights_ranges(rows: list) -> dict:
    """{(right_id, era): (first_row, last_row)} for the Urban Rights Detail sheet."""
    ranges = {}
    for i, row in enumerate(rows):
        r = URBAN_RIGHTS_FIRST_ROW + i
        ranges.setdefault((row["right_id"], row["era"]), [r, r])[1] = r
    return {key: tuple(span) for key, span in ranges.items()}


def _other_modifiers_text(right: dict) -> str:
    """The right's modifiers this workbook does not value, script values resolved."""
    parts = []
    for key, value in sorted(right["other"].items()):
        if key == "local_production_efficiency":
            continue
        value = _resolve_modifier(value)
        if isinstance(value, float):
            parts.append(f"{key} {value:+.0%}")
        else:
            parts.append(f"{key} {value:+g}" if isinstance(value, int) else f"{key} = {value}")
    return ", ".join(parts)


def build_urban_rights_detail(wb, rows):
    """Per (right, era, good) working: the building held, its output, and what
    the net bonus adds. Row order must match urban_rights_ranges.
    """
    ws = wb.create_sheet("Urban Rights Detail")
    _title(ws, "Urban Rights Detail",
           "Every boosted good, the building a location holds in each era, and the value the right adds to one level of it. "
           "Output bonuses and production efficiency are the same bucket and both lift output only, so value added = output x price x net bonus and input costs never enter it.")
    _put(ws, 3, 1,
         "Worked example: Tooling Rights on a Tools Guild running Blacksmiths. +30% tools output less the right's own -5% production efficiency is a net +25%. "
         "Output 1 tools at base price 3 is a base value of 3.00, so the right adds 0.75 per level. "
         "Counting inputs gives the same answer: (1 x 3 x 1.25) - (0.8333 x 3) = 1.25 against (1 x 3) - (0.8333 x 3) = 0.50.",
         font=NOTE_FONT)
    _put(ws, 4, 1,
         "Price is editable (yellow) and defaults to the good's base price. Every total on the Urban Rights Ranking sheet follows it.",
         font=NOTE_FONT)

    _header_row(ws, 6, ["Right", "Right id", "Era", "Good", "Output bonus", "Prod efficiency",
                        "Net bonus", "Building", "Building class", "Production method(s)",
                        "Output /level", "Price", "Base value /level", "Value added /level",
                        "Margin at base prices", "Method unlocked by"])
    net_col, out_col, price_col, base_col = (
        get_column_letter(c) for c in (COL_NET_BONUS, COL_OUTPUT, COL_PRICE, COL_BASE_VALUE))
    bonus_col, eff_col = get_column_letter(COL_BONUS), get_column_letter(COL_EFFICIENCY)
    r = URBAN_RIGHTS_FIRST_ROW - 1
    for row in rows:
        r += 1
        _put(ws, r, 1, row["right"], border=True)
        _put(ws, r, 2, row["right_id"], border=True)
        _put(ws, r, 3, row["era"], border=True, align=CENTER)
        _put(ws, r, 4, good_label(row["good"]), border=True)
        _put(ws, r, COL_BONUS, row["bonus"], fmt="0%", border=True, align=CENTER)
        _put(ws, r, COL_EFFICIENCY, row["efficiency"], fmt="+0%;-0%;0", border=True, align=CENTER)
        _put(ws, r, COL_NET_BONUS, f"=${bonus_col}{r}+${eff_col}{r}", fmt="0%",
             border=True, align=CENTER)
        _put(ws, r, 8, row["building"], border=True)
        _put(ws, r, 9, row["klass"], border=True, align=CENTER,
             fill=CALC_FILL if row["klass"] != row["era"] else None)
        _put(ws, r, 10, row["methods"], border=True)
        _put(ws, r, COL_OUTPUT, row["output"], fmt="0.000", border=True, align=CENTER)
        _put(ws, r, COL_PRICE, row["base_price"], fmt="0.00", border=True, align=CENTER,
             fill=INPUT_FILL)
        _put(ws, r, COL_BASE_VALUE, f"=${out_col}{r}*${price_col}{r}", fmt="0.000",
             border=True, align=CENTER)
        _put(ws, r, COL_VALUE_ADDED, f"=${base_col}{r}*${net_col}{r}", fmt="0.000",
             border=True, align=CENTER, fill=GOOD_FILL)
        _put(ws, r, 15, row["margin"], fmt="+0.000;-0.000;0", border=True, align=CENTER)
        _put(ws, r, 16, row["unlock"] or None, border=True)
    _put(ws, r + 2, 1,
         "A gray building class means that era has no building of its own for the good, so the location keeps the newest one it has.",
         font=NOTE_FONT)
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:P{r}"
    _autosize(ws)


def build_urban_rights_ranking(wb, rights, rows, ranges):
    """Rights ranked by the value they add per building level, one block per
    production era.
    """
    ws = wb.create_sheet("Urban Rights Ranking")
    _title(ws, "Urban Rights Ranking",
           "Urban Rights granting a goods output bonus, ranked by the gold per month they add per building level.")
    lines = [
        "The production specialization set off the Borough Privileges advance; culture and country restricted rights are left out, as are production methods behind a restricted advance.",
        "Each right lifts its own goods and drops everything else in the location by 5%. Only the boosted buildings are assumed built, so the two net into one effective bonus: a 20% output bonus with -5% production efficiency is a 15% bonus.",
        "Total counts one level of every building the right boosts, so a three-good right needs one of each to collect it. Best single is the strongest of those buildings on its own.",
        "Buildings run their highest-output production method, plus the enhancement slot where they have one.",
        "Value added = output x price x net bonus. Ranked at base prices; edit a price on the Urban Rights Detail sheet and the totals follow, though the row order stays as built.",
        "Gross value at the location: your own take is the tax share of it and market access scales output, and both scale the whole column evenly.",
    ]
    for i, line in enumerate(lines):
        _put(ws, 3 + i, 1, line, font=NOTE_FONT)

    detail = "'Urban Rights Detail'"
    value_col = get_column_letter(COL_VALUE_ADDED)
    by_right = {right["id"]: right for right in rights}
    best_good = {}
    for row in rows:
        key = (row["right_id"], row["era"])
        if row["value_added"] > best_good.get(key, (0, ""))[0]:
            best_good[key] = (row["value_added"], good_label(row["good"]))

    r = 3 + len(lines)
    for era in BUILDING_ERAS:
        totals = {}
        for row in rows:
            if row["era"] == era:
                totals[row["right_id"]] = totals.get(row["right_id"], 0) + row["value_added"]
        if not totals:
            continue
        r += 2
        _section(ws, r, f"{era} era", width=10)
        r += 1
        _header_row(ws, r, ["Rank", "Right", "Goods boosted", "Output bonus", "Prod efficiency",
                            "Net bonus", "Total value added /level", "Best single /level",
                            "Best good", "Buildings counted", "Also grants (not valued here)"])
        for rank, (right_id, _total) in enumerate(
                sorted(totals.items(), key=lambda kv: -kv[1]), start=1):
            right = by_right[right_id]
            first, last = ranges[(right_id, era)]
            bonuses = sorted(right["bonuses"].items(), key=lambda kv: good_label(kv[0]))
            r += 1
            _put(ws, r, 1, rank, border=True, align=CENTER)
            _put(ws, r, 2, right["name"], border=True)
            _put(ws, r, 3, ", ".join(good_label(g) for g, _ in bonuses), border=True)
            _put(ws, r, 4, ", ".join(f"{pct:.0%}" for _, pct in bonuses), border=True, align=CENTER)
            _put(ws, r, 5, right["efficiency"], fmt="+0%;-0%;0", border=True, align=CENTER)
            _put(ws, r, 6, ", ".join(f"{pct + right['efficiency']:.0%}" for _, pct in bonuses),
                 border=True, align=CENTER)
            _put(ws, r, 7, f"=SUM({detail}!${value_col}${first}:${value_col}${last})",
                 fmt="0.000", border=True, align=CENTER, fill=GOOD_FILL)
            _put(ws, r, 8, f"=MAX({detail}!${value_col}${first}:${value_col}${last})",
                 fmt="0.000", border=True, align=CENTER)
            _put(ws, r, 9, best_good[(right_id, era)][1], border=True)
            _put(ws, r, 10, last - first + 1, border=True, align=CENTER)
            _put(ws, r, 11, _other_modifiers_text(right) or None, border=True)
    _autosize(ws)


def build_price_curves(wb, ages):
    ws = wb.create_sheet("Price Curves")
    _title(ws, "Price Curves", "Target price multiplier by R = (1+S)/(1+D), per age. Gray = pinned at the age max price.")
    _header_row(ws, 4, ["R"] + [f"Age {a['num']} ({a['name']})  s={a['stability']:.2f} K={a['max_price']:.0f}" for a in ages])
    r = 4
    steps = [round(0.05 * i, 2) for i in range(1, 101)]
    for R in steps:
        r += 1
        _put(ws, r, 1, R, fmt="0.00", border=True, align=CENTER)
        for i, a in enumerate(ages):
            m = target_multiplier(R, a["stability"], a["max_price"])
            capped = (1 + a["stability"] * R) / (R + a["stability"]) >= a["max_price"]
            _put(ws, r, 2 + i, round(m, 4), fmt="0.000", border=True, align=CENTER,
                 fill=CALC_FILL if capped else None)
    ws.freeze_panes = "B5"
    _autosize(ws)


# =====================================================================
# Main
# =====================================================================

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Verifying price formula against the 9 in-game tooltip samples...")
    rows = verify_tooltip_samples()
    worst = max(r[6] for r in rows)
    print(f"  9/9 reproduce (worst diff {worst:.4f}, tolerance {SAMPLE_TOLERANCE})")

    print("Verifying closed-form equilibrium against fixed-point iteration...")
    n = verify_closed_form()
    print(f"  {n} parameter combinations match to 1e-6")

    print("Scraping game data...")
    ages = scrape_ages()
    goods_data = scrape_goods()
    recipes = scrape_production_recipes()
    for produced, lst in scrape_shared_methods().items():
        recipes.setdefault(produced, []).extend(lst)
    pop_demands = scrape_pop_demands()
    town_rights = scrape_town_right_output_bonuses(
        goods_data, scrape_advance_unlocks("unlock_town_rights"))
    producers = scrape_producing_buildings()
    rights_rows = urban_right_rows(town_rights, producers, goods_data,
                                   scrape_advance_unlocks("unlock_production_method"))
    print(f"  {len(ages)} ages, {len(goods_data)} goods, "
          f"{sum(len(v) for v in recipes.values())} recipes, {len(pop_demands['goods'])} pop-demanded goods")
    print(f"  {len(town_rights)} unrestricted urban rights with a goods output bonus, "
          f"{len(producers)} producing buildings, {len(rights_rows)} right/era/good rows")

    wb = Workbook()
    build_readme(wb)
    age_table_ref = build_price_model(wb, ages)
    build_price_calculator(wb, age_table_ref)
    build_build_calculator(wb, age_table_ref)
    build_turnoff_calculator(wb, age_table_ref)
    build_worked_examples(wb, goods_data)
    build_urban_rights_ranking(wb, town_rights, rights_rows, urban_rights_ranges(rights_rows))
    build_urban_rights_detail(wb, rights_rows)
    build_goods_reference(wb, goods_data, pop_demands)
    build_recipes_reference(wb, recipes, goods_data, age_table_ref)
    build_price_curves(wb, ages)

    out = OUTPUT_DIR / "eu5_market_analysis.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        if sys.platform == "win32":
            subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, timeout=5)
            import time
            time.sleep(1)
        wb.save(out)
    print(f"\nSaved to: {out}")
    if sys.platform == "win32" and os.environ.get("MARKET_ANALYSIS_NO_OPEN") != "1":
        os.startfile(out)


if __name__ == "__main__":
    main()
