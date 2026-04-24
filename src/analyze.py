"""Analyze EU5 army unit data and export to Excel spreadsheet."""

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
OUTPUT_DIR = DATA_DIR

AGE_ORDER = [
    "age_1_traditions",
    "age_2_renaissance",
    "age_3_discovery",
    "age_4_reformation",
    "age_5_absolutism",
    "age_6_revolutions",
]
AGE_LABELS = {
    "age_1_traditions": "1 - Traditions",
    "age_2_renaissance": "2 - Renaissance",
    "age_3_discovery": "3 - Discovery",
    "age_4_reformation": "4 - Reformation",
    "age_5_absolutism": "5 - Absolutism",
    "age_6_revolutions": "6 - Revolutions",
}
CAT_LABELS = {
    "army_infantry": "Infantry",
    "army_cavalry": "Cavalry",
    "army_artillery": "Artillery",
    "army_auxiliary": "Auxiliary",
    "navy_galley": "Galley",
    "navy_light_ship": "Light Ship",
    "navy_heavy_ship": "Heavy Ship",
    "navy_transport": "Transport",
}
LAND_CATS = ["army_infantry", "army_cavalry", "army_artillery", "army_auxiliary"]

# Styling
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, italic=True)
NUM_FMT_2 = "0.00"
NUM_FMT_3 = "0.000"
NUM_FMT_PCT = "0%"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Category highlight colors
CAT_FILLS = {
    "Infantry": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "Cavalry": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Artillery": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Auxiliary": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    "Galley": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "Light Ship": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Heavy Ship": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Transport": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
}


def load_data():
    with open(DATA_DIR / "land_units.json") as f:
        land_units = json.load(f)
    with open(DATA_DIR / "unit_categories.json") as f:
        categories = json.load(f)
    with open(DATA_DIR / "age_progression.json") as f:
        age_progression = json.load(f)
    with open(DATA_DIR / "unit_prices.json") as f:
        prices = json.load(f)
    with open(DATA_DIR / "combined_arms.json") as f:
        combined_arms = json.load(f)
    with open(DATA_DIR / "goods_demands.json") as f:
        goods_demands = json.load(f)
    with open(DATA_DIR / "production_recipes.json") as f:
        production_recipes = json.load(f)
    with open(DATA_DIR / "localizations.json") as f:
        localizations = json.load(f)
    with open(DATA_DIR / "naval_units.json") as f:
        naval_units = json.load(f)
    with open(DATA_DIR / "food_goods.json") as f:
        food_goods = json.load(f)
    with open(DATA_DIR / "food_buildings.json") as f:
        food_buildings = json.load(f)
    with open(DATA_DIR / "building_caps.json") as f:
        building_caps = json.load(f)
    with open(DATA_DIR / "terrain_food_modifiers.json") as f:
        terrain_food_modifiers = json.load(f)
    with open(DATA_DIR / "forts.json") as f:
        forts = json.load(f)
    with open(DATA_DIR / "pop_demands.json") as f:
        pop_demands = json.load(f)
    return (land_units, categories, age_progression, prices, combined_arms,
            goods_demands, production_recipes, localizations, naval_units,
            food_goods, food_buildings, building_caps, terrain_food_modifiers,
            forts, pop_demands)


LOC = {}  # populated in main(), used by loc() helper


def loc(internal_name: str) -> str:
    """Get display name for a unit, falling back to internal name."""
    return LOC.get(internal_name, internal_name)


def style_header_row(ws, row, num_cols):
    """Style a header row with blue background and white bold text."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def safe_num(v, default=0):
    """Safely extract a numeric value, handling lists from duplicate keys."""
    if isinstance(v, list):
        # Sum duplicate modifier values (e.g., strength_damage_taken appearing twice)
        return sum(x for x in v if isinstance(x, (int, float)))
    if isinstance(v, (int, float)):
        return v
    return default


def auto_width(ws, min_width=8, max_width=30):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def calc_flank_power(damage, sfd, dmg_taken, str_dmg_done=0, str_dmg_taken=0):
    """FlankPower = EffDamage * 10 / ((1 - SFD) * DamageTaken * (1 + str_dmg_taken))

    str_dmg_done/taken: unit-specific strength damage modifiers (negative = better).
    """
    eff_damage = damage * (1 + str_dmg_done)
    denom = (1 - sfd) * dmg_taken * (1 + str_dmg_taken)
    if denom == 0:
        return 0
    return eff_damage * 10 / denom


def calc_center_power(flank_power, sfd):
    """CenterPower = FlankPower * (1 + 2 * SFD)"""
    return flank_power * (1 + 2 * sfd)


def calc_cost(strength, category, prices):
    """BuildCost = max_strength * base_build_gold * 10 (costs scale by max_strength)."""
    base = prices.get(category, {}).get("build_gold", 100)
    return strength * base * 10


def calc_maintenance(strength, category, prices):
    """Maintenance = max_strength * base_maintenance_gold * 10."""
    base = prices.get(category, {}).get("maintenance_gold", 1)
    return strength * base * 10


BEST_FONT = Font(bold=True)
BEST_OVERALL_FONT = Font(bold=True, color="FF0000")
BEST_CAT_FONT = Font(color="0070C0")  # blue


def highlight_best_in_age(ws, data_rows, header_row, col_indices):
    """Bold the best value per age for specified columns.

    data_rows: list of (excel_row_number, age_key, {col_index: value}) dicts
    col_indices: list of column indices (1-based) to check for best values
    """
    by_age = {}
    for row_num, age, vals in data_rows:
        by_age.setdefault(age, []).append((row_num, vals))

    for age, rows in by_age.items():
        for col in col_indices:
            best_val = None
            best_row = None
            for row_num, vals in rows:
                v = vals.get(col, 0)
                if v is not None and (best_val is None or v > best_val):
                    best_val = v
                    best_row = row_num
            if best_row is not None and best_val is not None and best_val > 0:
                ws.cell(row=best_row, column=col).font = BEST_FONT


def highlight_best_in_age_by_cat(ws, data_rows, col_indices):
    """Highlight best overall (green) and best per category (blue) per age.

    data_rows: list of (excel_row_number, age_key, category, {col_index: value})
    """
    by_age = {}
    for row_num, age, cat, vals in data_rows:
        by_age.setdefault(age, []).append((row_num, cat, vals))

    for age, rows in by_age.items():
        for col in col_indices:
            # Find overall best
            best_val = None
            best_row = None
            # Find best per category
            cat_best = {}  # cat -> (row, val)
            for row_num, cat, vals in rows:
                v = vals.get(col, 0)
                if v is not None and v > 0:
                    if best_val is None or v > best_val:
                        best_val = v
                        best_row = row_num
                    cb = cat_best.get(cat)
                    if cb is None or v > cb[1]:
                        cat_best[cat] = (row_num, v)

            # Apply overall best = green
            if best_row is not None:
                ws.cell(row=best_row, column=col).font = BEST_OVERALL_FONT

            # Apply best per category = blue (skip the overall winner's category
            # and categories with only one unit)
            if best_row is not None:
                winner_cat = None
                for row_num, cat, vals in rows:
                    if row_num == best_row:
                        winner_cat = cat
                        break
                # Count units per category
                cat_counts = {}
                for row_num, cat, vals in rows:
                    cat_counts[cat] = cat_counts.get(cat, 0) + 1
                for cat, (cr, cv) in cat_best.items():
                    if cat != winner_cat and cv > 0 and cat_counts.get(cat, 0) > 1:
                        ws.cell(row=cr, column=col).font = BEST_CAT_FONT


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_army_meta(wb, age_data, categories, prices):
    """Sheet 1: Army Composition Meta - per-age flank/center analysis."""
    ws = wb.active
    ws.title = "Unit Power"

    # Title
    ws.cell(row=1, column=1, value="EU5 Army Composition Meta").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Per-age template stats with flank/center power analysis").font = SUBTITLE_FONT

    # Formula explanation
    ws.cell(row=3, column=1, value="Formulas:").font = Font(bold=True)
    ws.cell(row=4, column=1, value="  Damage = Strength x CombatPower")
    ws.cell(row=5, column=1, value="  Flank Power = Damage x 10 / ((1 - SecureFlanks) x DamageTaken)")
    ws.cell(row=6, column=1, value="  Center Power = Flank Power x (1 + 2 x SecureFlanks)")
    ws.cell(row=7, column=1, value="  Power Per Gold = Power / BuildCost x 100")
    ws.cell(row=8, column=1, value="  Gold cells = best in age for that column").font = Font(italic=True)

    # Headers
    headers = [
        "Age", "Category", "Strength", "Combat Power", "Combat Speed",
        "Flanking Ability", "Secure Flanks", "Damage Taken",
        "Build Cost", "Maintenance",
        "Damage",
        "Flank Power", "Center Power",
        "Flank Power/Gold", "Center Power/Gold",
    ]
    header_row = 10
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    # Column indices for highlighting (1-based): FP=12, CP=13, FP/G=14, CP/G=15
    HIGHLIGHT_COLS = [12, 13, 14, 15]

    tracked_rows = []  # (row_num, age_key, {col: value})
    row = header_row + 1
    for age in AGE_ORDER:
        age_rows = [r for r in age_data if r["age"] == age]
        for ar in age_rows:
            cat = ar["category"]
            if cat not in LAND_CATS:
                continue

            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            strength = ar["max_strength"]
            cp = ar["combat_power"]
            cs = cat_data.get("combat_speed", 1)
            fa = cat_data.get("flanking_ability", 1.0)
            sfd = cat_data.get("secure_flanks_defense", 0)
            dmg_taken = cat_data.get("damage_taken", 1.0)

            damage = strength * cp
            build_cost = calc_cost(strength, cat, prices)
            maint = calc_maintenance(strength, cat, prices)

            fp = calc_flank_power(damage, sfd, dmg_taken)
            cp_val = calc_center_power(fp, sfd)
            fp_gold = fp / build_cost * 100 if build_cost > 0 else 0
            cp_gold = cp_val / build_cost * 100 if build_cost > 0 else 0

            values = [
                AGE_LABELS[age], cat_label, strength, cp, cs,
                fa, sfd, dmg_taken,
                build_cost, maint,
                damage,
                fp, cp_val,
                fp_gold, cp_gold,
            ]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1  # blank row between ages

    highlight_best_in_age(ws, tracked_rows, header_row, HIGHLIGHT_COLS)
    auto_width(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def build_buildable_units(wb, land_units, categories, prices):
    """Sheet 2: All buildable units with detailed stats and power calculations."""
    ws = wb.create_sheet("Unique Units")

    ws.cell(row=1, column=1, value="Unique Units by Age").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Green = best overall in age, Blue = best in category").font = Font(italic=True)

    headers = [
        "Age", "Category", "Unit",
        "Flank Power", "Center Power", "Flank P/Gold", "Center P/Gold",
        "Light", "Special",
        "Strength", "Combat Power", "Combat Speed", "Initiative",
        "Flanking", "Secure Flanks", "Damage Taken",
        "Str Dmg Taken", "Morale Dmg Taken",
        "Str Dmg Done", "Morale Dmg Done",
        "Damage", "Build Cost",
        "Upgrades To",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    # Highlight columns: FP=4, CP=5, FP/G=6, CP/G=7
    HIGHLIGHT_COLS = [4, 5, 6, 7]

    units = [
        u for u in land_units
        if u.get("buildable", True)
        and not u.get("levy", False)
        and not u["name"].startswith("a_age_")
    ]

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            strength = u.get("max_strength", 0)
            cp = u.get("combat_power", 0)
            cs = u.get("combat_speed", cat_data.get("combat_speed", 1))
            fa = u.get("flanking_ability", cat_data.get("flanking_ability", 1.0))
            sfd = u.get("secure_flanks_defense", cat_data.get("secure_flanks_defense", 0))
            dmg_taken = u.get("damage_taken", cat_data.get("damage_taken", 1.0))

            sdd = safe_num(u.get("strength_damage_done", 0))
            sdt = safe_num(u.get("strength_damage_taken", 0))

            damage = strength * cp
            build_cost = calc_cost(strength, cat, prices)

            fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)
            center = calc_center_power(fp, sfd)
            fp_gold = fp / build_cost * 100 if build_cost > 0 else 0
            cp_gold = center / build_cost * 100 if build_cost > 0 else 0

            values = [
                AGE_LABELS.get(u.get("age", ""), "?"),
                cat_label,
                loc(u["name"]),
                fp, center, fp_gold, cp_gold,
                "Yes" if u.get("light") else "",
                "Yes" if u.get("is_special") else "",
                strength, cp, cs,
                safe_num(u.get("initiative", cat_data.get("initiative", 1))),
                fa, sfd, dmg_taken,
                safe_num(u.get("strength_damage_taken", 0)),
                safe_num(u.get("morale_damage_taken", 0)),
                safe_num(u.get("strength_damage_done", 0)),
                safe_num(u.get("morale_damage_done", 0)),
                damage, build_cost,
                loc(u.get("upgrades_to", "")),
            ]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1  # gap between ages

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def build_levy_units(wb, land_units, categories, prices):
    """Levy Units sheet: all levy units per age with power analysis."""
    ws = wb.create_sheet("Levy Units")

    ws.cell(row=1, column=1, value="Levy Units by Age").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Red = best overall in age, Blue = best in category").font = Font(italic=True)

    headers = [
        "Age", "Category", "Unit",
        "Flank Power", "Center Power", "Flank P/Gold", "Center P/Gold",
        "Light", "Special",
        "Strength", "Combat Power", "Combat Speed", "Initiative",
        "Flanking", "Secure Flanks", "Damage Taken",
        "Str Dmg Taken", "Morale Dmg Taken",
        "Str Dmg Done", "Morale Dmg Done",
        "Damage", "Build Cost",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [4, 5, 6, 7]

    levies = [
        u for u in land_units
        if u.get("levy", False)
    ]

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in levies if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            strength = u.get("max_strength", 0)
            cp = u.get("combat_power", 0)
            cs = u.get("combat_speed", cat_data.get("combat_speed", 1))
            fa = u.get("flanking_ability", cat_data.get("flanking_ability", 1.0))
            sfd = u.get("secure_flanks_defense", cat_data.get("secure_flanks_defense", 0))
            dmg_taken = u.get("damage_taken", cat_data.get("damage_taken", 1.0))

            sdd = safe_num(u.get("strength_damage_done", 0))
            sdt = safe_num(u.get("strength_damage_taken", 0))

            damage = strength * cp
            build_cost = calc_cost(strength, cat, prices)

            fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)
            center = calc_center_power(fp, sfd)
            fp_gold = fp / build_cost * 100 if build_cost > 0 else 0
            cp_gold = center / build_cost * 100 if build_cost > 0 else 0

            values = [
                AGE_LABELS.get(u.get("age", ""), "?"),
                cat_label,
                loc(u["name"]),
                fp, center, fp_gold, cp_gold,
                "Yes" if u.get("light") else "",
                "Yes" if u.get("is_special") else "",
                strength, cp, cs,
                safe_num(u.get("initiative", cat_data.get("initiative", 1))),
                fa, sfd, dmg_taken,
                safe_num(u.get("strength_damage_taken", 0)),
                safe_num(u.get("morale_damage_taken", 0)),
                safe_num(u.get("strength_damage_done", 0)),
                safe_num(u.get("morale_damage_done", 0)),
                damage, build_cost,
            ]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1  # gap between ages

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def build_upgrade_chains(wb, land_units, categories, prices):
    """Sheet 3: Upgrade chains laid out horizontally."""
    ws = wb.create_sheet("Upgrade Chains")

    ws.cell(row=1, column=1, value="Unit Upgrade Chains").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Each row shows a full upgrade path from earliest to latest age").font = SUBTITLE_FONT

    headers = ["Category", "Type"] + [AGE_LABELS[a] for a in AGE_ORDER]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    units_by_name = {u["name"]: u for u in land_units}
    buildable = [
        u for u in land_units
        if u.get("buildable", True)
        and not u.get("levy", False)
        and not u["name"].startswith("a_age_")
    ]

    upgraded_to = {u.get("upgrades_to") for u in buildable if u.get("upgrades_to")}
    starters = [u for u in buildable if u["name"] not in upgraded_to]

    row = header_row + 1
    for start in sorted(starters, key=lambda x: (x["category"], x.get("is_special", False), x["name"])):
        cat_label = CAT_LABELS.get(start["category"], start["category"])
        is_light = start.get("light", False)
        is_special = start.get("is_special", False)
        type_label = ("Special" if is_special else "Light" if is_light else "Heavy")

        ws.cell(row=row, column=1, value=cat_label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=type_label).border = THIN_BORDER

        # Walk the chain and place units in their age column
        current = start
        visited = set()
        while current and current["name"] not in visited:
            visited.add(current["name"])
            age = current.get("age", "")
            if age in AGE_ORDER:
                col = AGE_ORDER.index(age) + 3  # offset for Category, Type columns
                cat = current["category"]
                cat_data = categories.get(cat, {})
                sfd = cat_data.get("secure_flanks_defense", 0)
                dmg_taken = cat_data.get("damage_taken", 1.0)
                sdd = safe_num(current.get("strength_damage_done", 0))
                sdt = safe_num(current.get("strength_damage_taken", 0))
                strength = current.get("max_strength", 0)
                damage = strength * current.get("combat_power", 0)
                fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)
                cp = round(calc_center_power(fp, sfd), 1)
                maint = round(calc_maintenance(strength, cat, prices), 1)
                cell = ws.cell(row=row, column=col, value=f"{loc(current['name'])} ({cp}p, {maint}g)")
                cell.border = THIN_BORDER
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            next_name = current.get("upgrades_to", "")
            current = units_by_name.get(next_name) if next_name else None

        row += 1

    auto_width(ws, min_width=10, max_width=45)
    ws.freeze_panes = f"A{header_row + 1}"


def build_levy_upgrade_chains(wb, land_units, categories, prices):
    """Levy upgrade chains laid out horizontally, same format as professional chains."""
    ws = wb.create_sheet("Levy Upgrade Chains")

    ws.cell(row=1, column=1, value="Levy Unit Upgrade Chains").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Each row shows a levy unit or upgrade path from earliest to latest age").font = SUBTITLE_FONT

    headers = ["Category", "Type"] + [AGE_LABELS[a] for a in AGE_ORDER]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    units_by_name = {u["name"]: u for u in land_units}
    levies = [
        u for u in land_units
        if u.get("levy", False)
    ]

    upgraded_to = {u.get("upgrades_to") for u in levies if u.get("upgrades_to")}
    starters = [u for u in levies if u["name"] not in upgraded_to]

    row = header_row + 1
    for start in sorted(starters, key=lambda x: (x["category"], x.get("is_special", False), x["name"])):
        cat_label = CAT_LABELS.get(start["category"], start["category"])
        is_light = start.get("light", False)
        is_special = start.get("is_special", False)
        type_label = ("Special" if is_special else "Light" if is_light else "Heavy")

        ws.cell(row=row, column=1, value=cat_label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=type_label).border = THIN_BORDER

        current = start
        visited = set()
        while current and current["name"] not in visited:
            visited.add(current["name"])
            age = current.get("age", "")
            if age in AGE_ORDER:
                col = AGE_ORDER.index(age) + 3
                cat = current["category"]
                cat_data = categories.get(cat, {})
                sfd = cat_data.get("secure_flanks_defense", 0)
                dmg_taken = cat_data.get("damage_taken", 1.0)
                sdd = safe_num(current.get("strength_damage_done", 0))
                sdt = safe_num(current.get("strength_damage_taken", 0))
                strength = current.get("max_strength", 0)
                damage = strength * current.get("combat_power", 0)
                fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)
                cp = round(calc_center_power(fp, sfd), 1)
                maint = round(calc_maintenance(strength, cat, prices), 1)
                cell = ws.cell(row=row, column=col, value=f"{loc(current['name'])} ({cp}p, {maint}g)")
                cell.border = THIN_BORDER
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            next_name = current.get("upgrades_to", "")
            current = units_by_name.get(next_name) if next_name else None

        row += 1

    auto_width(ws, min_width=10, max_width=45)
    ws.freeze_panes = f"A{header_row + 1}"


def build_category_reference(wb, categories, prices):
    """Sheet 5: Base category stats reference."""
    ws = wb.create_sheet("Category Stats")

    ws.cell(row=1, column=1, value="Unit Category Base Stats").font = TITLE_FONT
    ws.cell(row=2, column=1, value="These are inherent stats from the category, before any unit-type modifiers").font = SUBTITLE_FONT

    stats = [
        ("damage_taken", "Damage Taken Multiplier"),
        ("combat_speed", "Combat Speed"),
        ("initiative", "Initiative"),
        ("frontage", "Frontage"),
        ("flanking_ability", "Flanking Ability"),
        ("secure_flanks_defense", "Secure Flanks Defense"),
        ("supply_weight", "Supply Weight"),
        ("attrition_loss", "Extra Attrition"),
        ("food_storage_per_strength", "Food Storage / Strength"),
        ("food_consumption_per_strength", "Food Consumption / Strength"),
        ("startup_amount", "Starting Army Amount"),
    ]

    header_row = 4
    ws.cell(row=header_row, column=1, value="Stat")
    for i, cat in enumerate(LAND_CATS, 2):
        ws.cell(row=header_row, column=i, value=CAT_LABELS[cat])
    style_header_row(ws, header_row, len(LAND_CATS) + 1)

    for r, (key, label) in enumerate(stats, header_row + 1):
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=1).font = Font(bold=True)
        for i, cat in enumerate(LAND_CATS, 2):
            val = categories.get(cat, {}).get(key, 0)
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = THIN_BORDER
            if isinstance(val, float):
                cell.number_format = NUM_FMT_2

    # Cost reference (scraped from prices/02_units.txt)
    cost_row = header_row + len(stats) + 2
    ws.cell(row=cost_row, column=1, value="Cost Reference (scraped from prices/02_units.txt)").font = Font(bold=True, size=12)
    cost_headers = ["Category", "Build Cost (gold)", "Reinforce Cost", "Maintenance (gold)"]
    for i, h in enumerate(cost_headers, 1):
        ws.cell(row=cost_row + 1, column=i, value=h)
    style_header_row(ws, cost_row + 1, len(cost_headers))

    for r, cat in enumerate(LAND_CATS, cost_row + 2):
        p = prices.get(cat, {})
        values = [
            CAT_LABELS[cat],
            p.get("build_gold", 0),
            p.get("reinforce_gold", 0),
            p.get("maintenance_gold", 0),
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = THIN_BORDER

    auto_width(ws)


def build_light_vs_heavy(wb, land_units, categories):
    """Sheet 6: Light vs Heavy comparison within same age/category."""
    ws = wb.create_sheet("Light vs Heavy")

    ws.cell(row=1, column=1, value="Light vs Heavy Unit Comparison").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Generic (non-special) units only - same age, same category").font = SUBTITLE_FONT

    headers = [
        "Age", "Category", "Unit", "Type",
        "Strength", "Combat Power", "Effective Power",
        "Combat Speed", "Initiative",
        "Flanking", "Secure Flanks",
        "Str Dmg Taken", "Morale Dmg Taken",
        "Str Dmg Done", "Morale Dmg Done",
        "Flank Power", "Center Power",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    buildable = [
        u for u in land_units
        if u.get("buildable", True)
        and not u.get("levy", False)
        and not u["name"].startswith("a_age_")
        and not u.get("is_special", False)
    ]

    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in buildable if u.get("age") == age]
        for cat in ["army_infantry", "army_cavalry"]:
            cat_units = [u for u in age_units if u["category"] == cat]
            lights = [u for u in cat_units if u.get("light")]
            heavies = [u for u in cat_units if not u.get("light")]

            if not lights or not heavies:
                continue

            for u in heavies + lights:
                cat_data = categories.get(cat, {})
                cat_label = CAT_LABELS.get(cat, cat)

                strength = u.get("max_strength", 0)
                cp = u.get("combat_power", 0)
                sfd = u.get("secure_flanks_defense", cat_data.get("secure_flanks_defense", 0))
                dmg_taken = u.get("damage_taken", cat_data.get("damage_taken", 1.0))
                sdd = safe_num(u.get("strength_damage_done", 0))
                sdt = safe_num(u.get("strength_damage_taken", 0))
                damage = strength * cp

                fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)
                center = calc_center_power(fp, sfd)

                values = [
                    AGE_LABELS[age], cat_label,
                    loc(u["name"]),
                    "Light" if u.get("light") else "Heavy",
                    strength, cp, round(damage, 2),
                    safe_num(u.get("combat_speed", cat_data.get("combat_speed", 1))),
                    safe_num(u.get("initiative", cat_data.get("initiative", 1))),
                    safe_num(u.get("flanking_ability", cat_data.get("flanking_ability", 1.0))),
                    safe_num(u.get("secure_flanks_defense", cat_data.get("secure_flanks_defense", 0))),
                    safe_num(u.get("strength_damage_taken", 0)),
                    safe_num(u.get("morale_damage_taken", 0)),
                    safe_num(u.get("strength_damage_done", 0)),
                    safe_num(u.get("morale_damage_done", 0)),
                    fp, center,
                ]
                for i, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=i, value=v)
                    cell.border = THIN_BORDER
                    if isinstance(v, float):
                        cell.number_format = NUM_FMT_2
                    cat_fill = CAT_FILLS.get(cat_label)
                    if cat_fill:
                        cell.fill = cat_fill

                row += 1
            row += 1  # gap between comparisons

    auto_width(ws)
    ws.freeze_panes = f"A{header_row + 1}"


# ---------------------------------------------------------------------------
# Combined Arms Optimization
# ---------------------------------------------------------------------------

# The 6 combined-arms types (light/heavy count as separate)
CA_TYPES = [
    ("Heavy Infantry", "army_infantry", False),
    ("Light Infantry", "army_infantry", True),
    ("Heavy Cavalry", "army_cavalry", False),
    ("Light Cavalry", "army_cavalry", True),
    ("Artillery", "army_artillery", None),  # None = no light/heavy split
    ("Auxiliary", "army_auxiliary", None),
]


def build_artillery_barrage(wb, land_units, forts, prices):
    """Sheet: Artillery barrage bonus requirements.

    Shows how many artillery units of each type are needed to achieve each
    barrage bonus level (+1..+8) against every fort at various strength levels.
    Formula: CEILING(bonus / ((barrage * strength - fort_level) / 3), 1)
    """
    ws = wb.create_sheet("Artillery Barrage")

    ws.cell(row=1, column=1, value="Artillery Barrage Requirements").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Units needed per barrage bonus level against forts at various strength levels"
            ).font = SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Formula: CEILING(bonus / ((barrage * strength - fort_level) / 3), 1)")

    HEADER_ROW = 6
    STRENGTHS = [1.0, 0.95, 0.90, 0.75]

    # Collect unique barrage levels from artillery units, grouped by age
    # Skip age 1 (no actual artillery exists in-game despite template)
    barrage_by_age = {}  # age -> (barrage, unit_name)
    units_by_barrage = {}  # barrage -> [unit_names]
    for u in land_units:
        if u["category"] != "army_artillery":
            continue
        if u.get("age") == "age_1_traditions":
            continue
        b = u.get("artillery_barrage", 0)
        if b <= 0:
            continue
        age = u.get("age", "")
        if not age:
            continue
        # Prefer the base buildable (non-special, non-levy) unit for the name
        is_base = u.get("buildable", True) and not u.get("is_special", False) and not u.get("levy", False)
        if age not in barrage_by_age or is_base:
            barrage_by_age[age] = (b, u["name"])
        units_by_barrage.setdefault(b, []).append(loc(u["name"]))

    # Build ordered list: (age_label, barrage)
    barrage_rows = []
    for age in AGE_ORDER:
        if age in barrage_by_age:
            b, unit_name = barrage_by_age[age]
            barrage_rows.append((f"{AGE_LABELS[age]} - {loc(unit_name)}", b))

    # Alternating row fills by barrage group
    fill_cycle = [
        CAT_FILLS["Infantry"],   # light blue
        CAT_FILLS["Cavalry"],    # light green
        CAT_FILLS["Artillery"],  # light orange
        CAT_FILLS["Auxiliary"],  # light gray
    ]

    # Header row
    label_headers = ["Artillery", "Barrage", "Strength", "Fort", "Fort Level"]
    for col_idx, h in enumerate(label_headers, 1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=h)
    for i in range(8):
        cell = ws.cell(row=HEADER_ROW, column=6 + i, value=i + 1)
        cell.number_format = '"+"0'
    style_header_row(ws, HEADER_ROW, 13)

    # Data rows
    row = HEADER_ROW + 1
    for b_idx, (label, barrage) in enumerate(barrage_rows):
        fill = fill_cycle[b_idx % len(fill_cycle)]

        for fort in forts:
            fort_name = loc(fort["name"])
            fort_level = fort["fort_level"]

            for strength in STRENGTHS:
                # Skip rows where +1 bonus is impossible
                if barrage * strength - fort_level <= 0:
                    continue

                ws.cell(row=row, column=1, value=label).fill = fill
                ws.cell(row=row, column=1).border = THIN_BORDER

                c = ws.cell(row=row, column=2, value=barrage)
                c.fill = fill
                c.alignment = Alignment(horizontal="center")
                c.border = THIN_BORDER

                c = ws.cell(row=row, column=3, value=strength)
                c.fill = fill
                c.alignment = Alignment(horizontal="center")
                c.number_format = NUM_FMT_PCT
                c.border = THIN_BORDER

                c = ws.cell(row=row, column=4, value=fort_name)
                c.fill = fill
                c.border = THIN_BORDER

                c = ws.cell(row=row, column=5, value=fort_level)
                c.fill = fill
                c.alignment = Alignment(horizontal="center")
                c.border = THIN_BORDER

                for i in range(8):
                    col = 6 + i
                    cl = get_column_letter(col)
                    formula = (
                        f'=CEILING({cl}${HEADER_ROW}/(($B{row}*$C{row}-$E{row})/3),1)'
                    )
                    c = ws.cell(row=row, column=col, value=formula)
                    c.fill = fill
                    c.alignment = Alignment(horizontal="center")
                    c.border = THIN_BORDER

                row += 1

    # Unit reference section
    row += 1
    ws.cell(row=row, column=1, value="Units by Barrage Level").font = SUBTITLE_FONT
    row += 1
    for barrage in sorted(units_by_barrage):
        ws.cell(row=row, column=1, value=f"Barrage {barrage}:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=", ".join(units_by_barrage[barrage]))
        row += 1

    # --- Effective Barrage per Gold table (to the right of main table) ---
    # Collect base buildable artillery stats per age
    arty_units = []  # [(label, barrage, build_gold)]
    for age in AGE_ORDER:
        if age not in barrage_by_age:
            continue
        b, unit_name = barrage_by_age[age]
        for u in land_units:
            if u["name"] == unit_name:
                cost = calc_cost(safe_num(u.get("max_strength", 0)), "army_artillery", prices)
                age_num = AGE_LABELS[age].split(" - ")[0]
                arty_units.append((f"{age_num} - {loc(unit_name)}", b, cost))
                break

    # Place to the right of the main table (cols 1-13), with a gap column
    R_COL = 15  # column O

    ws.cell(row=1, column=R_COL, value="Effective Barrage per Gold by Fort").font = TITLE_FONT
    ws.cell(row=2, column=R_COL,
            value="Effective barrage = artillery_barrage - fort_level.  "
                  "Values show (effective barrage / build cost) \u00d7 1000."
            ).font = SUBTITLE_FONT

    # Header at same row as main table header
    ws.cell(row=HEADER_ROW, column=R_COL, value="Fort")
    ws.cell(row=HEADER_ROW, column=R_COL + 1, value="Level")
    for ci, (label, _, _) in enumerate(arty_units):
        ws.cell(row=HEADER_ROW, column=R_COL + 2 + ci, value=label)
    epg_num_cols = 2 + len(arty_units)
    for col in range(R_COL, R_COL + epg_num_cols):
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    epg_row = HEADER_ROW + 1
    for f_idx, fort in enumerate(forts):
        fort_name = loc(fort["name"])
        fort_level = fort["fort_level"]
        fill = fill_cycle[f_idx % len(fill_cycle)]

        c = ws.cell(row=epg_row, column=R_COL, value=fort_name)
        c.fill = fill
        c.border = THIN_BORDER

        c = ws.cell(row=epg_row, column=R_COL + 1, value=fort_level)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

        best_val = -1
        best_col = -1
        values = []
        for ci, (_, barrage, cost) in enumerate(arty_units):
            eff = barrage - fort_level
            if eff <= 0 or cost == 0:
                values.append(None)
            else:
                val = eff / cost * 1000
                values.append(val)
                if val > best_val:
                    best_val = val
                    best_col = ci

        for ci, val in enumerate(values):
            c = ws.cell(row=epg_row, column=R_COL + 2 + ci)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER
            if val is None:
                c.value = "-"
            else:
                c.value = val
                c.number_format = NUM_FMT_3
                if ci == best_col:
                    c.font = BEST_FONT

        epg_row += 1

    ws.freeze_panes = "F7"
    auto_width(ws)
    # Widen Artillery column slightly for long names like "Chambered Cannon"
    ws.column_dimensions["A"].width = int(ws.column_dimensions["A"].width * 1.15)
    # Shrink numeric columns so the sheet is compact
    ws.column_dimensions["B"].width = 7   # Barrage
    ws.column_dimensions["C"].width = 7   # Strength
    ws.column_dimensions["E"].width = 5   # Fort Level
    for i in range(8):
        ws.column_dimensions[get_column_letter(6 + i)].width = 5  # +1..+8
    # Shrink right-side table: Level column and artillery value columns
    ws.column_dimensions[get_column_letter(R_COL + 1)].width = 5  # Level
    for ci in range(len(arty_units)):
        ws.column_dimensions[get_column_letter(R_COL + 2 + ci)].width = 14


def get_best_generic_units(land_units, categories):
    """For each age and each of the 6 CA types, find the best generic unit by flank power.

    Returns: { age: [ { "type_label", "unit_name", "flank_power", "center_power", ... }, ... ] }
    """
    result = {}
    for age in AGE_ORDER:
        age_types = []
        for type_label, cat, is_light in CA_TYPES:
            cat_data = categories.get(cat, {})
            sfd = cat_data.get("secure_flanks_defense", 0)
            dmg_taken = cat_data.get("damage_taken", 1.0)

            candidates = [
                u for u in land_units
                if u.get("age") == age
                and u["category"] == cat
                and u.get("buildable", True)
                and not u.get("levy", False)
                and not u["name"].startswith("a_age_")
                and not u.get("is_special", False)
                and (is_light is None or u.get("light", False) == is_light)
            ]

            if not candidates:
                age_types.append({
                    "type_label": type_label, "unit_name": "-",
                    "flank_power": 0, "center_power": 0,
                    "strength": 0, "combat_power": 0,
                    "str_dmg_done": 0, "str_dmg_taken": 0,
                    "initiative": 0,
                })
                continue

            best = None
            best_fp = -1
            for u in candidates:
                strength = u.get("max_strength", 0)
                cp = u.get("combat_power", 0)
                sdd = safe_num(u.get("strength_damage_done", 0))
                sdt = safe_num(u.get("strength_damage_taken", 0))
                damage = strength * cp
                fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)
                if fp > best_fp:
                    best_fp = fp
                    best = {
                        "type_label": type_label,
                        "unit_name": u["name"],
                        "flank_power": fp,
                        "center_power": calc_center_power(fp, sfd),
                        "strength": strength,
                        "combat_power": cp,
                        "str_dmg_done": sdd,
                        "str_dmg_taken": sdt,
                        "initiative": safe_num(u.get("initiative", cat_data.get("initiative", 0))),
                    }
            age_types.append(best)

        result[age] = age_types
    return result


CENTER_RATIO = 1 / 3  # ~33% of army is center, ~67% flanks
FLANK_RATIO = 1 - CENTER_RATIO


def calc_positional_power(pcts, flank_powers, center_powers):
    """Calculate total army power with optimal positional placement.

    Units are assigned to center (33%) or flanks (67%) to maximize power.
    Each unit type uses its center_power when in center, flank_power on flanks.
    The center/flank benefit ratio determines which types go where.
    """
    n = len(pcts)
    if sum(pcts) < 1e-9:
        return 0

    # For each type, compute the center advantage ratio:
    # how much MORE power it gets from center vs flank (relative)
    # Types with highest center advantage should go to center first.
    type_info = []
    for i in range(n):
        if pcts[i] < 1e-9:
            continue
        fp = flank_powers[i]
        cp = center_powers[i]
        # center_advantage = extra power gained per unit in center vs flank
        advantage = cp - fp  # positive means center is better
        type_info.append((i, pcts[i], fp, cp, advantage))

    # Sort by center advantage descending — best center units first
    type_info.sort(key=lambda x: x[4], reverse=True)

    # Fill center slots (33% of army), then flanks (67%)
    center_remaining = CENTER_RATIO
    total_power = 0

    for idx, pct, fp, cp, adv in type_info:
        # How much of this type goes to center?
        in_center = min(pct, center_remaining)
        in_flank = pct - in_center
        center_remaining -= in_center

        total_power += in_center * cp + in_flank * fp

    return total_power


def optimize_composition(flank_powers, center_powers, combined_arms):
    """Find optimal percentage allocation to maximize total positional power.

    Uses positional placement: 33% center, 67% flanks.
    Each type placed optimally based on its center vs flank advantage.

    Returns: (best_pcts, best_total, bonus_used, n_qualifying)
    """
    bonus_per_type = combined_arms["bonus_per_type"]
    min_pct = combined_arms["min_percent"]
    max_pct = combined_arms["max_threshold"] - 0.01  # must be strictly under threshold
    n = len(flank_powers)

    best_total = -1
    best_pcts = [0.0] * n
    best_bonus = 0.0
    best_n_qual = 0

    # Enumerate all 2^n subsets of qualifying types
    for mask in range(1 << n):
        qualifying = [i for i in range(n) if mask & (1 << i)]
        k = len(qualifying)

        if k * min_pct > 1.0:
            continue

        if any(flank_powers[i] <= 0 for i in qualifying):
            continue

        # Allocate minimums
        pcts = [0.0] * n
        for i in qualifying:
            pcts[i] = min_pct
        remaining = 1.0 - k * min_pct

        # Greedily allocate remaining to types that contribute most power.
        # Use a blended power estimate for greedy ordering:
        # assume the marginal unit goes ~67% flank, ~33% center
        blended = [FLANK_RATIO * flank_powers[i] + CENTER_RATIO * center_powers[i]
                    for i in range(n)]
        sorted_indices = sorted(range(n), key=lambda i: blended[i], reverse=True)
        for i in sorted_indices:
            if blended[i] <= 0:
                continue
            room = max_pct - pcts[i]
            add = min(remaining, room)
            if add > 0:
                pcts[i] += add
                remaining -= add
            if remaining <= 1e-9:
                break

        bonus = bonus_per_type * k
        if any(p > max_pct + 1e-9 for p in pcts):
            bonus = 0

        weighted = calc_positional_power(pcts, flank_powers, center_powers)
        total = weighted * (1 + bonus)

        if total > best_total:
            best_total = total
            best_pcts = pcts[:]
            best_bonus = bonus
            best_n_qual = k

    # Also try no-bonus: 100% in best type
    for i in range(n):
        # 100% of one type — it fills both center and flank
        power = CENTER_RATIO * center_powers[i] + FLANK_RATIO * flank_powers[i]
        if power > best_total:
            best_total = power
            best_pcts = [0.0] * n
            best_pcts[i] = 1.0
            best_bonus = 0.0
            best_n_qual = 0

    return best_pcts, best_total, best_bonus, best_n_qual


def build_optimal_composition(wb, land_units, categories, combined_arms):
    """Sheet: Optimal army composition per age using combined arms."""
    ws = wb.create_sheet("Optimal Comp (K-D)")

    ws.cell(row=1, column=1, value="Optimal Army Composition (K/D)").font = TITLE_FONT

    # Show scraped combined arms values
    ca = combined_arms
    ws.cell(row=2, column=1,
            value=f"Scraped: bonus/type={ca['bonus_per_type']:.1%}, "
                  f"min threshold={ca['min_percent']:.0%}, "
                  f"max cap={ca['max_threshold']:.0%}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Positional placement: 33% center / 67% flanks. "
                  "Units assigned to center/flank to maximize power.").font = Font(italic=True)
    ws.cell(row=4, column=1,
            value="Formulas: EffDmg = Str*CP*(1+StrDmgDone), "
                  "Power = EffDmg*10/((1-SFD)*DmgTaken*(1+StrDmgTaken)), "
                  "Total = sum(positional_power) * (1+bonus)").font = Font(italic=True)

    best_units = get_best_generic_units(land_units, categories)

    row = 6
    for age in AGE_ORDER:
        types = best_units[age]

        # Age header
        ws.cell(row=row, column=1, value=AGE_LABELS[age]).font = Font(bold=True, size=13)
        row += 1

        # Column headers
        headers = [
            "Type", "Unit", "Optimal %", "Positional Power",
            "Flank Power", "Center Power",
            "Strength", "CombatPower",
            "Str Dmg Done", "Str Dmg Taken", "Initiative",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        flank_powers = [t["flank_power"] for t in types]
        center_powers = [t["center_power"] for t in types]

        pcts, total, bonus, nq = optimize_composition(
            flank_powers, center_powers, combined_arms
        )

        # Compute per-type positional contribution
        center_remaining = CENTER_RATIO
        type_positional = [0.0] * len(types)
        for idx, pct, fp, cp, adv in sorted(
            [(i, pcts[i], flank_powers[i], center_powers[i], center_powers[i] - flank_powers[i])
             for i in range(len(types)) if pcts[i] > 1e-9],
            key=lambda x: x[4], reverse=True,
        ):
            in_center = min(pct, center_remaining)
            in_flank = pct - in_center
            center_remaining -= in_center
            type_positional[idx] = in_center * cp + in_flank * fp

        for i, t in enumerate(types):
            pct = pcts[i]
            values = [
                t["type_label"], loc(t["unit_name"]),
                pct, type_positional[i],
                t["flank_power"], t["center_power"],
                t["strength"], t["combat_power"],
                t["str_dmg_done"], t["str_dmg_taken"], t["initiative"],
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                if j == 3:  # Optimal %
                    cell.number_format = "0%"
            if pct >= combined_arms["min_percent"] - 1e-9:
                ws.cell(row=row, column=1).font = BEST_FONT
                ws.cell(row=row, column=3).font = BEST_FONT
            row += 1

        # Summary rows
        base_power = calc_positional_power(pcts, flank_powers, center_powers)

        ws.cell(row=row, column=1, value="Base Power (no bonus)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=base_power).number_format = NUM_FMT_2
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value=f"Combined Arms Bonus ({nq} types)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=bonus).number_format = "0.0%"
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value="Total Power (with bonus)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=total).number_format = NUM_FMT_2
        ws.cell(row=row, column=4).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 2

    auto_width(ws, max_width=35)
    ws.freeze_panes = "A6"


def calc_morale_flank_power(damage, sfd, dmg_taken, morale_dmg_done=0, morale_dmg_taken=0):
    """Same formula as strength flank power but with morale modifiers."""
    eff_damage = damage * (1 + morale_dmg_done)
    denom = (1 - sfd) * dmg_taken * (1 + morale_dmg_taken)
    if denom == 0:
        return 0
    return eff_damage * 10 / denom


def get_best_generic_units_morale(land_units, categories):
    """Same as get_best_generic_units but ranked by morale power."""
    result = {}
    for age in AGE_ORDER:
        age_types = []
        for type_label, cat, is_light in CA_TYPES:
            cat_data = categories.get(cat, {})
            sfd = cat_data.get("secure_flanks_defense", 0)
            dmg_taken = cat_data.get("damage_taken", 1.0)

            candidates = [
                u for u in land_units
                if u.get("age") == age
                and u["category"] == cat
                and u.get("buildable", True)
                and not u.get("levy", False)
                and not u["name"].startswith("a_age_")
                and not u.get("is_special", False)
                and (is_light is None or u.get("light", False) == is_light)
            ]

            if not candidates:
                age_types.append({
                    "type_label": type_label, "unit_name": "-",
                    "flank_power": 0, "center_power": 0,
                    "strength": 0, "combat_power": 0,
                    "morale_dmg_done": 0, "morale_dmg_taken": 0,
                    "initiative": 0,
                })
                continue

            best = None
            best_fp = -1
            for u in candidates:
                strength = u.get("max_strength", 0)
                cp = u.get("combat_power", 0)
                mdd = safe_num(u.get("morale_damage_done", 0))
                mdt = safe_num(u.get("morale_damage_taken", 0))
                damage = strength * cp
                fp = calc_morale_flank_power(damage, sfd, dmg_taken, mdd, mdt)
                if fp > best_fp:
                    best_fp = fp
                    best = {
                        "type_label": type_label,
                        "unit_name": u["name"],
                        "flank_power": fp,
                        "center_power": calc_center_power(fp, sfd),
                        "strength": strength,
                        "combat_power": cp,
                        "morale_dmg_done": mdd,
                        "morale_dmg_taken": mdt,
                        "initiative": safe_num(u.get("initiative", cat_data.get("initiative", 0))),
                    }
            age_types.append(best)

        result[age] = age_types
    return result


def build_optimal_composition_morale(wb, land_units, categories, combined_arms):
    """Sheet: Optimal army composition per age by morale power."""
    ws = wb.create_sheet("Optimal Comp (Morale)")

    ws.cell(row=1, column=1, value="Optimal Army Composition (Morale Power)").font = TITLE_FONT

    ca = combined_arms
    ws.cell(row=2, column=1,
            value=f"Scraped: bonus/type={ca['bonus_per_type']:.1%}, "
                  f"min threshold={ca['min_percent']:.0%}, "
                  f"max cap={ca['max_threshold']:.0%}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Same formula as strength but using morale_damage_done/taken. "
                  "Morale is flat per unit (not scaled by strength).").font = Font(italic=True)

    best_units = get_best_generic_units_morale(land_units, categories)

    row = 5
    for age in AGE_ORDER:
        types = best_units[age]

        ws.cell(row=row, column=1, value=AGE_LABELS[age]).font = Font(bold=True, size=13)
        row += 1

        headers = [
            "Type", "Unit", "Optimal %", "Positional Power",
            "Flank Power", "Center Power",
            "Strength", "CombatPower",
            "Morale Dmg Done", "Morale Dmg Taken", "Initiative",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        flank_powers = [t["flank_power"] for t in types]
        center_powers = [t["center_power"] for t in types]

        pcts, total, bonus, nq = optimize_composition(
            flank_powers, center_powers, combined_arms
        )

        center_remaining = CENTER_RATIO
        type_positional = [0.0] * len(types)
        for idx, pct, fp, cp, adv in sorted(
            [(i, pcts[i], flank_powers[i], center_powers[i], center_powers[i] - flank_powers[i])
             for i in range(len(types)) if pcts[i] > 1e-9],
            key=lambda x: x[4], reverse=True,
        ):
            in_center = min(pct, center_remaining)
            in_flank = pct - in_center
            center_remaining -= in_center
            type_positional[idx] = in_center * cp + in_flank * fp

        for i, t in enumerate(types):
            pct = pcts[i]
            values = [
                t["type_label"], loc(t["unit_name"]),
                pct, type_positional[i],
                t["flank_power"], t["center_power"],
                t["strength"], t["combat_power"],
                t["morale_dmg_done"], t["morale_dmg_taken"], t["initiative"],
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                if j == 3:
                    cell.number_format = "0%"
            if pct >= combined_arms["min_percent"] - 1e-9:
                ws.cell(row=row, column=1).font = BEST_FONT
                ws.cell(row=row, column=3).font = BEST_FONT
            row += 1

        base_power = calc_positional_power(pcts, flank_powers, center_powers)

        ws.cell(row=row, column=1, value="Base Morale Power (no bonus)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=base_power).number_format = NUM_FMT_2
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value=f"Combined Arms Bonus ({nq} types)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=bonus).number_format = "0.0%"
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value="Total Morale Power (with bonus)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=total).number_format = NUM_FMT_2
        ws.cell(row=row, column=4).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 2

    auto_width(ws, max_width=35)
    ws.freeze_panes = "A5"


def get_cheapest_units(land_units, categories):
    """For each CA type, find the cheapest generic unit across ALL ages."""
    result = {}
    for type_label, cat, is_light in CA_TYPES:
        cat_data = categories.get(cat, {})
        sfd = cat_data.get("secure_flanks_defense", 0)
        dmg_taken = cat_data.get("damage_taken", 1.0)

        candidates = [
            u for u in land_units
            if u["category"] == cat
            and u.get("buildable", True)
            and not u.get("levy", False)
            and not u["name"].startswith("a_age_")
            and not u.get("is_special", False)
            and (is_light is None or u.get("light", False) == is_light)
        ]
        if not candidates:
            result[type_label] = None
            continue

        cheapest = min(candidates, key=lambda u: u.get("max_strength", 999))
        strength = cheapest.get("max_strength", 0)
        cp = cheapest.get("combat_power", 0)
        sdd = safe_num(cheapest.get("strength_damage_done", 0))
        sdt = safe_num(cheapest.get("strength_damage_taken", 0))
        damage = strength * cp
        fp = calc_flank_power(damage, sfd, dmg_taken, sdd, sdt)

        result[type_label] = {
            "type_label": type_label,
            "unit_name": cheapest["name"],
            "age": cheapest.get("age", ""),
            "flank_power": fp,
            "center_power": calc_center_power(fp, sfd),
            "strength": strength,
            "combat_power": cp,
        }
    return result


def optimize_budget(best_units_age, cheapest_units, prices, combined_arms):
    """Find optimal composition allowing cheap filler units from any age.

    For each qualifying set, each type can use the best (current age) or cheapest
    (any age) unit. Types at minimum 10% use cheapest to save gold. Types getting
    extra allocation use best for power.

    Maximizes: sum(r_i * power_i) / sum(r_i * cost_i) * (1 + bonus)
    """
    ca = combined_arms
    bonus_per_type = ca["bonus_per_type"]
    min_pct = ca["min_percent"]
    max_pct = ca["max_threshold"] - 0.01  # must be strictly under threshold
    n = len(best_units_age)

    # Precompute costs and powers for best and cheapest variants
    best_fp = [t["flank_power"] for t in best_units_age]
    best_cp = [t["center_power"] for t in best_units_age]
    best_costs = []
    cheap_fp = []
    cheap_cp = []
    cheap_costs = []
    for t in best_units_age:
        cat_key = [c for label, c, _ in CA_TYPES if label == t["type_label"]][0]
        best_costs.append(calc_maintenance(t["strength"], cat_key, prices))
        ch = cheapest_units.get(t["type_label"])
        if ch:
            cheap_fp.append(ch["flank_power"])
            cheap_cp.append(ch["center_power"])
            cheap_costs.append(calc_maintenance(ch["strength"], cat_key, prices))
        else:
            cheap_fp.append(0)
            cheap_cp.append(0)
            cheap_costs.append(0)

    best_result = None
    best_ratio = -1

    # Enumerate all qualifying subsets
    for qual_mask in range(1 << n):
        qualifying = [i for i in range(n) if qual_mask & (1 << i)]
        k = len(qualifying)
        if k * min_pct > 1.0:
            continue
        # Skip if any qualifying type has no unit available
        if any(best_fp[i] <= 0 and cheap_fp[i] <= 0 for i in qualifying):
            continue

        bonus = bonus_per_type * k

        # Enumerate all filler subsets within qualifying set
        # Filler types use cheapest unit at exactly 10%
        # Power types use best unit, allocated greedily
        for filler_mask in range(1 << k):
            fillers = [qualifying[j] for j in range(k) if filler_mask & (1 << j)]
            powers = [qualifying[j] for j in range(k) if not (filler_mask & (1 << j))]

            # Skip if a filler type has no cheapest unit
            if any(cheap_costs[i] <= 0 for i in fillers):
                continue
            # Skip if a power type has no best unit
            if any(best_fp[i] <= 0 for i in powers):
                continue

            # Allocate: fillers at 10%, power types get remaining greedily
            pcts = [0.0] * n
            use_cheap = [False] * n
            for i in fillers:
                pcts[i] = min_pct
                use_cheap[i] = True
            for i in powers:
                pcts[i] = min_pct  # minimum for qualifying

            remaining = 1.0 - k * min_pct

            # Greedily fill power types (by best unit power, descending) up to 50%
            for i in sorted(powers, key=lambda i: best_fp[i], reverse=True):
                room = max_pct - pcts[i]
                add = min(remaining, room)
                if add > 0:
                    pcts[i] += add
                    remaining -= add
                if remaining <= 1e-9:
                    break

            # If remaining, try filling non-qualifying power types
            if remaining > 1e-9:
                non_qual = [i for i in range(n) if i not in qualifying and best_fp[i] > 0]
                for i in sorted(non_qual, key=lambda i: best_fp[i], reverse=True):
                    room = max_pct - pcts[i]
                    add = min(remaining, room)
                    if add > 0:
                        pcts[i] += add
                        remaining -= add
                    if remaining <= 1e-9:
                        break

            # Check max threshold
            if any(p > max_pct + 1e-9 for p in pcts):
                continue

            # Compute true ratio with positional placement
            unit_fp = [cheap_fp[i] if use_cheap[i] else best_fp[i] for i in range(n)]
            unit_cp = [cheap_cp[i] if use_cheap[i] else best_cp[i] for i in range(n)]
            unit_costs = [cheap_costs[i] if use_cheap[i] else best_costs[i] for i in range(n)]

            total_power = calc_positional_power(pcts, unit_fp, unit_cp)
            total_cost = sum(pcts[i] * unit_costs[i] for i in range(n))
            if total_cost <= 0:
                continue

            ratio = total_power / total_cost * (1 + bonus)
            if ratio > best_ratio:
                best_ratio = ratio
                best_result = {
                    "pcts": pcts[:],
                    "use_cheap": use_cheap[:],
                    "bonus": bonus,
                    "nq": k,
                    "ratio": ratio,
                    "total_power": total_power,
                    "total_cost": total_cost,
                }

    # Also try no-bonus: 100% of single best type
    for i in range(n):
        if best_costs[i] <= 0:
            continue
        power = CENTER_RATIO * best_cp[i] + FLANK_RATIO * best_fp[i]
        ratio = power / best_costs[i]
        if ratio > best_ratio:
            best_ratio = ratio
            pcts = [0.0] * n
            pcts[i] = 1.0
            best_result = {
                "pcts": pcts,
                "use_cheap": [False] * n,
                "bonus": 0,
                "nq": 0,
                "ratio": ratio,
                "total_power": power,
                "total_cost": best_costs[i],
            }

    return best_result


def build_optimal_composition_budget(wb, land_units, categories, combined_arms, prices):
    """Sheet: Optimal composition allowing cheap filler units from any age."""
    ws = wb.create_sheet("Optimal Comp (Gold-Mixed Age)")

    ws.cell(row=1, column=1, value="Optimal Army Composition (Gold-Mixed Age)").font = TITLE_FONT

    ca = combined_arms
    ws.cell(row=2, column=1,
            value=f"Scraped: bonus/type={ca['bonus_per_type']:.1%}, "
                  f"min threshold={ca['min_percent']:.0%}, "
                  f"max cap={ca['max_threshold']:.0%}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Uses cheapest unit from ANY age as filler to hit combined arms thresholds. "
                  "Maximizes power per gold for equal-budget armies.").font = Font(italic=True)

    best_by_age = get_best_generic_units(land_units, categories)
    cheapest = get_cheapest_units(land_units, categories)

    row = 5
    for age in AGE_ORDER:
        types = best_by_age[age]

        ws.cell(row=row, column=1, value=AGE_LABELS[age]).font = Font(bold=True, size=13)
        row += 1

        headers = [
            "Type", "Unit", "Optimal %", "Role",
            "Power", "Maintenance",
            "Strength", "Age",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        result = optimize_budget(types, cheapest, prices, combined_arms)
        if not result:
            ws.cell(row=row, column=1, value="No valid composition found")
            row += 2
            continue

        pcts = result["pcts"]
        use_cheap = result["use_cheap"]

        for i, t in enumerate(types):
            pct = pcts[i]
            if pct < 1e-9:
                continue

            ch = cheapest.get(t["type_label"])
            if use_cheap[i] and ch:
                unit_name = ch["unit_name"]
                unit_age = ch["age"]
                cat_key = [c for label, c, _ in CA_TYPES if label == t["type_label"]][0]
                power = ch["center_power"]
                cost = calc_maintenance(ch["strength"], cat_key, prices)
                strength = ch["strength"]
                role = "Filler"
            else:
                unit_name = t["unit_name"]
                unit_age = age
                cat_key = [c for label, c, _ in CA_TYPES if label == t["type_label"]][0]
                power = t["center_power"]
                cost = calc_maintenance(t["strength"], cat_key, prices)
                strength = t["strength"]
                role = "Power"

            values = [
                t["type_label"], loc(unit_name),
                pct, role,
                power, cost,
                strength, AGE_LABELS.get(unit_age, "?"),
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                if j == 3:
                    cell.number_format = "0%"
            if role == "Power":
                ws.cell(row=row, column=1).font = BEST_FONT
            row += 1

        # Summary
        base_ratio = result["ratio"] / (1 + result["bonus"]) if result["bonus"] > 0 else result["ratio"]

        ws.cell(row=row, column=1, value="Avg Maintenance/Regiment").font = Font(bold=True)
        ws.cell(row=row, column=6, value=result["total_cost"]).number_format = NUM_FMT_2
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=6).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value=f"Combined Arms Bonus ({result['nq']} types)").font = Font(bold=True)
        ws.cell(row=row, column=5, value=result["bonus"]).number_format = "0.0%"
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=5).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value="Power/Maintenance Ratio").font = Font(bold=True, size=12)
        ws.cell(row=row, column=5, value=result["ratio"]).number_format = NUM_FMT_2
        ws.cell(row=row, column=5).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=5).border = THIN_BORDER
        row += 2

    auto_width(ws, max_width=35)
    ws.freeze_panes = "A5"


def build_optimal_composition_gold(wb, land_units, categories, combined_arms, prices):
    """Sheet: Optimal army composition per age by power-per-gold."""
    ws = wb.create_sheet("Optimal Comp (Gold-Same Age)")

    ws.cell(row=1, column=1, value="Optimal Army Composition (Gold-Same Age)").font = TITLE_FONT

    ca = combined_arms
    ws.cell(row=2, column=1,
            value=f"Scraped: bonus/type={ca['bonus_per_type']:.1%}, "
                  f"min threshold={ca['min_percent']:.0%}, "
                  f"max cap={ca['max_threshold']:.0%}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Positional placement: 33% center / 67% flanks. "
                  "Optimizes power/gold instead of raw power.").font = Font(italic=True)
    ws.cell(row=4, column=1,
            value="Power/Gold = Power / (Strength * BaseBuildCost * 10) * 100").font = Font(italic=True)

    best_units = get_best_generic_units(land_units, categories)

    row = 6
    for age in AGE_ORDER:
        types = best_units[age]

        ws.cell(row=row, column=1, value=AGE_LABELS[age]).font = Font(bold=True, size=13)
        row += 1

        headers = [
            "Type", "Unit", "Optimal %", "Positional P/Gold",
            "Flank P/Gold", "Center P/Gold",
            "Strength", "CombatPower",
            "Str Dmg Done", "Str Dmg Taken", "Initiative",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        # Compute per-type power/gold for display, and raw power + costs for true ratio
        flank_pg = []
        center_pg = []
        costs = []
        flank_powers = [t["flank_power"] for t in types]
        center_powers = [t["center_power"] for t in types]
        for t in types:
            cat_key = [c for label, c, _ in CA_TYPES if label == t["type_label"]][0]
            cost = calc_cost(t["strength"], cat_key, prices)
            costs.append(cost)
            if cost > 0:
                flank_pg.append(t["flank_power"] / cost * 100)
                center_pg.append(t["center_power"] / cost * 100)
            else:
                flank_pg.append(0)
                center_pg.append(0)

        # Optimize using power/cost to find best composition
        pcts, _, bonus, nq = optimize_composition(
            flank_pg, center_pg, combined_arms
        )

        # Compute true ratio: total_power / total_cost
        total_power = calc_positional_power(pcts, flank_powers, center_powers)
        total_cost = sum(pcts[i] * costs[i] for i in range(len(types)))
        true_ratio = total_power / total_cost * 100 if total_cost > 0 else 0
        base_ratio = true_ratio / (1 + bonus) if bonus > 0 else true_ratio

        # Per-type positional contribution (using power/gold for display)
        center_remaining = CENTER_RATIO
        type_positional = [0.0] * len(types)
        for idx, pct, fpg, cpg, adv in sorted(
            [(i, pcts[i], flank_pg[i], center_pg[i], center_pg[i] - flank_pg[i])
             for i in range(len(types)) if pcts[i] > 1e-9],
            key=lambda x: x[4], reverse=True,
        ):
            in_center = min(pct, center_remaining)
            in_flank = pct - in_center
            center_remaining -= in_center
            type_positional[idx] = in_center * cpg + in_flank * fpg

        for i, t in enumerate(types):
            pct = pcts[i]
            values = [
                t["type_label"], loc(t["unit_name"]),
                pct, type_positional[i],
                flank_pg[i], center_pg[i],
                t["strength"], t["combat_power"],
                t["str_dmg_done"], t["str_dmg_taken"], t["initiative"],
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                if j == 3:
                    cell.number_format = "0%"
            if pct >= combined_arms["min_percent"] - 1e-9:
                ws.cell(row=row, column=1).font = BEST_FONT
                ws.cell(row=row, column=3).font = BEST_FONT
            row += 1

        ws.cell(row=row, column=1, value="Base P/Gold (no bonus)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=base_ratio).number_format = NUM_FMT_2
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value=f"Combined Arms Bonus ({nq} types)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=bonus).number_format = "0.0%"
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value="Total P/Gold (with bonus)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=true_ratio).number_format = NUM_FMT_2
        ws.cell(row=row, column=4).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 2

    auto_width(ws, max_width=35)
    ws.freeze_panes = "A6"


def calc_unit_iron(unit, categories, goods_demands, production_recipes):
    """Get total iron cost for a unit by resolving its construction demand."""
    ref = (unit.get("construction_demand", "")
           or categories.get(unit["category"], {}).get("construction_demand", ""))
    goods = goods_demands.get(ref, {})
    if not goods:
        return 0
    raw = resolve_raw_materials(goods, production_recipes)
    return raw.get("iron", 0)


def build_optimal_composition_iron(wb, land_units, categories, combined_arms,
                                    goods_demands, production_recipes):
    """Sheet: Optimal army composition per age by power-per-iron."""
    ws = wb.create_sheet("Optimal Comp (Iron)")

    ws.cell(row=1, column=1, value="Optimal Army Composition (Power per Iron)").font = TITLE_FONT

    ca = combined_arms
    ws.cell(row=2, column=1,
            value=f"Scraped: bonus/type={ca['bonus_per_type']:.1%}, "
                  f"min threshold={ca['min_percent']:.0%}, "
                  f"max cap={ca['max_threshold']:.0%}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Positional placement: 33% center / 67% flanks. "
                  "Optimizes power/iron (resolved via workshop recipes).").font = Font(italic=True)

    best_units = get_best_generic_units(land_units, categories)

    row = 5
    for age in AGE_ORDER:
        types = best_units[age]

        ws.cell(row=row, column=1, value=AGE_LABELS[age]).font = Font(bold=True, size=13)
        row += 1

        headers = [
            "Type", "Unit", "Optimal %", "Positional P/Iron",
            "Iron Cost", "Flank P/Iron", "Center P/Iron",
            "Flank Power", "Center Power",
            "Strength", "CombatPower",
            "Str Dmg Done", "Str Dmg Taken", "Initiative",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        # Compute power-per-iron for each type, and raw power + iron costs for true ratio
        flank_pi = []
        center_pi = []
        iron_costs = []
        flank_powers = [t["flank_power"] for t in types]
        center_powers = [t["center_power"] for t in types]
        for t in types:
            if t["unit_name"] == "-":
                flank_pi.append(0)
                center_pi.append(0)
                iron_costs.append(0)
                continue
            u = next((u for u in land_units if u["name"] == t["unit_name"]), None)
            iron = calc_unit_iron(u, categories, goods_demands, production_recipes) if u else 0
            iron_costs.append(iron)
            if iron > 0:
                flank_pi.append(t["flank_power"] / iron)
                center_pi.append(t["center_power"] / iron)
            else:
                flank_pi.append(t["flank_power"] * 1000 if t["flank_power"] > 0 else 0)
                center_pi.append(t["center_power"] * 1000 if t["center_power"] > 0 else 0)

        # Optimize using power/iron to find best composition
        pcts, _, bonus, nq = optimize_composition(
            flank_pi, center_pi, combined_arms
        )

        # Compute true ratio: total_power / total_iron
        total_power = calc_positional_power(pcts, flank_powers, center_powers)
        total_iron = sum(pcts[i] * iron_costs[i] for i in range(len(types)))
        true_ratio = total_power / total_iron if total_iron > 0 else 0
        base_ratio = true_ratio / (1 + bonus) if bonus > 0 else true_ratio

        # Per-type positional contribution (using power/iron for display)
        center_remaining = CENTER_RATIO
        type_positional = [0.0] * len(types)
        for idx, pct, fpi, cpi, adv in sorted(
            [(i, pcts[i], flank_pi[i], center_pi[i], center_pi[i] - flank_pi[i])
             for i in range(len(types)) if pcts[i] > 1e-9],
            key=lambda x: x[4], reverse=True,
        ):
            in_center = min(pct, center_remaining)
            in_flank = pct - in_center
            center_remaining -= in_center
            type_positional[idx] = in_center * cpi + in_flank * fpi

        for i, t in enumerate(types):
            pct = pcts[i]
            values = [
                t["type_label"], loc(t["unit_name"]),
                pct, type_positional[i],
                iron_costs[i],
                flank_pi[i] if iron_costs[i] > 0 else "",
                center_pi[i] if iron_costs[i] > 0 else "",
                t["flank_power"], t["center_power"],
                t["strength"], t["combat_power"],
                t["str_dmg_done"], t["str_dmg_taken"], t["initiative"],
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                if j == 3:  # optimal %
                    cell.number_format = "0%"
                if j == 5:  # iron cost
                    cell.number_format = NUM_FMT_3
            if pct >= combined_arms["min_percent"] - 1e-9:
                ws.cell(row=row, column=1).font = BEST_FONT
                ws.cell(row=row, column=3).font = BEST_FONT
            row += 1

        ws.cell(row=row, column=1, value="Base P/Iron (no bonus)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=base_ratio).number_format = NUM_FMT_2
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value=f"Combined Arms Bonus ({nq} types)").font = Font(bold=True)
        ws.cell(row=row, column=4, value=bonus).number_format = "0.0%"
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

        ws.cell(row=row, column=1, value="Total P/Iron (with bonus)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=true_ratio).number_format = NUM_FMT_2
        ws.cell(row=row, column=4).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 2

    auto_width(ws, max_width=35)
    ws.freeze_panes = "A5"


def build_goods_demands(wb, land_units, categories, goods_demands):
    """Sheet: Resource requirements per buildable unit per age."""
    ws = wb.create_sheet("Goods Demands")

    ws.cell(row=1, column=1, value="Unit Goods Demands (Construction & Maintenance)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Scraped from goods_demand/army_demands.txt. "
                  "Quantities are per unit at max_strength.").font = SUBTITLE_FONT

    # Collect all goods that appear across all demands
    all_goods = set()
    for goods in goods_demands.values():
        all_goods.update(goods.keys())
    all_goods = sorted(all_goods)

    units = [
        u for u in land_units
        if u.get("buildable", True)
        and not u.get("levy", False)
        and not u["name"].startswith("a_age_")
    ]

    headers = ["Age", "Unit", "Category", "Light", "Demand Ref"] + all_goods

    row = 4
    for table_type, demand_field in [
        ("Construction", "construction_demand"),
        ("Maintenance", "maintenance_demand"),
    ]:
        ws.cell(row=row, column=1, value=f"{table_type} Demands").font = Font(bold=True, size=13)
        row += 1

        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        header_row = row
        row += 1

        for age in AGE_ORDER:
            age_units = [u for u in units if u.get("age") == age]
            if not age_units:
                continue

            for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
                cat_label = CAT_LABELS.get(u["category"], u["category"])
                demand_ref = u.get(demand_field, "") or categories.get(u["category"], {}).get(demand_field, "")
                goods = goods_demands.get(demand_ref, {})

                values = [
                    AGE_LABELS.get(u.get("age", ""), "?"),
                    loc(u["name"]),
                    cat_label,
                    "Yes" if u.get("light") else "",
                    demand_ref,
                ] + [goods.get(g, "") for g in all_goods]

                for j, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=j, value=v)
                    cell.border = THIN_BORDER
                    if isinstance(v, float):
                        cell.number_format = NUM_FMT_3
                    cat_fill = CAT_FILLS.get(cat_label)
                    if cat_fill:
                        cell.fill = cat_fill

                row += 1
            row += 1  # gap between ages

        row += 1  # gap between tables

    auto_width(ws, min_width=6, max_width=20)
    ws.freeze_panes = f"F{6}"


def build_goods_demands_generic(wb, land_units, categories, goods_demands, production_recipes):
    """Sheet: Goods demands for generic (non-special) infantry/cav/art/aux only."""
    ws = wb.create_sheet("Goods (Generic)")

    ws.cell(row=1, column=1, value="Generic Unit Goods Demands").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Standard upgrade-chain units only (no specials). "
                  "iron (total) = resolved from produced goods via workshop recipes.").font = SUBTITLE_FONT

    best_units = get_best_generic_units(land_units, categories)

    # Collect only goods that these units actually use
    relevant_goods = set()
    for age_types in best_units.values():
        for t in age_types:
            unit_name = t["unit_name"]
            if unit_name == "-":
                continue
            u = next((u for u in land_units if u["name"] == unit_name), None)
            if not u:
                continue
            for field in ["construction_demand", "maintenance_demand"]:
                ref = u.get(field, "") or categories.get(u["category"], {}).get(field, "")
                for g in goods_demands.get(ref, {}):
                    relevant_goods.add(g)
    relevant_goods = sorted(relevant_goods)

    headers = ["Age", "Type", "Unit", "Demand Ref"] + relevant_goods + ["iron (total)"]

    row = 4
    for table_type, demand_field in [
        ("Construction", "construction_demand"),
        ("Maintenance", "maintenance_demand"),
    ]:
        ws.cell(row=row, column=1, value=f"{table_type} Demands").font = Font(bold=True, size=13)
        row += 1

        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for age in AGE_ORDER:
            types = best_units[age]
            for t in types:
                unit_name = t["unit_name"]
                if unit_name == "-":
                    continue
                u = next((u for u in land_units if u["name"] == unit_name), None)
                if not u:
                    continue

                cat_label = CAT_LABELS.get(u["category"], u["category"])
                demand_ref = u.get(demand_field, "") or categories.get(u["category"], {}).get(demand_field, "")
                goods = goods_demands.get(demand_ref, {})

                # Resolve raw iron from all produced goods
                raw = resolve_raw_materials(goods, production_recipes) if goods else {}
                total_iron = round(raw.get("iron", 0), 4) or ""

                values = [
                    AGE_LABELS.get(age, "?"),
                    t["type_label"],
                    loc(unit_name),
                    demand_ref,
                ] + [goods.get(g, "") for g in relevant_goods] + [total_iron]

                for j, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=j, value=v)
                    cell.border = THIN_BORDER
                    if isinstance(v, float):
                        cell.number_format = NUM_FMT_3
                    cat_fill = CAT_FILLS.get(cat_label)
                    if cat_fill:
                        cell.fill = cat_fill

                row += 1
            row += 1  # gap between ages

        row += 1  # gap between tables

    auto_width(ws, min_width=6, max_width=20)
    ws.freeze_panes = "E6"


def pick_recipe(recipes, good, prefer_tier="workshop", prefer_variant="iron"):
    """Pick a single production recipe for a good.

    Prefers: workshop tier, iron-based variant. Falls back through tiers.
    Returns dict of { input_good: amount_per_unit_output } or None.
    """
    candidates = recipes.get(good, [])
    if not candidates:
        return None

    tier_order = [prefer_tier, "guild", "manufactory", "factory", "unknown"]

    for tier in tier_order:
        tier_recipes = [r for r in candidates if r["tier"] == tier]
        if not tier_recipes:
            continue
        # Prefer iron-based variant
        for r in tier_recipes:
            if prefer_variant in r["method"]:
                return {k: v / r["output"] for k, v in r["inputs"].items()}
        # Prefer livestock over wild_game for leather
        for r in tier_recipes:
            if "livestock" in r["method"] or "cotton" in r["method"]:
                return {k: v / r["output"] for k, v in r["inputs"].items()}
        # Fall back to first recipe without "ammunition" or "stone" or "obsidian"
        for r in tier_recipes:
            if not any(x in r["method"] for x in ["ammunition", "stone", "obsidian", "pre_columbian"]):
                return {k: v / r["output"] for k, v in r["inputs"].items()}
        # Last resort
        return {k: v / tier_recipes[0]["output"] for k, v in tier_recipes[0]["inputs"].items()}

    return None


def resolve_raw_materials(goods_needed, recipes, max_depth=3):
    """Resolve produced goods into raw materials recursively.

    goods_needed: { good_name: amount }
    Returns: { raw_material: total_amount }
    """
    raw = {}
    to_resolve = dict(goods_needed)

    for _ in range(max_depth):
        next_resolve = {}
        for good, amount in to_resolve.items():
            recipe = pick_recipe(recipes, good)
            if recipe is None:
                # It's a raw material (or unknown) — keep as-is
                raw[good] = raw.get(good, 0) + amount
            else:
                # Produced good — break down into inputs
                for input_good, input_amt in recipe.items():
                    needed = input_amt * amount
                    next_resolve[input_good] = next_resolve.get(input_good, 0) + needed
        if not next_resolve:
            break
        to_resolve = next_resolve

    # Any remaining unresolved go into raw
    for good, amount in to_resolve.items():
        raw[good] = raw.get(good, 0) + amount

    return raw


def build_raw_materials(wb, land_units, categories, goods_demands, production_recipes):
    """Sheet: Raw material breakdown for generic units per age."""
    ws = wb.create_sheet("Raw Materials")

    ws.cell(row=1, column=1, value="Raw Material Requirements (Generic Units)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Produced goods resolved to raw materials using workshop-tier iron-based recipes. "
                  "Recursive: tools->iron, etc.").font = SUBTITLE_FONT

    best_units = get_best_generic_units(land_units, categories)

    # First pass: collect all raw materials that appear
    all_raws = set()
    unit_raws = {}  # (age, type_label, demand_field) -> raw dict
    for age in AGE_ORDER:
        for t in best_units[age]:
            if t["unit_name"] == "-":
                continue
            u = next((u for u in land_units if u["name"] == t["unit_name"]), None)
            if not u:
                continue
            for demand_field in ["construction_demand", "maintenance_demand"]:
                ref = u.get(demand_field, "") or categories.get(u["category"], {}).get(demand_field, "")
                goods = goods_demands.get(ref, {})
                if goods:
                    raw = resolve_raw_materials(goods, production_recipes)
                    unit_raws[(age, t["type_label"], demand_field)] = raw
                    all_raws.update(raw.keys())

    all_raws = sorted(all_raws)
    headers = ["Age", "Type", "Unit"] + all_raws

    row = 4
    for table_type, demand_field in [
        ("Construction", "construction_demand"),
        ("Maintenance", "maintenance_demand"),
    ]:
        ws.cell(row=row, column=1, value=f"{table_type} - Raw Materials").font = Font(bold=True, size=13)
        row += 1

        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for age in AGE_ORDER:
            for t in best_units[age]:
                if t["unit_name"] == "-":
                    continue
                u = next((u for u in land_units if u["name"] == t["unit_name"]), None)
                if not u:
                    continue

                cat_label = CAT_LABELS.get(u["category"], u["category"])
                raw = unit_raws.get((age, t["type_label"], demand_field), {})

                values = [
                    AGE_LABELS.get(age, "?"),
                    t["type_label"],
                    loc(t["unit_name"]),
                ] + [round(raw.get(g, 0), 4) if raw.get(g, 0) else "" for g in all_raws]

                for j, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=j, value=v)
                    cell.border = THIN_BORDER
                    if isinstance(v, float):
                        cell.number_format = NUM_FMT_3
                    cat_fill = CAT_FILLS.get(cat_label)
                    if cat_fill:
                        cell.fill = cat_fill

                row += 1
            row += 1  # gap between ages

        row += 1  # gap between tables

    auto_width(ws, min_width=6, max_width=15)
    ws.freeze_panes = "D6"


NAVAL_CATS = ["navy_galley", "navy_light_ship", "navy_heavy_ship", "navy_transport"]


def _filter_naval(naval_units, generic_only=False):
    """Filter naval units to buildable, non-template ships."""
    return [
        u for u in naval_units
        if u.get("buildable", True)
        and not u.get("levy", False)
        and not u["name"].startswith("n_age_")
        and u["category"] in NAVAL_CATS
        and (not generic_only or not u.get("is_special", False))
    ]


def build_navy(wb, naval_units, categories):
    """Sheet: Naval units per age with combat power ranking."""
    ws = wb.create_sheet("Navy")

    ws.cell(row=1, column=1, value="Naval Units by Age").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="NavalPower = Cannons x HullSize. "
                  "Red = best overall in age, Blue = best in category.").font = Font(italic=True)

    headers = [
        "Age", "Category", "Unit",
        "Naval Power", "Power/Crew",
        "Cannons", "Hull Size", "Crew Size",
        "Movement", "Blockade", "Transport Cap",
        "Anti-Piracy", "Initiative", "Combat Speed",
        "Terrain", "Upgrades To",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    # Highlight: Naval Power=4, Power/Crew=5
    HIGHLIGHT_COLS = [4, 5]

    units = _filter_naval(naval_units, generic_only=True)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            crew = u.get("crew_size", cat_data.get("crew_size", 0)) or 0
            naval_power = cannons * hull
            power_per_crew = naval_power / crew if crew > 0 else 0

            # Terrain from unit, then fall back to category
            terrain = u.get("terrain_combat", {})
            if not terrain:
                terrain = cat_data.get("combat", {})
            terrain_str = ", ".join(
                f"{k}: {v:+.1f}" for k, v in terrain.items()
            ) if terrain else ""

            values = [
                AGE_LABELS.get(u.get("age", ""), "?"),
                cat_label,
                loc(u["name"]),
                naval_power,
                round(power_per_crew, 1),
                cannons, hull, crew,
                u.get("movement_speed", cat_data.get("movement_speed", 0)),
                u.get("blockade_capacity", cat_data.get("blockade_capacity", 0)),
                u.get("transport_capacity", cat_data.get("transport_capacity", 0)),
                u.get("anti_piracy_warfare", cat_data.get("anti_piracy_warfare", 0)),
                safe_num(u.get("initiative", cat_data.get("initiative", 0))),
                u.get("combat_speed", cat_data.get("combat_speed", 0)),
                terrain_str,
                loc(u.get("upgrades_to", "")),
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1  # gap between ages

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=40)
    ws.freeze_panes = f"A{header_row + 1}"


WATER_TYPES = ["deep_ocean", "ocean", "inland_sea", "narrows"]


def _all_naval_up_to_age(naval_units, age):
    """Get all buildable generic naval units available up to and including the given age."""
    age_idx = AGE_ORDER.index(age)
    valid_ages = set(AGE_ORDER[:age_idx + 1])
    return [
        u for u in naval_units
        if u.get("buildable", True)
        and not u.get("levy", False)
        and not u["name"].startswith("n_age_")
        and u["category"] in NAVAL_CATS
        and u.get("age", "") in valid_ages
    ]


def build_maritime_per_sailor(wb, naval_units, categories, prices):
    """Sheet: Maritime presence per sailor, considering all ages for each age."""
    ws = wb.create_sheet("Maritime (Sailor)")

    ws.cell(row=1, column=1, value="Maritime Presence per Sailor").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="All ships available up to each age. Best = most presence per sailor. "
                  "Crew = sailors per ship.").font = Font(italic=True)

    headers = [
        "Available At", "Category", "Unit", "Ship Age",
        "MP/Sailor", "Maritime Presence", "Crew (Sailors)",
        "Gold Cost",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [5]

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        units = _all_naval_up_to_age(naval_units, age)
        if not units:
            continue

        ws.cell(row=row, column=1, value=f"Available by {AGE_LABELS[age]}").font = Font(bold=True, size=12)
        row += 1

        for u in sorted(units, key=lambda x: (
            x.get("maritime_presence", 0) / x.get("crew_size", 999) if x.get("crew_size", 0) > 0 else 0
        ), reverse=True):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            mp = u.get("maritime_presence", 0) or 0
            crew = u.get("crew_size", cat_data.get("crew_size", 0)) or 0
            mp_per_sailor = mp / crew if crew > 0 else 0
            cost = calc_cost(u.get("max_strength", 1.0), cat, prices)

            values = [
                AGE_LABELS.get(age, "?"),
                cat_label, loc(u["name"]),
                AGE_LABELS.get(u.get("age", ""), "?"),
                mp_per_sailor, mp, crew,
                cost,
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {5: mp_per_sailor}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


def build_maritime_per_gold(wb, naval_units, categories, prices):
    """Sheet: Maritime presence per gold, considering all ages for each age."""
    ws = wb.create_sheet("Maritime (Gold)")

    ws.cell(row=1, column=1, value="Maritime Presence per Gold").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="All ships available up to each age. Best = most presence per gold spent.").font = Font(italic=True)

    headers = [
        "Available At", "Category", "Unit", "Ship Age",
        "MP/Gold", "Maritime Presence", "Gold Cost",
        "Crew (Sailors)",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [5]

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        units = _all_naval_up_to_age(naval_units, age)
        if not units:
            continue

        ws.cell(row=row, column=1, value=f"Available by {AGE_LABELS[age]}").font = Font(bold=True, size=12)
        row += 1

        for u in sorted(units, key=lambda x: (
            x.get("maritime_presence", 0) / calc_cost(x.get("max_strength", 1.0), x["category"], prices)
            if calc_cost(x.get("max_strength", 1.0), x["category"], prices) > 0 else 0
        ), reverse=True):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            mp = u.get("maritime_presence", 0) or 0
            crew = u.get("crew_size", cat_data.get("crew_size", 0)) or 0
            cost = calc_cost(u.get("max_strength", 1.0), cat, prices)
            mp_per_gold = mp / cost * 100 if cost > 0 else 0

            values = [
                AGE_LABELS.get(age, "?"),
                cat_label, loc(u["name"]),
                AGE_LABELS.get(u.get("age", ""), "?"),
                mp_per_gold, mp, cost,
                crew,
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {5: mp_per_gold}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


def build_navy_unique_terrain(wb, naval_units, categories):
    """Sheet: All naval units (incl. unique) with power by water type."""
    ws = wb.create_sheet("Navy (Unique Terrain)")

    ws.cell(row=1, column=1, value="All Naval Units - Power by Water Type").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Terrain Power = NavalPower * (1 + terrain_modifier). "
                  "Includes all unique/special units.").font = Font(italic=True)

    headers = (
        ["Age", "Category", "Unit", "Special", "Base Power"]
        + [f"Power ({t.replace('_', ' ').title()})" for t in WATER_TYPES]
        + ["Cannons", "Hull Size"]
    )
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [6, 7, 8, 9]

    units = _filter_naval(naval_units, generic_only=False)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            naval_power = cannons * hull

            terrain = u.get("terrain_combat", {})
            if not terrain:
                terrain = cat_data.get("combat", {})

            terrain_powers = []
            for wt in WATER_TYPES:
                mod = terrain.get(wt, 0)
                terrain_powers.append(round(naval_power * (1 + mod), 1))

            values = (
                [AGE_LABELS.get(u.get("age", ""), "?"),
                 cat_label, loc(u["name"]),
                 "Yes" if u.get("is_special") else "",
                 naval_power]
                + terrain_powers
                + [cannons, hull]
            )
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


def build_navy_unique(wb, naval_units, categories):
    """Sheet: Unique/special naval units with power ranking."""
    ws = wb.create_sheet("Navy (Unique)")

    ws.cell(row=1, column=1, value="All Naval Units by Age (incl. Unique)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Red = best overall in age, Blue = best in category.").font = Font(italic=True)

    headers = [
        "Age", "Category", "Unit", "Special",
        "Naval Power", "Power/Crew",
        "Cannons", "Hull Size", "Crew Size",
        "Movement", "Blockade", "Transport Cap",
        "Anti-Piracy", "Initiative", "Combat Speed",
        "Terrain", "Upgrades To",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [5, 6]

    units = _filter_naval(naval_units, generic_only=False)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            crew = u.get("crew_size", cat_data.get("crew_size", 0)) or 0
            naval_power = cannons * hull
            power_per_crew = naval_power / crew if crew > 0 else 0

            terrain = u.get("terrain_combat", {})
            if not terrain:
                terrain = cat_data.get("combat", {})
            terrain_str = ", ".join(
                f"{k}: {v:+.1f}" for k, v in terrain.items()
            ) if terrain else ""

            values = [
                AGE_LABELS.get(u.get("age", ""), "?"),
                cat_label,
                loc(u["name"]),
                "Yes" if u.get("is_special") else "",
                naval_power,
                round(power_per_crew, 1),
                cannons, hull, crew,
                u.get("movement_speed", cat_data.get("movement_speed", 0)),
                u.get("blockade_capacity", cat_data.get("blockade_capacity", 0)),
                u.get("transport_capacity", cat_data.get("transport_capacity", 0)),
                u.get("anti_piracy_warfare", cat_data.get("anti_piracy_warfare", 0)),
                safe_num(u.get("initiative", cat_data.get("initiative", 0))),
                u.get("combat_speed", cat_data.get("combat_speed", 0)),
                terrain_str,
                loc(u.get("upgrades_to", "")),
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=40)
    ws.freeze_panes = f"A{header_row + 1}"


def build_navy_gold(wb, naval_units, categories, prices):
    """Sheet: Naval units ranked by power per gold."""
    ws = wb.create_sheet("Navy (Gold)")

    ws.cell(row=1, column=1, value="Naval Units - Power per Gold").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Power/Gold = NavalPower / BuildCost. "
                  "Red = best overall, Blue = best in category.").font = Font(italic=True)

    headers = [
        "Age", "Category", "Unit",
        "Power/Gold", "Naval Power", "Build Cost",
        "Cannons", "Hull Size", "Crew Size",
        "Movement", "Blockade", "Transport Cap",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [4]

    units = _filter_naval(naval_units, generic_only=True)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            crew = u.get("crew_size", cat_data.get("crew_size", 0)) or 0
            naval_power = cannons * hull
            # Naval max_strength is always 1.0, so cost = base * 10
            build_cost = calc_cost(u.get("max_strength", 1.0), cat, prices)
            power_gold = naval_power / build_cost * 100 if build_cost > 0 else 0

            values = [
                AGE_LABELS.get(u.get("age", ""), "?"),
                cat_label,
                loc(u["name"]),
                round(power_gold, 2),
                naval_power, build_cost,
                cannons, hull, crew,
                u.get("movement_speed", cat_data.get("movement_speed", 0)),
                u.get("blockade_capacity", cat_data.get("blockade_capacity", 0)),
                u.get("transport_capacity", cat_data.get("transport_capacity", 0)),
            ]
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


def build_navy_unique_terrain_gold(wb, naval_units, categories, prices):
    """Sheet: All naval units (incl. unique) with power/gold by water type."""
    ws = wb.create_sheet("Navy (Uniq Terrain-Gold)")

    ws.cell(row=1, column=1, value="All Naval Units - Power per Gold by Water Type").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Terrain P/Gold = NavalPower * (1 + terrain_mod) / BuildCost * 100. "
                  "Includes all unique/special units.").font = Font(italic=True)

    headers = (
        ["Age", "Category", "Unit", "Special", "Build Cost"]
        + [f"P/Gold ({t.replace('_', ' ').title()})" for t in WATER_TYPES]
        + ["Cannons", "Hull Size"]
    )
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [6, 7, 8, 9]

    units = _filter_naval(naval_units, generic_only=False)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            naval_power = cannons * hull
            build_cost = calc_cost(u.get("max_strength", 1.0), cat, prices)

            terrain = u.get("terrain_combat", {})
            if not terrain:
                terrain = cat_data.get("combat", {})

            terrain_pg = []
            for wt in WATER_TYPES:
                mod = terrain.get(wt, 0)
                tp = naval_power * (1 + mod) / build_cost * 100 if build_cost > 0 else 0
                terrain_pg.append(round(tp, 2))

            values = (
                [AGE_LABELS.get(u.get("age", ""), "?"),
                 cat_label, loc(u["name"]),
                 "Yes" if u.get("is_special") else "",
                 build_cost]
                + terrain_pg
                + [cannons, hull]
            )
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


def build_navy_terrain(wb, naval_units, categories):
    """Sheet: Naval power adjusted per water type."""
    ws = wb.create_sheet("Navy (Terrain)")

    ws.cell(row=1, column=1, value="Naval Units - Power by Water Type").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Terrain Power = NavalPower * (1 + terrain_modifier). "
                  "Galleys get bonuses in shallow water, penalties in deep ocean.").font = Font(italic=True)

    headers = (
        ["Age", "Category", "Unit", "Base Power"]
        + [f"Power ({t.replace('_', ' ').title()})" for t in WATER_TYPES]
        + ["Cannons", "Hull Size"]
    )
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    # Highlight per-terrain power columns: 5, 6, 7, 8
    HIGHLIGHT_COLS = [5, 6, 7, 8]

    units = _filter_naval(naval_units, generic_only=True)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            naval_power = cannons * hull

            # Get terrain modifiers (unit level, then category fallback)
            terrain = u.get("terrain_combat", {})
            if not terrain:
                terrain = cat_data.get("combat", {})

            terrain_powers = []
            for wt in WATER_TYPES:
                mod = terrain.get(wt, 0)
                terrain_powers.append(round(naval_power * (1 + mod), 1))

            values = (
                [AGE_LABELS.get(u.get("age", ""), "?"),
                 cat_label, loc(u["name"]), naval_power]
                + terrain_powers
                + [cannons, hull]
            )
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


def build_navy_terrain_gold(wb, naval_units, categories, prices):
    """Sheet: Naval power per gold adjusted per water type."""
    ws = wb.create_sheet("Navy (Terrain-Gold)")

    ws.cell(row=1, column=1, value="Naval Units - Power per Gold by Water Type").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Terrain P/Gold = NavalPower * (1 + terrain_mod) / BuildCost * 100.").font = Font(italic=True)

    headers = (
        ["Age", "Category", "Unit", "Build Cost"]
        + [f"P/Gold ({t.replace('_', ' ').title()})" for t in WATER_TYPES]
        + ["Cannons", "Hull Size"]
    )
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    HIGHLIGHT_COLS = [5, 6, 7, 8]

    units = _filter_naval(naval_units, generic_only=True)

    tracked_rows = []
    row = header_row + 1
    for age in AGE_ORDER:
        age_units = [u for u in units if u.get("age") == age]
        if not age_units:
            continue

        for u in sorted(age_units, key=lambda x: (x["category"], x["name"])):
            cat = u["category"]
            cat_data = categories.get(cat, {})
            cat_label = CAT_LABELS.get(cat, cat)

            cannons = u.get("cannons", cat_data.get("cannons", 0)) or 0
            hull = u.get("hull_size", cat_data.get("hull_size", 0)) or 0
            naval_power = cannons * hull
            build_cost = calc_cost(u.get("max_strength", 1.0), cat, prices)

            terrain = u.get("terrain_combat", {})
            if not terrain:
                terrain = cat_data.get("combat", {})

            terrain_pg = []
            for wt in WATER_TYPES:
                mod = terrain.get(wt, 0)
                tp = naval_power * (1 + mod) / build_cost * 100 if build_cost > 0 else 0
                terrain_pg.append(round(tp, 2))

            values = (
                [AGE_LABELS.get(u.get("age", ""), "?"),
                 cat_label, loc(u["name"]), build_cost]
                + terrain_pg
                + [cannons, hull]
            )
            for j, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_2
                cat_fill = CAT_FILLS.get(cat_label)
                if cat_fill:
                    cell.fill = cat_fill

            tracked_rows.append((row, age, cat_label, {c: values[c - 1] for c in HIGHLIGHT_COLS}))
            row += 1

        row += 1

    highlight_best_in_age_by_cat(ws, tracked_rows, HIGHLIGHT_COLS)
    auto_width(ws, max_width=35)
    ws.freeze_panes = f"A{header_row + 1}"


# ---------------------------------------------------------------------------
#  Food analysis helpers
# ---------------------------------------------------------------------------

FOOD_BUILDING_ORDER = [
    "farming_village", "fishing_village", "forest_village", "fruit_orchard",
    "sheep_farms", "windmill", "irrigation_systems", "pound_lock_canal_infrastructure",
    "market_village", "elephant_hunting_grounds",
]

FOOD_BUILDING_LABELS = {
    "farming_village": "Farming Village",
    "fishing_village": "Fishing Village",
    "forest_village": "Forest Village",
    "fruit_orchard": "Fruit Orchard",
    "sheep_farms": "Sheep Farms",
    "windmill": "Windmill",
    "irrigation_systems": "Irrigation",
    "pound_lock_canal_infrastructure": "Canal",
    "market_village": "Market Village",
    "elephant_hunting_grounds": "Elephant Grounds",
}

VEG_ORDER = ["farmland", "grasslands", "woods", "jungle", "forest", "sparse", "desert"]
TOPO_ORDER = ["flatland", "hills", "plateau", "wetlands", "mountains"]


def calc_building_cap(cap_name, caps, dev, rgo_workers, has_river):
    """Compute max building levels from a cap formula."""
    cap = caps.get(cap_name, {})
    if not cap:
        return 0
    levels = cap.get("base", 0)
    levels += cap.get("per_development", 0) * dev
    levels += cap.get("per_max_rgo_workers", 0) * rgo_workers
    if has_river:
        levels += cap.get("if_river", 0)
    return int(levels)


def resolve_max_levels(bld, caps, dev, rgo_workers, has_river):
    """Resolve a building's max_levels to an integer."""
    ml = bld.get("max_levels", 1)
    if isinstance(ml, int):
        return ml
    if isinstance(ml, str):
        return calc_building_cap(ml, caps, dev, rgo_workers, has_river)
    # Complex expression (e.g., elephant_hunting_grounds with conditional)
    if isinstance(ml, dict):
        base_ref = ml.get("value", "")
        if isinstance(base_ref, str):
            return calc_building_cap(base_ref, caps, dev, rgo_workers, has_river)
    return 1


def building_available(bld, rgo, vegetation, has_river, is_coastal):
    """Check if a building is available for a given location profile."""
    reqs = bld.get("requirements", {})

    # RGO check: if building requires specific RGOs, location must have one
    req_rgos = reqs.get("rgo", [])
    if req_rgos and rgo not in req_rgos:
        return False

    # Vegetation check: if building requires specific vegetation, location must match
    req_veg = reqs.get("vegetation", [])
    if req_veg and vegetation not in req_veg:
        return False

    # Feature checks: OR logic — if building needs any of these features, at least one must match
    req_features = reqs.get("features", [])
    if req_features:
        has_any = False
        for feat in req_features:
            if feat == "is_coastal" and is_coastal:
                has_any = True
            elif feat == "has_river" and has_river:
                has_any = True
            elif feat == "is_adjacent_to_lake":
                # Lakes are rare; treat as equivalent to river for analysis
                has_any = has_any or has_river
        if not has_any:
            return False

    return True


def calc_building_food(bld, food_goods):
    """Calculate food contribution per level for a building.

    Returns (goods_food_per_level, food_modifier_per_level, flat_food_per_level, rgo_output_mod_note).
    """
    goods_food = 0.0
    produces = bld.get("produces")
    if produces:
        good = produces["good"]
        output = produces["output_per_level"]
        food_val = food_goods.get(good, {}).get("food_value", 0)
        goods_food = output * food_val

    food_mod = bld.get("food_modifiers", {}).get("local_monthly_food_modifier", 0)
    flat_food = bld.get("food_modifiers", {}).get("local_monthly_food", 0)

    rgo_notes = ""
    rgo_mods = bld.get("rgo_output_modifiers", {})
    if rgo_mods:
        parts = []
        for good, mod in rgo_mods.items():
            parts.append(f"+{mod:.0%} {good}")
        rgo_notes = ", ".join(parts)

    return goods_food, food_mod, flat_food, rgo_notes


def build_food_reference(wb, food_goods, food_buildings, caps, terrain_mods):
    """Reference sheet listing all food goods and food buildings."""
    ws = wb.create_sheet("Food Reference")

    # --- Food Goods section ---
    ws.cell(row=1, column=1, value="Food Goods Reference").font = TITLE_FONT

    goods_headers = ["Good", "Food Value", "Method", "Price"]
    header_row = 3
    for i, h in enumerate(goods_headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(goods_headers))

    sorted_goods = sorted(food_goods.items(), key=lambda x: -x[1]["food_value"])
    for r, (name, data) in enumerate(sorted_goods, header_row + 1):
        values = [name, data["food_value"], data["method"], data["price"]]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = THIN_BORDER
            if isinstance(v, float):
                cell.number_format = NUM_FMT_2

    # --- Food Buildings section ---
    bld_start = header_row + len(sorted_goods) + 3
    ws.cell(row=bld_start, column=1, value="Food Buildings Reference").font = TITLE_FONT

    bld_headers = [
        "Building", "Max Levels", "RGO Req", "Vegetation Req", "Feature Req",
        "Produces", "Output/Lvl", "Food/Lvl (goods)", "Food Mod%/Lvl", "Flat Food/Lvl",
        "RGO Output Mod", "Food Capacity/Lvl",
    ]
    bld_hdr_row = bld_start + 1
    for i, h in enumerate(bld_headers, 1):
        ws.cell(row=bld_hdr_row, column=i, value=h)
    style_header_row(ws, bld_hdr_row, len(bld_headers))

    row = bld_hdr_row + 1
    for bld_name in FOOD_BUILDING_ORDER:
        bld = food_buildings.get(bld_name)
        if not bld:
            continue
        reqs = bld.get("requirements", {})
        goods_food, food_mod, flat_food, rgo_note = calc_building_food(bld, food_goods)
        produces = bld.get("produces")
        ml = bld.get("max_levels", 1)
        ml_display = ml if isinstance(ml, int) else str(ml) if isinstance(ml, str) else "varies"

        values = [
            FOOD_BUILDING_LABELS.get(bld_name, bld_name),
            ml_display,
            ", ".join(reqs.get("rgo", [])) or "-",
            ", ".join(reqs.get("vegetation", [])) or "-",
            ", ".join(reqs.get("features", [])) or "-",
            produces["good"] if produces else "-",
            produces["output_per_level"] if produces else 0,
            goods_food,
            food_mod,
            flat_food,
            rgo_note or "-",
            bld.get("food_modifiers", {}).get("local_food_capacity", 0),
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
            if isinstance(v, float):
                cell.number_format = NUM_FMT_3 if abs(v) < 1 else NUM_FMT_2
        row += 1

    # --- Terrain Modifiers section ---
    terr_start = row + 2
    ws.cell(row=terr_start, column=1, value="Terrain Food Modifiers").font = TITLE_FONT

    terr_headers = ["Type", "Category", "Food Modifier"]
    terr_hdr_row = terr_start + 1
    for i, h in enumerate(terr_headers, 1):
        ws.cell(row=terr_hdr_row, column=i, value=h)
    style_header_row(ws, terr_hdr_row, len(terr_headers))

    row = terr_hdr_row + 1
    for category, entries in terrain_mods.items():
        for name, data in entries.items():
            mod = data.get("local_monthly_food_modifier", 0)
            values = [name, category, mod]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.border = THIN_BORDER
                if isinstance(v, float):
                    cell.number_format = "+0%;-0%;0%"
            row += 1

    # --- Building Caps section ---
    cap_start = row + 2
    ws.cell(row=cap_start, column=1, value="Building Cap Formulas").font = TITLE_FONT

    cap_headers = ["Cap Name", "Base", "Per Dev", "Per RGO Workers", "If River"]
    cap_hdr_row = cap_start + 1
    for i, h in enumerate(cap_headers, 1):
        ws.cell(row=cap_hdr_row, column=i, value=h)
    style_header_row(ws, cap_hdr_row, len(cap_headers))

    row = cap_hdr_row + 1
    for cap_name, cap_data in caps.items():
        values = [
            cap_name,
            cap_data.get("base", 0),
            cap_data.get("per_development", 0),
            cap_data.get("per_max_rgo_workers", 0),
            cap_data.get("if_river", 0),
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
            if isinstance(v, float):
                cell.number_format = NUM_FMT_2
        row += 1

    auto_width(ws, max_width=40)


def _calc_crossover(rgo_food_val, goods_f, mod_f, flat_f, rgo_note):
    """Calculate the RGO level at which a building becomes better than another RGO level.

    At N RGO levels producing N * rgo_food_val total RGO food:
    - Another RGO level gives: +rgo_food_val
    - This building gives: goods_f + flat_f + mod_f * total_base_food
    - Crossover: goods_f + flat_f + mod_f * N * rgo_food_val >= rgo_food_val

    For windmill (+X% RGO output): 0.1 * N * rgo_food_val >= rgo_food_val → N >= 10

    Returns crossover N (int), or None if building is always better or never better via modifier.
    """
    constant = goods_f + flat_f

    if mod_f > 0:
        # goods_f + flat_f + mod_f * N * rgo_food_val >= rgo_food_val
        # mod_f * N * rgo_food_val >= rgo_food_val - constant
        shortfall = rgo_food_val - constant
        if shortfall <= 0:
            return 0  # always better than RGO (constant food alone beats it)
        if rgo_food_val == 0:
            return None
        n = shortfall / (mod_f * rgo_food_val)
        return int(n) + (1 if n != int(n) else 0)  # ceiling

    if rgo_note:
        # Windmill: extracts the modifier from rgo_note (10% = 0.1 per good)
        # Simplified: assume +10% RGO output → 0.1 * N * rgo_food_val >= rgo_food_val → N >= 10
        return 10

    if constant >= rgo_food_val:
        return 0  # always better
    return None  # never better via modifier (pure constant food, less than RGO)


def _rank_buildings_for_profile(available, food_buildings, food_goods, rgo_good, rgo_workers):
    """Rank available buildings by build priority for a location profile.

    Includes the base RGO and crossover points.

    Returns list of (label, max_levels, food_display, crossover) sorted by initial priority.
    crossover = RGO level at which this building becomes better than another RGO level.
    """
    rgo_food_val = food_goods.get(rgo_good, {}).get("food_value", 0)
    entries = []

    # Add the base RGO
    if rgo_food_val > 0:
        entries.append({
            "name": "_rgo",
            "label": f"RGO ({rgo_good})",
            "levels": rgo_workers,
            "goods_f": rgo_food_val,
            "mod_f": 0,
            "flat_f": 0,
            "rgo_note": "",
            "crossover": None,  # RGO is the baseline
            "sort_key": (0, -rgo_food_val),
        })

    for bld_name, levels in available.items():
        bld = food_buildings[bld_name]
        label = FOOD_BUILDING_LABELS.get(bld_name, bld_name)
        goods_f, mod_f, flat_f, rgo_note = calc_building_food(bld, food_goods)
        constant = goods_f + flat_f

        if constant == 0 and mod_f == 0 and not rgo_note:
            continue
        if mod_f < 0 and constant == 0:
            continue

        crossover = _calc_crossover(rgo_food_val, goods_f, mod_f, flat_f, rgo_note)

        # Sort: constant food producers first (by food desc), then modifier buildings
        if rgo_note and constant == 0 and mod_f == 0:
            sort_key = (2, 0)
        elif constant > 0:
            sort_key = (1, -constant)
        else:
            sort_key = (3, -mod_f)

        entries.append({
            "name": bld_name,
            "label": label,
            "levels": levels,
            "goods_f": goods_f,
            "mod_f": mod_f,
            "flat_f": flat_f,
            "rgo_note": rgo_note,
            "crossover": crossover,
            "sort_key": sort_key,
        })

    # Sort by crossover point: RGO first (baseline), then by when each building
    # becomes worth building (lower crossover = build sooner).
    # "always" (0) before numeric, numeric ascending, "never" (None) last.
    def sort_key(e):
        if e["name"] == "_rgo":
            return (-1, 0)  # always first
        xover = e["crossover"]
        if xover is None:
            return (10000, 0)  # never beats RGO — last
        return (xover, -e["goods_f"] - e["flat_f"])

    entries.sort(key=sort_key)
    return entries


def build_food_location_buildup(wb, food_goods, food_buildings, caps, terrain_mods):
    """Main sheet: build order / ratio table per location profile."""
    ws = wb.create_sheet("Food Build Order")

    ws.cell(row=1, column=1, value="Food Production Build Order by Location").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="'Build After' = the RGO level at which this building gives more food than another RGO level. "
                  "Topo adjusts Terrain Mod%: hills -10%, wetlands -10%, mountains -20%.").font = SUBTITLE_FONT

    DEV = 10
    RGO_WORKERS = 3

    relevant_buildings = [b for b in FOOD_BUILDING_ORDER if b in food_buildings]

    MAX_SLOTS = 7  # RGO + up to 6 buildings

    fixed_headers = ["RGO", "Food Val", "Vegetation", "River", "Coastal", "Terrain\nMod%"]
    slot_headers = []
    for i in range(1, MAX_SLOTS + 1):
        slot_headers.extend([f"#{i} Building", f"#{i} Build After\n(RGO lvl)", f"#{i} Ratio\nvs RGO", f"#{i} Food/Lvl"])
    tail_headers = ["Build Order"]

    all_headers = fixed_headers + slot_headers + tail_headers
    header_row = 4
    for i, h in enumerate(all_headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(all_headers))

    food_rgo_goods = sorted(
        [(name, data["food_value"]) for name, data in food_goods.items()],
        key=lambda x: -x[1]
    )

    VEG_FILLS = {
        "farmland": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "grasslands": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "woods": PatternFill(start_color="E8E0D0", end_color="E8E0D0", fill_type="solid"),
        "jungle": PatternFill(start_color="D5F5D5", end_color="D5F5D5", fill_type="solid"),
        "forest": PatternFill(start_color="C5D9C5", end_color="C5D9C5", fill_type="solid"),
        "sparse": PatternFill(start_color="F5F0E0", end_color="F5F0E0", fill_type="solid"),
        "desert": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }

    row = header_row + 1

    for rgo, rgo_food_val in food_rgo_goods:
        for veg in VEG_ORDER:
            for has_river in [True, False]:
                for is_coastal in [True, False]:
                    veg_mod = terrain_mods.get("vegetation", {}).get(veg, {}).get("local_monthly_food_modifier", 0)
                    rank_mod = terrain_mods.get("location_ranks", {}).get("rural_settlement", {}).get("local_monthly_food_modifier", 0)
                    terrain_total_mod = veg_mod + rank_mod

                    available = {}
                    for bld_name in relevant_buildings:
                        bld = food_buildings[bld_name]
                        if building_available(bld, rgo, veg, has_river, is_coastal):
                            levels = resolve_max_levels(bld, caps, DEV, RGO_WORKERS, has_river)
                            if levels > 0:
                                available[bld_name] = levels

                    if not available:
                        continue

                    ranked = _rank_buildings_for_profile(available, food_buildings, food_goods, rgo, RGO_WORKERS)
                    if not ranked:
                        continue

                    fixed_vals = [
                        rgo, rgo_food_val, veg,
                        "Yes" if has_river else "No",
                        "Yes" if is_coastal else "No",
                        terrain_total_mod,
                    ]

                    slot_vals = []
                    order_parts = []
                    for i in range(MAX_SLOTS):
                        if i < len(ranked):
                            e = ranked[i]
                            constant = e["goods_f"] + e["flat_f"]

                            # Food/level display
                            if e["rgo_note"]:
                                food_display = e["rgo_note"]
                            elif constant > 0 and e["mod_f"] > 0:
                                food_display = f"{constant:.1f} + {e['mod_f']:.0%}/lvl"
                            elif constant > 0:
                                food_display = f"{constant:.1f}"
                            elif e["mod_f"] > 0:
                                food_display = f"{e['mod_f']:.0%}/lvl"
                            else:
                                food_display = f"{constant:.1f}"

                            # Crossover display
                            xover = e["crossover"]
                            if e["name"] == "_rgo":
                                xover_display = "baseline"
                            elif xover is None:
                                xover_display = "never"
                            elif xover == 0:
                                xover_display = "always"
                            else:
                                xover_display = xover

                            # Ratio vs RGO:
                            # - RGO itself: baseline
                            # - Modifier buildings (food_mod > 0): 1:1 after threshold
                            #   (math: optimal M = N - threshold, so each RGO level past
                            #   threshold pairs with 1 building level)
                            # - One-time buildings (max 1 level): build once
                            # - Constant-food-only (no modifier): after RGO maxed
                            if e["name"] == "_rgo":
                                ratio_display = "-"
                            elif e["mod_f"] > 0 and e["levels"] > 1:
                                ratio_display = "1:1"
                            elif e["levels"] == 1:
                                ratio_display = "build once"
                            elif e["rgo_note"]:
                                ratio_display = "build once"
                            else:
                                ratio_display = "after max RGO"

                            slot_vals.extend([e["label"], xover_display, ratio_display, food_display])

                            # Build order string with crossover info
                            lvl_str = f"x{e['levels']}" if e["name"] != "_rgo" else "scale"
                            if e["name"] == "_rgo":
                                order_parts.append(f"{e['label']}")
                            elif xover is not None and xover > 0:
                                order_parts.append(f"{e['label']} ({lvl_str}, @{xover} RGO)")
                            else:
                                order_parts.append(f"{e['label']} ({lvl_str})")
                        else:
                            slot_vals.extend(["", "", "", ""])

                    build_order_str = " > ".join(order_parts)

                    all_vals = fixed_vals + slot_vals + [build_order_str]

                    for i, v in enumerate(all_vals, 1):
                        cell = ws.cell(row=row, column=i, value=v)
                        cell.border = THIN_BORDER
                        if isinstance(v, float):
                            if abs(v) < 1 and v != 0:
                                cell.number_format = "+0%;-0%;0%"
                            else:
                                cell.number_format = NUM_FMT_2

                    veg_fill = VEG_FILLS.get(veg)
                    if veg_fill:
                        for c in range(1, len(fixed_headers) + 1):
                            ws.cell(row=row, column=c).fill = veg_fill

                    row += 1

    auto_width(ws, max_width=25)
    ws.freeze_panes = f"A{header_row + 1}"


def build_food_build_order_by_rgo(wb, food_goods, food_buildings):
    """Simplified build order: one row per RGO, all buildings ranked.

    Terrain only affects availability, not crossover math, so this sheet
    lists every possible food building per RGO in priority order.
    Skip buildings not available at your location.
    """
    ws = wb.create_sheet("Build Order by RGO")

    ws.cell(row=1, column=1, value="Food Build Order by RGO").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Per-building in isolation (ignores modifier interactions). "
                  "'Build After' = the RGO level at which this building beats another RGO level. "
                  "See 'Compound Build Order' sheet for accurate ratios when multiple modifier buildings interact. "
                  "Skip buildings unavailable at your location.").font = SUBTITLE_FONT

    # Exclude market_village (no food) and elephant_hunting_grounds (negative)
    bld_names = [b for b in FOOD_BUILDING_ORDER
                 if b in food_buildings and b not in ("market_village", "elephant_hunting_grounds")]

    MAX_SLOTS = len(bld_names)

    fixed_headers = ["RGO", "RGO\nFood/Lvl"]
    slot_headers = []
    for i in range(1, MAX_SLOTS + 1):
        slot_headers.extend([
            f"#{i} Building",
            f"#{i} Build After\n(RGO lvl)",
            f"#{i} Food/Lvl",
            f"#{i} Requires",
        ])

    all_headers = fixed_headers + slot_headers
    header_row = 4
    for i, h in enumerate(all_headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(all_headers))

    sorted_goods = sorted(food_goods.items(), key=lambda x: -x[1]["food_value"])

    row = header_row + 1
    for rgo_name, rgo_data in sorted_goods:
        rgo_fv = rgo_data["food_value"]

        entries = []
        for bld_name in bld_names:
            bld = food_buildings[bld_name]

            # Filter out buildings that require a different RGO
            req_rgos = bld.get("requirements", {}).get("rgo", [])
            if req_rgos and rgo_name not in req_rgos:
                continue

            label = FOOD_BUILDING_LABELS.get(bld_name, bld_name)
            goods_f, mod_f, flat_f, rgo_note = calc_building_food(bld, food_goods)
            constant = goods_f + flat_f

            xover = _calc_crossover(rgo_fv, goods_f, mod_f, flat_f, rgo_note)

            # Food display
            if rgo_note:
                food_display = rgo_note
            elif constant > 0 and mod_f > 0:
                food_display = f"{constant:.1f} + {mod_f:.0%}/lvl"
            elif constant > 0:
                food_display = f"{constant:.1f}"
            elif mod_f > 0:
                food_display = f"{mod_f:.0%}/lvl"
            else:
                food_display = f"{constant:.1f}"

            # Build After display
            if xover is None:
                build_after = "after max RGO"
            elif xover == 0:
                build_after = "always"
            else:
                build_after = xover

            # Requirements summary
            reqs = bld.get("requirements", {})
            req_parts = []
            if reqs.get("rgo"):
                req_parts.append("RGO: " + ", ".join(reqs["rgo"]))
            if reqs.get("vegetation"):
                req_parts.append("veg: " + ", ".join(reqs["vegetation"]))
            if reqs.get("features"):
                req_parts.append(", ".join(reqs["features"]))
            requires = "; ".join(req_parts) if req_parts else "any"

            # Sort key
            if xover is None:
                sort_val = 10000
            elif xover == 0:
                sort_val = -1
            else:
                sort_val = xover

            entries.append({
                "label": label,
                "build_after": build_after,
                "food_display": food_display,
                "requires": requires,
                "crossover_sort": (sort_val, -constant),
            })

        entries.sort(key=lambda x: x["crossover_sort"])

        # Write row
        ws.cell(row=row, column=1, value=rgo_name).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=rgo_fv).border = THIN_BORDER

        col = 3
        for e in entries:
            ws.cell(row=row, column=col, value=e["label"]).border = THIN_BORDER
            ws.cell(row=row, column=col + 1, value=e["build_after"]).border = THIN_BORDER
            ws.cell(row=row, column=col + 2, value=e["food_display"]).border = THIN_BORDER
            ws.cell(row=row, column=col + 3, value=e["requires"]).border = THIN_BORDER
            col += 4

        row += 1

    auto_width(ws, max_width=30)
    ws.freeze_panes = f"A{header_row + 1}"


def _simulate_compound_build_order(rgo_fv, buildings, max_rgo=40):
    """Simulate greedy build order accounting for modifier interactions.

    At each step, build whichever option (RGO level or building level) gives
    the highest marginal food increase.

    Total food = (N*fv*(1+rgo_mod) + sum(goods_food*lvls)) * (1 + sum(food_mod*lvls))

    Returns (first_built, phases):
      first_built: {building_name: rgo_level_when_first_built}
      phases: list of (description_string) showing the build pattern
    """
    n_rgo = 0
    rgo_mod = 0.0
    lvls = {b["name"]: 0 for b in buildings}
    first_built = {}
    labels = {b["name"]: FOOD_BUILDING_LABELS.get(b["name"], b["name"]) for b in buildings}

    # Track sequence for phase extraction
    sequence = []  # list of ("_rgo" or building_name) per step

    for _ in range(300):
        base_food = n_rgo * rgo_fv * (1 + rgo_mod)
        for b in buildings:
            base_food += lvls[b["name"]] * b["goods_food"]

        total_mod = sum(lvls[b["name"]] * b["food_mod"] for b in buildings)

        best_val = -1
        best_name = None

        if n_rgo < max_rgo:
            m = rgo_fv * (1 + rgo_mod) * (1 + total_mod)
            if m > best_val:
                best_val = m
                best_name = "_rgo"

        for b in buildings:
            if lvls[b["name"]] >= b["max_levels"]:
                continue
            if b["rgo_output_mod"] > 0 and lvls[b["name"]] == 0:
                m = b["rgo_output_mod"] * n_rgo * rgo_fv * (1 + total_mod)
            else:
                m = b["goods_food"] * (1 + total_mod) + b["food_mod"] * base_food
            if m > best_val:
                best_val = m
                best_name = b["name"]

        if best_name is None:
            break

        if best_name == "_rgo":
            n_rgo += 1
        else:
            lvls[best_name] += 1
            b_data = next(b for b in buildings if b["name"] == best_name)
            if b_data["rgo_output_mod"] > 0:
                rgo_mod += b_data["rgo_output_mod"]
            if best_name not in first_built:
                first_built[best_name] = n_rgo

        sequence.append((best_name, n_rgo, dict(lvls)))

        if n_rgo >= max_rgo and all(lvls[b["name"]] >= b["max_levels"] for b in buildings):
            break

    # Extract phases using cumulative totals for clarity
    # First, build raw phases: each entry is (action, rgo_level, building_totals)
    raw = []
    i = 0
    while i < len(sequence):
        name = sequence[i][0]
        if name == "_rgo":
            j = i
            while j < len(sequence) and sequence[j][0] == "_rgo":
                j += 1
            raw.append(("_rgo", sequence[j - 1][1], dict(sequence[j - 1][2])))
            i = j
        else:
            j = i
            while j < len(sequence) and sequence[j][0] == name:
                j += 1
            total = sequence[j - 1][2][name]
            raw.append((name, sequence[j - 1][1], total))
            i = j

    # Detect alternating RGO/building patterns and compress
    phases = []
    i = 0
    while i < len(raw):
        entry = raw[i]

        if entry[0] == "_rgo":
            # Look ahead for alternating RGO > Building > RGO > Building
            if i + 1 < len(raw) and raw[i + 1][0] != "_rgo":
                bld_name = raw[i + 1][0]
                j = i
                alt_count = 0
                while (j + 1 < len(raw)
                       and raw[j][0] == "_rgo"
                       and raw[j + 1][0] == bld_name):
                    # Check it's single-level increments
                    prev_total = raw[j - 1][2] if j > 0 and raw[j - 1][0] == bld_name else 0
                    if isinstance(raw[j + 1][2], int) and raw[j + 1][2] - (prev_total if isinstance(prev_total, int) else 0) == 1:
                        alt_count += 1
                        j += 2
                    else:
                        break

                if alt_count >= 3:
                    start_rgo = entry[1]
                    end_rgo = raw[j - 2][1]
                    label = labels.get(bld_name, bld_name)
                    phases.append(f"alternate RGO/{label} from RGO {start_rgo} to {end_rgo}")
                    i = j
                    continue

            phases.append(f"RGO to {entry[1]}")
            i += 1
        else:
            name, rgo_at, total = entry
            label = labels.get(name, name)
            phases.append(f"{label} to {total}")
            i += 1

    # Merge trailing buildings after max RGO
    tail = []
    while phases and not phases[-1].startswith("RGO") and not phases[-1].startswith("alternate"):
        tail.insert(0, phases.pop())
    if tail:
        phases.append("then " + ", ".join(tail))

    return first_built, phases


def build_food_compound_order(wb, food_goods, food_buildings):
    """Compound build order: accounts for modifier interactions between buildings."""
    ws = wb.create_sheet("Compound Build Order")

    ws.cell(row=1, column=1, value="Compound Food Build Order by RGO").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Simulates greedy optimization: at each step, build whichever gives the most "
                  "marginal food (accounting for how modifiers boost each other). "
                  "Building caps scale with RGO. Skip unavailable buildings.").font = SUBTITLE_FONT

    bld_names = [b for b in FOOD_BUILDING_ORDER
                 if b in food_buildings and b not in ("market_village", "elephant_hunting_grounds")]

    MAX_SLOTS = len(bld_names)
    MAX_RGO = 40  # simulation limit

    fixed_headers = ["RGO", "RGO\nFood/Lvl"]
    slot_headers = []
    for i in range(1, MAX_SLOTS + 1):
        slot_headers.extend([
            f"#{i} Building",
            f"#{i} Build After\n(RGO lvl)",
            f"#{i} Food/Lvl",
            f"#{i} Requires",
        ])

    tail_headers = ["Build Pattern"]
    all_headers = fixed_headers + slot_headers + tail_headers
    header_row = 4
    for i, h in enumerate(all_headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(all_headers))

    sorted_goods = sorted(food_goods.items(), key=lambda x: -x[1]["food_value"])

    row = header_row + 1
    for rgo_name, rgo_data in sorted_goods:
        rgo_fv = rgo_data["food_value"]

        # Build simulation input for compatible buildings
        sim_buildings = []
        for bld_name in bld_names:
            bld = food_buildings[bld_name]
            req_rgos = bld.get("requirements", {}).get("rgo", [])
            if req_rgos and rgo_name not in req_rgos:
                continue

            goods_f, mod_f, flat_f, rgo_note = calc_building_food(bld, food_goods)
            rgo_out_mod = 0.0
            if rgo_note:
                # Extract the modifier for this specific RGO good
                rgo_mods = bld.get("rgo_output_modifiers", {})
                rgo_out_mod = rgo_mods.get(rgo_name, 0)

            ml = bld.get("max_levels", 1)
            # Fixed caps (windmill=1, canal=1) stay fixed;
            # scaling caps (rural_building_cap, irrigant_cap) track RGO
            max_lvl = ml if isinstance(ml, int) else MAX_RGO

            sim_buildings.append({
                "name": bld_name,
                "goods_food": goods_f + flat_f,
                "food_mod": mod_f,
                "rgo_output_mod": rgo_out_mod,
                "max_levels": max_lvl,
            })

        # Run simulation
        first_built, phases = _simulate_compound_build_order(rgo_fv, sim_buildings)

        # Build display entries sorted by compound crossover
        entries = []
        for b in sim_buildings:
            bld_name = b["name"]
            bld = food_buildings[bld_name]
            label = FOOD_BUILDING_LABELS.get(bld_name, bld_name)
            goods_f, mod_f, flat_f, rgo_note = calc_building_food(bld, food_goods)
            constant = goods_f + flat_f

            compound_after = first_built.get(bld_name)
            if compound_after is None:
                build_after = "after max RGO"
                sort_val = 10000
            else:
                build_after = compound_after
                sort_val = compound_after

            if rgo_note:
                food_display = rgo_note
            elif constant > 0 and mod_f > 0:
                food_display = f"{constant:.1f} + {mod_f:.0%}/lvl"
            elif constant > 0:
                food_display = f"{constant:.1f}"
            elif mod_f > 0:
                food_display = f"{mod_f:.0%}/lvl"
            else:
                food_display = f"{constant:.1f}"

            reqs = bld.get("requirements", {})
            req_parts = []
            if reqs.get("vegetation"):
                req_parts.append("veg: " + ", ".join(reqs["vegetation"]))
            if reqs.get("features"):
                req_parts.append(", ".join(reqs["features"]))
            requires = "; ".join(req_parts) if req_parts else "any"

            entries.append({
                "label": label,
                "build_after": build_after,
                "food_display": food_display,
                "requires": requires,
                "sort_val": (sort_val, -constant),
            })

        entries.sort(key=lambda x: x["sort_val"])

        # Write row
        ws.cell(row=row, column=1, value=rgo_name).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=rgo_fv).border = THIN_BORDER

        col = 3
        for e in entries:
            ws.cell(row=row, column=col, value=e["label"]).border = THIN_BORDER
            ws.cell(row=row, column=col + 1, value=e["build_after"]).border = THIN_BORDER
            ws.cell(row=row, column=col + 2, value=e["food_display"]).border = THIN_BORDER
            ws.cell(row=row, column=col + 3, value=e["requires"]).border = THIN_BORDER
            col += 4

        # Build Pattern column (fixed position matching header)
        pattern_col = len(fixed_headers) + MAX_SLOTS * 4 + 1
        pattern_str = " > ".join(phases)
        ws.cell(row=row, column=pattern_col, value=pattern_str).border = THIN_BORDER

        row += 1

    auto_width(ws, max_width=80)
    ws.freeze_panes = f"A{header_row + 1}"


def _greedy_allocate(budget, rgo_fv, buildings):
    """Allocate a fixed number of build levels optimally using greedy marginal.

    budget: total levels to allocate across RGO + buildings.
    Returns (n_rgo, rgo_mod, lvls_dict, food_production, food_capacity, total_food_mod).
    """
    n_rgo = 0
    rgo_mod = 0.0
    lvls = {b["name"]: 0 for b in buildings}
    BASE_FOOD_CAP = 50  # rural settlement base

    for _ in range(budget):
        # Current state
        base_food = n_rgo * rgo_fv * (1 + rgo_mod)
        for b in buildings:
            base_food += lvls[b["name"]] * b["goods_food"]
        total_mod = sum(lvls[b["name"]] * b["food_mod"] for b in buildings)

        best_val = -1
        best_name = None

        # RGO
        m = rgo_fv * (1 + rgo_mod) * (1 + total_mod)
        if m > best_val:
            best_val = m
            best_name = "_rgo"

        # Buildings
        for b in buildings:
            if b["rgo_output_mod"] > 0 and lvls[b["name"]] == 0:
                m = b["rgo_output_mod"] * n_rgo * rgo_fv * (1 + total_mod)
            elif b["max_levels"] is not None and lvls[b["name"]] >= b["max_levels"]:
                continue
            else:
                m = b["goods_food"] * (1 + total_mod) + b["food_mod"] * base_food
            if m > best_val:
                best_val = m
                best_name = b["name"]

        if best_name == "_rgo":
            n_rgo += 1
        elif best_name:
            lvls[best_name] += 1
            b_data = next(b for b in buildings if b["name"] == best_name)
            if b_data["rgo_output_mod"] > 0:
                rgo_mod += b_data["rgo_output_mod"]

    # Compute final food production
    base_food = n_rgo * rgo_fv * (1 + rgo_mod)
    for b in buildings:
        base_food += lvls[b["name"]] * b["goods_food"]
    total_mod = sum(lvls[b["name"]] * b["food_mod"] for b in buildings)
    food_production = base_food * (1 + total_mod)

    # Food capacity
    food_cap = BASE_FOOD_CAP
    for b in buildings:
        food_cap += lvls[b["name"]] * b.get("food_capacity", 0)

    return n_rgo, rgo_mod, lvls, food_production, food_cap, total_mod


def _calc_rgo_max(dev, pops_units):
    """RGO max levels: (2 + 0.025×pop_in_thousands + 0.1×dev) × (1 + 1.0 rural rank).

    Ignoring literacy for simplicity.
    """
    base = 2 + 0.025 * pops_units + 0.1 * dev
    return int(base * 2)  # ×2 from rural_settlement +100%


def _calc_rural_building_cap(dev, rgo_max, has_river):
    """rural_building_cap = 1 + 0.1×dev + 0.5×max_rgo_workers + river."""
    cap = 1 + 0.1 * dev + 0.5 * rgo_max
    if has_river:
        cap += 1
    return int(cap)


def _calc_irrigant_cap(dev, has_river):
    """irrigant_cap = 2 + 0.2×dev + 2 if river."""
    cap = 2 + 0.2 * dev
    if has_river:
        cap += 2
    return int(cap)


def _calc_pop_capacity(dev, irr_levels):
    """Population capacity for wheat/farmland/flatland/oceanic/river location.

    Base: 100K (farmland)
    Multipliers: oceanic +100%, river +10%, dev +2.5%/dev, traditional economy +22.5%
    Additive: irrigation +1K/level, equator ~2.7K (fixed estimate)
    """
    base = 100 + 2.7 + irr_levels * 1.0  # in thousands
    mult = 1 + 1.0 + 0.1 + 0.025 * dev + 0.225  # oceanic + river + dev + trad economy
    return base * mult


def _run_full_simulation(allocator_fn, rgo_fv, sim_buildings, months=1200,
                         start_pops=5.0, start_dev=5.0, has_river=True):
    """Run monthly simulation with co-evolving dev, pops, and building caps.

    Scenario: wheat, farmland, flatland, oceanic, river, rural settlement.

    Growth sources (annual rates, applied monthly as rate/12):
    - Rural settlement: +0.1%
    - Prosperity (scaled, ~half): +0.1%
    - Food storage: +0.08% per 12 months stored (capped at 10× base)
    - Free land: +0.25% × (1 - pops/pop_capacity)

    Dev growth: ~0.004/mo from prosperity × (1 + road 5% + farmland 10% + river 5%)
    """
    RURAL_GROWTH = 0.001         # annual +0.1%
    PROSPERITY_GROWTH = 0.0005   # annual +0.05%
    FOOD_GROWTH_PER_12 = 0.0008  # annual +0.08% per 12 months stored
    FOOD_GROWTH_CAP_MULT = 10    # max 10× base from food
    FREE_LAND_GROWTH = 0.0025    # annual +0.25% at full free land
    FOOD_DECAY = 0.005           # 0.5%/mo on stored food
    POP_CONSUMPTION = 1.0        # per 1K pops per month
    # Dev growth: prosperity base ~0.004/mo × (1 + 0.05 road + 0.10 farmland + 0.05 river)
    DEV_GROWTH_PER_MONTH = 0.004 * 1.20
    BASE_FOOD_CAP = 50           # rural settlement base

    pops = start_pops
    dev = start_dev
    food_stored = 0.0
    cumulative_food = 0.0
    yearly = []

    for month in range(1, months + 1):
        # Derive caps
        rgo_max = _calc_rgo_max(dev, pops)
        rural_cap = _calc_rural_building_cap(dev, rgo_max, has_river)
        irr_cap = _calc_irrigant_cap(dev, has_river)

        caps = {}
        for b in sim_buildings:
            if b["name"] == "irrigation_systems":
                caps[b["name"]] = irr_cap
            elif b["max_levels"] is not None:
                caps[b["name"]] = b["max_levels"]
            else:
                caps[b["name"]] = rural_cap

        pop_budget = max(int(pops), 1)

        n_rgo, rgo_mod, lvls, food_prod, food_cap_bld, total_mod = allocator_fn(
            pop_budget, rgo_fv, sim_buildings, rgo_max, caps
        )

        # Food capacity
        food_cap = BASE_FOOD_CAP
        for b in sim_buildings:
            food_cap += lvls[b["name"]] * b.get("food_capacity", 0)

        # Food storage
        food_consumption = pops * POP_CONSUMPTION
        surplus = food_prod - food_consumption
        food_stored = food_stored * (1 - FOOD_DECAY) + surplus
        food_stored = max(0, min(food_stored, food_cap))
        months_stored = food_stored / food_consumption if food_consumption > 0 else 0

        # Population capacity
        irr_levels = lvls.get("irrigation_systems", 0)
        pop_cap = _calc_pop_capacity(dev, irr_levels)

        # Annual growth rate (applied monthly as /12)
        base_annual = RURAL_GROWTH + PROSPERITY_GROWTH
        food_bonus = min(FOOD_GROWTH_PER_12 * months_stored / 12,
                         base_annual * FOOD_GROWTH_CAP_MULT)
        free_land_ratio = max(0, (pop_cap - pops) / pop_cap) if pop_cap > 0 else 0
        free_land_bonus = FREE_LAND_GROWTH * free_land_ratio
        growth_rate = base_annual + food_bonus + free_land_bonus

        # Apply monthly, cap at population capacity
        new_pops = pops + pops * growth_rate / 12
        if new_pops > pop_cap:
            new_pops = max(pops, pop_cap)
        pops = new_pops

        dev += DEV_GROWTH_PER_MONTH
        cumulative_food += food_prod

        if month % 12 == 0:
            yearly.append({
                "year": month // 12,
                "pops": pops,
                "dev": dev,
                "rgo_max": rgo_max,
                "rural_cap": rural_cap,
                "irr_cap": irr_cap,
                "pop_cap": pop_cap,
                "n_rgo": n_rgo,
                "farming_village": lvls.get("farming_village", 0),
                "windmill": lvls.get("windmill", 0),
                "irrigation": lvls.get("irrigation_systems", 0),
                "sheep_farms": lvls.get("sheep_farms", 0),
                "fruit_orchard": lvls.get("fruit_orchard", 0),
                "food_prod": food_prod,
                "food_cap": food_cap,
                "months_stored": months_stored,
                "growth_rate": growth_rate,
                "cumulative_food": cumulative_food,
            })

    return yearly


def _capped_greedy_allocate(pop_budget, rgo_fv, buildings, rgo_max, caps):
    """Greedy allocation respecting both pop budget and per-building caps."""
    n_rgo = 0
    rgo_mod = 0.0
    lvls = {b["name"]: 0 for b in buildings}
    used = 0

    for _ in range(pop_budget):
        base_food = n_rgo * rgo_fv * (1 + rgo_mod)
        for b in buildings:
            base_food += lvls[b["name"]] * b["goods_food"]
        total_mod = sum(lvls[b["name"]] * b["food_mod"] for b in buildings)

        best_val = -1
        best_name = None

        # RGO
        if n_rgo < rgo_max:
            m = rgo_fv * (1 + rgo_mod) * (1 + total_mod)
            if m > best_val:
                best_val = m
                best_name = "_rgo"

        for b in buildings:
            bld_cap = caps.get(b["name"], 0)
            if lvls[b["name"]] >= bld_cap:
                continue
            if b["rgo_output_mod"] > 0 and lvls[b["name"]] == 0:
                m = b["rgo_output_mod"] * n_rgo * rgo_fv * (1 + total_mod)
            else:
                m = b["goods_food"] * (1 + total_mod) + b["food_mod"] * base_food
            if m > best_val:
                best_val = m
                best_name = b["name"]

        if best_name is None:
            break
        if best_name == "_rgo":
            n_rgo += 1
        else:
            lvls[best_name] += 1
            bd = next(b for b in buildings if b["name"] == best_name)
            if bd["rgo_output_mod"] > 0:
                rgo_mod += bd["rgo_output_mod"]
        used += 1

    # Compute food production
    base_food = n_rgo * rgo_fv * (1 + rgo_mod)
    for b in buildings:
        base_food += lvls[b["name"]] * b["goods_food"]
    total_mod = sum(lvls[b["name"]] * b["food_mod"] for b in buildings)
    food_prod = base_food * (1 + total_mod)

    food_cap = 50
    for b in buildings:
        food_cap += lvls[b["name"]] * b.get("food_capacity", 0)

    return n_rgo, rgo_mod, lvls, food_prod, food_cap, total_mod


def _capped_farming_first_allocate(pop_budget, rgo_fv, buildings, rgo_max, caps):
    """Farming-first allocation respecting building caps."""
    n_rgo = 0
    rgo_mod = 0.0
    lvls = {b["name"]: 0 for b in buildings}
    remaining = pop_budget

    fv_bld = next((b for b in buildings if b["name"] == "farming_village"), None)
    wm_bld = next((b for b in buildings if b["name"] == "windmill"), None)

    # 1. Max farming village (up to cap)
    if fv_bld and remaining > 0:
        fv_cap = caps.get("farming_village", 0)
        fv_lvl = min(remaining, fv_cap)
        lvls["farming_village"] = fv_lvl
        remaining -= fv_lvl

    # 2. Windmill
    if wm_bld and remaining > 0 and caps.get("windmill", 0) >= 1:
        lvls["windmill"] = 1
        rgo_mod += wm_bld["rgo_output_mod"]
        remaining -= 1

    # 3. RGO with remaining (up to rgo_max)
    rgo_lvl = min(remaining, rgo_max)
    n_rgo = rgo_lvl
    remaining -= rgo_lvl

    # 4. Fill any remaining into irrigation, sheep, etc.
    for b in buildings:
        if remaining <= 0:
            break
        if b["name"] in ("farming_village", "windmill"):
            continue
        bld_cap = caps.get(b["name"], 0)
        add = min(remaining, bld_cap - lvls[b["name"]])
        if add > 0:
            lvls[b["name"]] += add
            remaining -= add

    # Compute food
    base_food = n_rgo * rgo_fv * (1 + rgo_mod)
    for b in buildings:
        base_food += lvls[b["name"]] * b["goods_food"]
    total_mod = sum(lvls[b["name"]] * b["food_mod"] for b in buildings)
    food_prod = base_food * (1 + total_mod)

    food_cap = 50
    for b in buildings:
        food_cap += lvls[b["name"]] * b.get("food_capacity", 0)

    return n_rgo, rgo_mod, lvls, food_prod, food_cap, total_mod


def build_food_simulation(wb, food_goods, food_buildings):
    """100-year simulation comparing optimal vs farming-first with real caps."""
    ws = wb.create_sheet("100-Year Simulation")

    ws.cell(row=1, column=1, value="100-Year Food Simulation: Optimal vs Farming-First (Wheat)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Wheat, farmland, flatland, oceanic, river, rural settlement. 5k pops, dev 5. "
                  "Growth: +0.1% rural + 0.1% prosperity + 0.25%×free_land + 0.08%/12mo stored. "
                  "Pop cap ~240K+ (farmland+oceanic+river). Dev grows ~0.005/mo.").font = SUBTITLE_FONT

    RGO_FV = 8.0

    bld_names = [b for b in FOOD_BUILDING_ORDER
                 if b in food_buildings and b not in ("market_village", "elephant_hunting_grounds")]

    sim_buildings = []
    for bld_name in bld_names:
        bld = food_buildings[bld_name]
        req_rgos = bld.get("requirements", {}).get("rgo", [])
        if req_rgos and "wheat" not in req_rgos:
            continue
        goods_f, mod_f, flat_f, rgo_note = calc_building_food(bld, food_goods)
        rgo_out = bld.get("rgo_output_modifiers", {}).get("wheat", 0)
        ml = bld.get("max_levels", 1)
        max_lvl = ml if isinstance(ml, int) else None
        food_cap_val = bld.get("food_modifiers", {}).get("local_food_capacity", 0)

        sim_buildings.append({
            "name": bld_name,
            "goods_food": goods_f + flat_f,
            "food_mod": mod_f,
            "rgo_output_mod": rgo_out,
            "max_levels": max_lvl,
            "food_capacity": food_cap_val,
        })

    # Run both strategies
    optimal = _run_full_simulation(_capped_greedy_allocate, RGO_FV, sim_buildings)
    fv_first = _run_full_simulation(_capped_farming_first_allocate, RGO_FV, sim_buildings)

    # Write comparison
    headers = [
        "Year", "Dev",
        "OPT\nPops", "OPT\nRGO", "OPT\nFarm Vil", "OPT\nWindmill", "OPT\nIrrigation",
        "OPT\nFood/Mo", "OPT\nFood Cap", "OPT\nMo Stored", "OPT\nGrowth/Yr",
        "OPT\nCum. Food",
        "",
        "FV1st\nPops", "FV1st\nRGO", "FV1st\nFarm Vil", "FV1st\nWindmill", "FV1st\nIrrigation",
        "FV1st\nFood/Mo", "FV1st\nFood Cap", "FV1st\nMo Stored", "FV1st\nGrowth/Yr",
        "FV1st\nCum. Food",
        "",
        "OPT\nAdvantage",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    OPT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    FV_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for r, (opt, fv1) in enumerate(zip(optimal, fv_first), header_row + 1):
        adv = opt["cumulative_food"] - fv1["cumulative_food"]
        adv_pct = adv / fv1["cumulative_food"] if fv1["cumulative_food"] > 0 else 0

        opt_vals = [
            opt["year"], round(opt["dev"], 1),
            round(opt["pops"] * 1000),
            opt["n_rgo"], opt["farming_village"], opt["windmill"], opt["irrigation"],
            round(opt["food_prod"], 1), round(opt["food_cap"]),
            round(opt["months_stored"], 1), opt["growth_rate"],
            round(opt["cumulative_food"]),
        ]
        fv_vals = [
            round(fv1["pops"] * 1000),
            fv1["n_rgo"], fv1["farming_village"], fv1["windmill"], fv1["irrigation"],
            round(fv1["food_prod"], 1), round(fv1["food_cap"]),
            round(fv1["months_stored"], 1), fv1["growth_rate"],
            round(fv1["cumulative_food"]),
        ]

        for i, v in enumerate(opt_vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = THIN_BORDER
            if i == 2:
                cell.number_format = NUM_FMT_2
            elif i == 11:
                cell.number_format = "0.000%"
            elif isinstance(v, float):
                cell.number_format = NUM_FMT_2
            if i >= 3:
                cell.fill = OPT_FILL

        ws.cell(row=r, column=13, value="").border = THIN_BORDER

        for i, v in enumerate(fv_vals, 14):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = THIN_BORDER
            cell.fill = FV_FILL
            if i == 22:  # growth rate column
                cell.number_format = "0.000%"
            elif isinstance(v, float):
                cell.number_format = NUM_FMT_2

        ws.cell(row=r, column=25, value="").border = THIN_BORDER

        cell = ws.cell(row=r, column=26, value=adv_pct)
        cell.border = THIN_BORDER
        cell.number_format = "+0.0%;-0.0%;0%"
        if adv_pct > 0:
            cell.font = Font(bold=True, color="008000")
        elif adv_pct < 0:
            cell.font = Font(bold=True, color="FF0000")

    auto_width(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def build_vassal_breakeven(wb):
    """Break-even analysis for vassal annexation decisions.

    Compares monthly income from keeping a vassal vs annexing at various
    control levels, accounting for subject income rates and the economical
    base increase that raises all slider costs.

    Model (per unit of vassal monthly income V):
      As vassal:  you receive income_rate × V
      If annexed at control C:  you receive C × V, but slider costs rise
      Break-even control = income_rate + slider_overhead
    """
    ws = wb.create_sheet("Vassal Break-Even")

    # --------------- Game constants ---------------
    ECON_TAX_WEIGHT = 0.5       # ECONOMICAL_BASE_FROM_TAX_BASE
    ECON_SUBJ_WEIGHT = 0.05     # ECONOMICAL_BASE_FROM_SUBJECT
    # Net econ base change when annexing (per unit vassal tax base):
    #   gain 0.5 from direct ownership, lose 0.05*0.5 from dropping subject bonus
    ECON_CHANGE = ECON_TAX_WEIGHT - ECON_SUBJ_WEIGHT * ECON_TAX_WEIGHT  # 0.475

    SUBJECTS = [
        # (name, scaled_gold, can_annex, min_years, min_opinion, stall_opinion)
        ("Vassal",           0.20,  True,  10, 150, 125),
        ("Fiefdom",          0.20,  True,  10, 150, 125),
        ("March",            0.10,  False, None, None, None),
        ("Tributary",        0.20,  False, None, None, None),
        ("Samanta",          0.10,  True,  10, 150, 125),
        ("Maha Samanta",     0.20,  True,  10, 150, 125),
        ("P. Maha Samanta",  0.33,  True,  10, 150, 125),
        ("Trade Company",    0.50,  False, None, None, None),
        ("Colonial Nation",  0.025, False, None, None, None),
        ("Hanseatic Member", 0.05,  False, None, None, None),
        ("UC Bey",           0.05,  False, None, None, None),
    ]

    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def _breakeven(income_rate, slider_spend_frac):
        """Min control % for annexation to beat keeping the vassal.

        income_rate:       fraction of vassal income you receive (e.g. 0.20)
        slider_spend_frac: your monthly slider spend as fraction of income
        Slider overhead = 0.95 × slider_spend_frac (econ base increase
        makes all slider costs go up by that proportion).
        """
        return max(0, min(1, income_rate + 0.95 * slider_spend_frac))

    row = 1

    # ===== Title =====
    ws.cell(row=row, column=1,
            value="Vassal Annexation Break-Even Analysis").font = TITLE_FONT
    row += 2

    # ===== Section 1: Subject Income Rates =====
    ws.cell(row=row, column=1, value="Subject Income Rates").font = SUBTITLE_FONT
    ws.cell(row=row, column=4,
            value="Source: in_game/common/prices/03_diplomacy.txt").font = Font(
                italic=True, color="808080")
    row += 1

    s_headers = ["Subject Type", "Income Rate", "Annexable",
                 "Min Years", "Min Opinion", "Stall Opinion"]
    for i, h in enumerate(s_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(s_headers))
    row += 1

    for name, rate, annex, years, opinion, stall in SUBJECTS:
        vals = [name, rate, "Yes" if annex else "No",
                years if years is not None else "-",
                opinion if opinion is not None else "-",
                stall if stall is not None else "-"]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
            if i == 2 and isinstance(v, float):
                cell.number_format = "0%"
        row += 1

    # ===== Section 2: Key Game Constants =====
    row += 2
    ws.cell(row=row, column=1, value="Key Game Constants").font = SUBTITLE_FONT
    ws.cell(row=row, column=3,
            value="Source: loading_screen/common/defines/00_defines.txt").font = Font(
                italic=True, color="808080")
    row += 1

    constants = [
        ("ECONOMICAL_BASE_FROM_TAX_BASE", 0.5),
        ("ECONOMICAL_BASE_FROM_SUBJECT", 0.05),
        ("STABILITY_INVEST_FACTOR", 0.5),
        ("STABILIY_EXPENSE_FACTOR", 0.1),
        ("LOW_CONTROL_THRESHOLD_FOR_BEST_TAX", 0.95),
    ]

    c_headers = ["Define", "Value"]
    for i, h in enumerate(c_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(c_headers))
    row += 1

    for name, val in constants:
        for i, v in enumerate([name, val], 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
        row += 1

    # ===== Section 3: Break-Even Model =====
    row += 2
    ws.cell(row=row, column=1, value="Break-Even Model").font = SUBTITLE_FONT
    row += 1
    notes = [
        "All values expressed per 1 gold/month of vassal income (multiply by "
        "vassal's actual monthly income for gold amounts).",
        "As VASSAL:  you receive income_rate \u00d7 V  "
        "(e.g. 20% \u00d7 V for vassal type)",
        "If ANNEXED at control C:  you receive C \u00d7 V, but all slider "
        "costs increase proportionally to your econ base growth",
        "Break-even control = income_rate + 0.95 \u00d7 slider_spend_frac",
        "slider_spend_frac = your monthly slider expenses / your monthly "
        "income (e.g. 0.10 if 10% of income goes to sliders)",
        "Econ base grows by ~0.95 \u00d7 (vassal/overlord economy ratio), "
        "raising ALL gold-based costs by that %",
        "Econ base increase is independent of economy ratio for the "
        "break-even %, but ratio determines absolute gold amounts",
    ]
    for n in notes:
        ws.cell(row=row, column=1, value=n).font = Font(italic=True, size=10)
        row += 1

    # ===== Section 4: Break-Even by Subject Type =====
    row += 2
    ws.cell(row=row, column=1,
            value="Break-Even Control % by Subject Type").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Minimum average control in annexed land to beat vassal "
            "income, at different slider spending levels."
            ).font = Font(italic=True, size=10)
    row += 1

    annexable = [(n, r) for n, r, a, *_ in SUBJECTS if a]
    slider_spends = [0.00, 0.05, 0.10, 0.15, 0.20, 0.30, 0.50]

    headers = ["Slider Spend\n(% of income)"] + [n for n, _ in annexable]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for ss in slider_spends:
        cell = ws.cell(row=row, column=1, value=ss)
        cell.number_format = "0%"
        cell.border = THIN_BORDER
        for j, (_, rate) in enumerate(annexable, 2):
            be = _breakeven(rate, ss)
            cell = ws.cell(row=row, column=j)
            cell.border = THIN_BORDER
            cell.value = be
            cell.number_format = "0%"
            if be < 0.40:
                cell.fill = GREEN
            elif be < 0.70:
                cell.fill = YELLOW
            else:
                cell.fill = RED
        row += 1

    # ===== Section 5: Monthly Slider Cost Increase =====
    row += 2
    ws.cell(row=row, column=1,
            value="Monthly Slider Cost Increase from Annexation").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Shows how much your monthly slider expenses rise after "
            "annexing, by vassal/overlord economy ratio and current slider "
            "spending.  Values are gold/month per 1 gold of vassal income."
            ).font = Font(italic=True, size=10)
    row += 1

    ratios2 = [0.10, 0.20, 0.30, 0.50, 0.75, 1.00]
    slider_spends2 = [0.05, 0.10, 0.15, 0.20, 0.30, 0.50]

    headers = ["Slider Spend\n(% of income)"] + [f"Ratio\n{r:.0%}" for r in ratios2]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for ss in slider_spends2:
        cell = ws.cell(row=row, column=1, value=ss)
        cell.number_format = "0%"
        cell.border = THIN_BORDER
        for j, ratio in enumerate(ratios2, 2):
            # Econ base increase = 0.95 × ratio, so slider costs rise by
            # that fraction.  Extra gold/month = slider_spend × econ_increase
            # Per unit vassal income: 0.95 × ratio × ss × (O/V) = 0.95 × ss
            # (ratio cancels).  But absolute gold = 0.95 × ratio × ss × O.
            # Here show per unit overlord income:
            extra = 0.95 * ratio * ss
            cell = ws.cell(row=row, column=j, value=extra)
            cell.number_format = NUM_FMT_3
            cell.border = THIN_BORDER
        row += 1

    # ===== Section 6: Net Income Comparison by Control Level =====
    row += 2
    ws.cell(row=row, column=1,
            value="Monthly Net Income: Annex vs Keep").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="% of vassal income gained (+) or lost (-) by annexing.  "
            "Slider spend = 10% of income."
            ).font = Font(italic=True, size=10)
    row += 1

    control_levels = [0.05, 0.10, 0.15, 0.20, 0.30, 0.40,
                      0.50, 0.60, 0.70, 0.80, 0.90, 0.95, 1.00]
    ref_ss = 0.10
    # Slider overhead per unit V: 0.95 × slider_spend_frac
    overhead = 0.95 * ref_ss

    headers = ["Control"] + [n for n, _ in annexable]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for C in control_levels:
        cell = ws.cell(row=row, column=1, value=C)
        cell.number_format = "0%"
        cell.border = THIN_BORDER
        for j, (_, rate) in enumerate(annexable, 2):
            # You lose rate × V from vassal, gain C × V from control,
            # pay overhead × V in higher slider costs
            diff = (C - rate) - overhead
            cell = ws.cell(row=row, column=j, value=diff)
            cell.number_format = "+0%;-0%;0%"
            cell.border = THIN_BORDER
            if diff > 0.001:
                cell.fill = GREEN
            elif diff < -0.001:
                cell.fill = RED
            else:
                cell.fill = YELLOW
        row += 1

    # ===== Section 7: Econ Base Impact =====
    row += 2
    ws.cell(row=row, column=1,
            value="Economical Base Impact of Annexation").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Annexing adds 0.5\u00d7vassal_tax_base to your econ base, "
            "minus the 0.05 lost subject contribution.  Net: +0.475 per unit "
            "vassal tax base.  Econ base scales ALL gold-based slider costs."
            ).font = Font(italic=True, size=10)
    row += 1

    e_headers = ["Economy\nRatio", "Your Econ Base\nIncrease",
                 "Slider Cost\nIncrease"]
    for i, h in enumerate(e_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(e_headers))
    row += 1

    for ratio in [0.05, 0.10, 0.20, 0.30, 0.50, 0.75, 1.00, 1.50, 2.00]:
        # Econ base change as % of overlord's:
        # 0.475 * T_v / (0.5 * T_o) = 0.95 * ratio
        econ_pct = 0.95 * ratio
        for i, v in enumerate([ratio, econ_pct, econ_pct], 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
            cell.number_format = "0%"
        row += 1

    auto_width(ws, max_width=50)


def build_annex_batching(wb):
    """Optimal number of concurrent subject annexations by base speed.

    Each concurrent annexation adds MULTIPLE_ANNEX_PENALTY (-0.5) to speed.
    Throughput = N × (S − 0.5N).  Optimal N (continuous) = S.
    """
    ws = wb.create_sheet("Annexation Batching")

    PENALTY = 0.5  # |MULTIPLE_ANNEX_PENALTY|

    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def eff_speed(s, n):
        """Speed per annexation with n concurrent at base speed s."""
        return max(0.0, s - PENALTY * n)

    def tp(s, n):
        """Total throughput with n concurrent at base speed s."""
        return n * eff_speed(s, n)

    def best_n(s):
        """Integer N maximising throughput (lower N wins ties)."""
        n_lo = max(1, int(s))
        n_hi = n_lo + 1
        return n_hi if tp(s, n_hi) > tp(s, n_lo) else n_lo

    def best_ns(s):
        """Set of equally-optimal integer N values."""
        n_lo = max(1, int(s))
        n_hi = n_lo + 1
        t_lo, t_hi = tp(s, n_lo), tp(s, n_hi)
        if t_hi > 0 and abs(t_hi - t_lo) < 0.001:
            return {n_lo, n_hi}
        return {n_hi} if t_hi > t_lo else {n_lo}

    row = 1

    # ===== Title =====
    ws.cell(row=row, column=1,
            value="Optimal Concurrent Annexation Count").font = TITLE_FONT
    row += 2

    # ===== Notes =====
    notes = [
        "Each concurrent annexation applies MULTIPLE_ANNEX_PENALTY = \u22120.5 "
        "(\u221250%) to annexation speed.",
        "Effective speed per annexation = Base Speed \u2212 50% \u00d7 N, where "
        "N = number of concurrent annexations.",
        "Throughput = N \u00d7 Effective Speed.  Optimal N \u2248 Base Speed "
        "(as a multiplier, e.g. 5 at 500%).",
        "At the optimal, each annexation runs at half the base speed "
        "and total throughput = Base Speed\u00b2 \u00f7 2.",
        "MAX_ANNEX_SIZE = 2 caps something else (not concurrent count).",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=10)
        row += 1
    ws.cell(row=row, column=1,
            value="Source: loading_screen/common/defines/00_defines.txt"
            ).font = Font(italic=True, color="808080")
    row += 2

    # ===== Section 1: Quick Reference =====
    ws.cell(row=row, column=1,
            value="Optimal Count by Base Speed").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="When two N values give equal throughput, the lower is shown "
            "(faster per-subject completion)."
            ).font = Font(italic=True, size=10, color="808080")
    row += 1

    q_headers = ["Base Speed", "Optimal N", "Speed per\nAnnexation",
                  "Total\nThroughput", "Speed-Up\nvs N=1"]
    for i, h in enumerate(q_headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="center")
    style_header_row(ws, row, len(q_headers))
    row += 1

    speeds_qr = [s / 2 for s in range(2, 31)]  # 1.0 to 15.0, step 0.5

    for S in speeds_qr:
        n_opt = best_n(S)
        spd = eff_speed(S, n_opt)
        t_opt = tp(S, n_opt)
        t_1 = tp(S, 1)
        ratio = t_opt / t_1 if t_1 > 0 else 0

        vals = [S, n_opt, spd, t_opt, ratio]
        fmts = [NUM_FMT_PCT, "0", NUM_FMT_PCT, NUM_FMT_2, '0.00"x"']
        for i, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
            cell.number_format = fmt
            if i >= 2:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # ===== Section 2: Throughput Matrix =====
    row += 2
    ws.cell(row=row, column=1,
            value="Throughput Matrix").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Total annexation progress per month across all concurrent "
            "annexations.  Green = optimal for that base speed."
            ).font = Font(italic=True, size=10)
    row += 1

    max_n = 15
    matrix_speeds = [float(s) for s in range(1, 16)]

    # Label above header
    ws.cell(row=row, column=2,
            value="Concurrent Annexations \u2192").font = Font(
                italic=True, size=10, color="808080")
    row += 1

    # Header row
    ws.cell(row=row, column=1, value="Base Speed")
    for n in range(1, max_n + 1):
        ws.cell(row=row, column=n + 1, value=n)
    style_header_row(ws, row, max_n + 1)
    row += 1

    for S in matrix_speeds:
        cell = ws.cell(row=row, column=1, value=S)
        cell.number_format = NUM_FMT_PCT
        cell.border = THIN_BORDER

        opt_set = best_ns(S)

        for n in range(1, max_n + 1):
            spd = eff_speed(S, n)
            cell = ws.cell(row=row, column=n + 1)
            cell.border = THIN_BORDER
            if spd <= 0:
                cell.value = "\u2014"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="C0C0C0")
            else:
                cell.value = round(tp(S, n), 2)
                cell.number_format = NUM_FMT_2
                if n in opt_set:
                    cell.fill = GREEN
        row += 1

    auto_width(ws)


def build_pop_demands(wb, pop_demands):
    """Pop demands per pop type, total demand value, and noble equivalency."""
    ws = wb.create_sheet("Pop Demands")

    POP_TYPES = ["nobles", "clergy", "burghers", "soldiers", "laborers",
                 "peasants", "slaves", "tribesmen"]
    POP_LABELS = {
        "nobles": "Nobles", "clergy": "Clergy", "burghers": "Burghers",
        "soldiers": "Soldiers", "laborers": "Laborers", "peasants": "Peasants",
        "slaves": "Slaves", "tribesmen": "Tribesmen",
    }

    POP_FILLS = {
        "nobles": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "clergy": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "burghers": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "soldiers": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        "laborers": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
        "peasants": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "slaves": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "tribesmen": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    }

    goods = pop_demands["goods"]
    pop_info = pop_demands["pop_types"]

    # Pre-compute demand value per good per pop type and sort goods by noble value desc
    for g in goods:
        g["values"] = {p: g["demands"][p] * g["price"] for p in POP_TYPES}
    goods.sort(key=lambda g: g["values"]["nobles"], reverse=True)

    # Compute totals
    totals = {p: sum(g["values"][p] for g in goods) for p in POP_TYPES}

    row = 1

    # ===== Title =====
    ws.cell(row=row, column=1,
            value="Pop Type Goods Demands").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Baseline demand per pop (no culture/religion/location modifiers).  "
            "Demand = demand_add \u00d7 demand_multiply from goods definitions."
            ).font = Font(italic=True, size=10)
    row += 1
    ws.cell(row=row, column=1,
            value="Source: in_game/common/goods/*.txt, in_game/common/pop_types/00_default.txt"
            ).font = Font(italic=True, color="808080", size=10)
    row += 2

    # ===== Section 1: Summary & Equivalency =====
    ws.cell(row=row, column=1, value="Summary").font = SUBTITLE_FONT
    row += 1

    # Header
    sum_headers = ["Pop Type", "Food Consumption", "Total Demand Value",
                   "Noble Equivalency"]
    for i, h in enumerate(sum_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(sum_headers))
    sum_header_row = row
    row += 1

    noble_total = totals["nobles"]
    for p in POP_TYPES:
        food = pop_info[p]["food_consumption"]
        total = totals[p]
        equiv = noble_total / total if total > 0 else 0

        vals = [POP_LABELS[p], food, total, equiv]
        fmts = [None, NUM_FMT_2, "0.0000", NUM_FMT_2]
        for i, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN_BORDER
            cell.fill = POP_FILLS[p]
            if fmt:
                cell.number_format = fmt
            if i >= 2:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value="Noble Equivalency: how many of each pop type equals 1 noble "
            "in total goods demand value."
            ).font = Font(italic=True, size=10, color="808080")
    row += 2

    # ===== Section 2: Demand Value Table =====
    ws.cell(row=row, column=1, value="Demand Value by Good").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Value = Demand Quantity \u00d7 Market Price (gold per pop)"
            ).font = Font(italic=True, size=10)
    row += 1

    # Header: Good | Price | Nobles | Clergy | ... | Tribesmen
    val_headers = ["Good", "Price"] + [POP_LABELS[p] for p in POP_TYPES]
    for i, h in enumerate(val_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(val_headers))
    val_header_row = row
    row += 1

    GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    # Track max value per column for highlighting
    col_maxes = {p: 0 for p in POP_TYPES}
    col_max_rows = {p: None for p in POP_TYPES}
    data_start_row = row

    for g in goods:
        name = g["name"].replace("_", " ").replace("goods ", "").title()
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=g["price"])
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_2

        for j, p in enumerate(POP_TYPES):
            val = g["values"][p]
            cell = ws.cell(row=row, column=3 + j, value=val)
            cell.border = THIN_BORDER
            cell.number_format = "0.0000" if val < 0.01 else NUM_FMT_2
            cell.alignment = Alignment(horizontal="center")
            if val > col_maxes[p]:
                col_maxes[p] = val
                col_max_rows[p] = row
        row += 1

    # Total row
    for i, h in enumerate(val_headers, 1):
        ws.cell(row=row, column=i).border = THIN_BORDER
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    for j, p in enumerate(POP_TYPES):
        cell = ws.cell(row=row, column=3 + j, value=totals[p])
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_2
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Highlight max value per pop type column
    for j, p in enumerate(POP_TYPES):
        r = col_max_rows[p]
        if r:
            ws.cell(row=r, column=3 + j).fill = GOLD_FILL

    row += 2

    # ===== Section 3: Demand Quantity Table =====
    ws.cell(row=row, column=1, value="Demand Quantity by Good").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Raw demand quantity per pop (before price weighting)"
            ).font = Font(italic=True, size=10)
    row += 1

    qty_headers = ["Good", "Price"] + [POP_LABELS[p] for p in POP_TYPES]
    for i, h in enumerate(qty_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(qty_headers))
    row += 1

    for g in goods:
        name = g["name"].replace("_", " ").replace("goods ", "").title()
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=g["price"])
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_2

        for j, p in enumerate(POP_TYPES):
            val = g["demands"][p]
            cell = ws.cell(row=row, column=3 + j, value=val)
            cell.border = THIN_BORDER
            cell.number_format = "0.0000" if val < 0.01 else NUM_FMT_2
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # Freeze at value table header
    ws.freeze_panes = ws.cell(row=val_header_row + 1, column=1)

    auto_width(ws)


def main():
    if not DATA_DIR.exists():
        print("No data/ directory found. Run scraper.py first.")
        return

    (land_units, categories, age_progression, prices, combined_arms,
     goods_demands, production_recipes, localizations, naval_units,
     food_goods, food_buildings, building_caps, terrain_food_modifiers,
     forts, pop_demands) = load_data()

    global LOC
    LOC = localizations

    wb = Workbook()

    print("Building Unit Power sheet...")
    build_army_meta(wb, age_progression, categories, prices)

    print("Building Buildable Units sheet...")
    build_buildable_units(wb, land_units, categories, prices)

    print("Building Levy Units sheet...")
    build_levy_units(wb, land_units, categories, prices)

    print("Building Optimal Composition sheet...")
    build_optimal_composition(wb, land_units, categories, combined_arms)

    print("Building Optimal Composition (Morale) sheet...")
    build_optimal_composition_morale(wb, land_units, categories, combined_arms)

    print("Building Optimal Composition (Budget) sheet...")
    build_optimal_composition_budget(wb, land_units, categories, combined_arms, prices)

    print("Building Optimal Composition (Gold) sheet...")
    build_optimal_composition_gold(wb, land_units, categories, combined_arms, prices)

    print("Building Optimal Composition (Iron) sheet...")
    build_optimal_composition_iron(wb, land_units, categories, combined_arms,
                                    goods_demands, production_recipes)

    print("Building Goods Demands sheet...")
    build_goods_demands(wb, land_units, categories, goods_demands)

    print("Building Goods (Generic) sheet...")
    build_goods_demands_generic(wb, land_units, categories, goods_demands, production_recipes)

    print("Building Raw Materials sheet...")
    build_raw_materials(wb, land_units, categories, goods_demands, production_recipes)

    print("Building Upgrade Chains sheet...")
    build_upgrade_chains(wb, land_units, categories, prices)

    print("Building Levy Upgrade Chains sheet...")
    build_levy_upgrade_chains(wb, land_units, categories, prices)

    print("Building Category Stats sheet...")
    build_category_reference(wb, categories, prices)

    print("Building Light vs Heavy sheet...")
    build_light_vs_heavy(wb, land_units, categories)

    print("Building Artillery Barrage sheet...")
    build_artillery_barrage(wb, land_units, forts, prices)

    # --- Navy workbook (separate file) ---
    navy_wb = Workbook()
    navy_ws = navy_wb.active
    navy_ws.title = "placeholder"

    print("Building Navy sheet...")
    build_navy(navy_wb, naval_units, categories)

    print("Building Navy (Gold) sheet...")
    build_navy_gold(navy_wb, naval_units, categories, prices)

    print("Building Navy (Terrain) sheet...")
    build_navy_terrain(navy_wb, naval_units, categories)

    print("Building Navy (Terrain/Gold) sheet...")
    build_navy_terrain_gold(navy_wb, naval_units, categories, prices)

    print("Building Maritime (Sailor) sheet...")
    build_maritime_per_sailor(navy_wb, naval_units, categories, prices)

    print("Building Maritime (Gold) sheet...")
    build_maritime_per_gold(navy_wb, naval_units, categories, prices)

    print("Building Navy (Unique) sheet...")
    build_navy_unique(navy_wb, naval_units, categories)

    print("Building Navy (Unique Terrain) sheet...")
    build_navy_unique_terrain(navy_wb, naval_units, categories)

    print("Building Navy (Unique Terrain-Gold) sheet...")
    build_navy_unique_terrain_gold(navy_wb, naval_units, categories, prices)

    # Remove the placeholder sheet
    navy_wb.remove(navy_ws)

    # --- Food workbook (separate file) ---
    food_wb = Workbook()
    food_ws = food_wb.active
    food_ws.title = "placeholder"

    print("Building Food Reference sheet...")
    build_food_reference(food_wb, food_goods, food_buildings, building_caps, terrain_food_modifiers)

    print("Building Food Location Buildup sheet...")
    build_food_location_buildup(food_wb, food_goods, food_buildings, building_caps, terrain_food_modifiers)

    print("Building Build Order by RGO sheet...")
    build_food_build_order_by_rgo(food_wb, food_goods, food_buildings)

    print("Building Compound Build Order sheet...")
    build_food_compound_order(food_wb, food_goods, food_buildings)

    print("Building 100-Year Simulation sheet...")
    build_food_simulation(food_wb, food_goods, food_buildings)

    # Remove placeholder
    food_wb.remove(food_ws)

    # --- Economy workbook (separate file) ---
    econ_wb = Workbook()
    econ_ws = econ_wb.active
    econ_ws.title = "placeholder"

    print("Building Vassal Break-Even sheet...")
    build_vassal_breakeven(econ_wb)

    print("Building Annexation Batching sheet...")
    build_annex_batching(econ_wb)

    print("Building Pop Demands sheet...")
    build_pop_demands(econ_wb, pop_demands)

    # Remove placeholder
    econ_wb.remove(econ_ws)

    army_path = OUTPUT_DIR / "eu5_army_analysis.xlsx"
    navy_path = OUTPUT_DIR / "eu5_navy_analysis.xlsx"
    food_path = OUTPUT_DIR / "eu5_food_analysis.xlsx"
    econ_path = OUTPUT_DIR / "eu5_economy_analysis.xlsx"

    # Close Excel if it has our files open
    if sys.platform == "win32":
        import time
        subprocess.run(
            ["taskkill", "/F", "/IM", "EXCEL.EXE"],
            capture_output=True, timeout=5,
        )
        time.sleep(1)

    wb.save(army_path)
    print(f"\nSaved to: {army_path}")

    navy_wb.save(navy_path)
    print(f"Saved to: {navy_path}")

    food_wb.save(food_path)
    print(f"Saved to: {food_path}")

    econ_wb.save(econ_path)
    print(f"Saved to: {econ_path}")

    # Open all in Excel
    if sys.platform == "win32":
        os.startfile(army_path)
        os.startfile(navy_path)
        os.startfile(food_path)
        os.startfile(econ_path)


if __name__ == "__main__":
    main()
