"""Scrape EU5 Christian religion modifiers and export to Excel.

Covers every religion in group = christian, with Catholicism as the base for the papal
mechanics, and every modifier source that touches them: base religion stats, religious aspects,
papal bulls (Catholic curia actions + papal authority), religious laws, and religion-gated
advances. Named magnitude tokens (large_permanent_target_satisfaction, societal_value_monthly_move,
etc.) resolve to numbers via default_values.txt.

Run: python src/religion_analysis.py  ->  data/eu5_religion_analysis.xlsx
"""

import os
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from parser import parse_file, parse_directory

GAME_DIR = Path("C:/Steam/steamapps/common/Europa Universalis V/game")
COMMON_DIR = GAME_DIR / "in_game" / "common"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"

BASE_RELIGION = "catholic"

# Religions tagged `protestant` in christian.txt, used for is_protestant advance gating.
PROTESTANT = ["lutheran", "calvinist", "anglican", "hussite", "lollardy", "catharism", "waldensian"]

# Every religion in group = christian. Ordered for presentation; any added by a later patch is
# appended alphabetically so the set stays complete. ASPECT_RELIGIONS = those with aspect slots.
_CHRISTIAN = {k: v for k, v in parse_file(COMMON_DIR / "religions" / "christian.txt").items()
              if isinstance(v, dict) and v.get("group") == "christian"}
_ORDER = ["catholic", "orthodox", "lutheran", "calvinist", "anglican",
          "hussite", "lollardy", "catharism", "waldensian",
          "bogomilism", "bosnian_church", "paulicianism",
          "miaphysite", "nestorianism", "strigolniki"]
RELIGION_ORDER = [r for r in _ORDER if r in _CHRISTIAN] + sorted(set(_CHRISTIAN) - set(_ORDER))
ASPECT_RELIGIONS = {k for k, v in _CHRISTIAN.items() if v.get("religious_aspects")}

# Reform-era grouping (editorial, not a game field). Krstjani = the Bosnian Church.
REFORM_TYPE = {
    "catholic": "Catholic (base)",
    "orthodox": "Orthodox",
    "lutheran": "Reformation",
    "calvinist": "Reformation (Reformed)",
    "anglican": "Reformation",
    "hussite": "Proto-Protestant",
    "lollardy": "Proto-Protestant",
    "catharism": "Western Heresy",
    "waldensian": "Western Heresy",
    "bogomilism": "Dualist (Eastern)",
    "bosnian_church": "Bosnian Church (Krstjani)",
    "paulicianism": "Dualist (Eastern)",
    "miaphysite": "Miaphysite (Oriental)",
    "nestorianism": "Nestorian",
    "strigolniki": "Strigolniki",
}

# Top-level religion keys that are gameplay flags/values rather than modifiers.
RELIGION_FLAG_KEYS = [
    ("religious_aspects", "Aspect Slots"),
    ("needs_reform", "Needs Reform"),
    ("has_religious_influence", "Religious Influence"),
    ("has_religious_head", "Religious Head"),
    ("has_cardinals", "Cardinals"),
    ("has_canonization", "Canonization"),
    ("has_patriarchs", "Patriarchs"),
    ("has_autocephalous_patriarchates", "Autocephaly"),
    ("use_icons", "Icons"),
    ("ai_wants_convert", "AI Wants Convert"),
    ("tithe", "Tithe"),
]

# The Catholic-faith static modifiers, grouped by mechanic (main_menu/common/static_modifiers/religion.txt).
CURIA_BULLS = [
    "christiana_pietas", "illius_qui_se_pro_divini", "apostolicae_servitutis",
    "immensa_aeterni_dei", "libertas_ecclesiae", "dei_gratia_rex",
    "in_coena_domini", "benedictus_deus", "inter_gravissimas",
]
PAPAL_AUTHORITY_MODS = ["high_papal_authority", "low_papal_authority", "papacy_blocked_modifier"]
REFORM_DESIRE_MODS = [
    "limited_indulgences_religion", "tolerated_simony_religion", "western_schism_modifier",
    "papal_conclave_modifier", "saint_peter_basilica_being_built_with_indulgences",
    "unlimited_sale_indulgences_religion", "no_financial_indulgences_religion",
    "endorsed_simony_religion", "banned_simony_religion",
]

SOCIETAL_PREFIX = "monthly_towards_"

# Keys inside an advance entry that are not country modifiers.
ADVANCE_META_KEYS = {
    "age", "icon", "requires", "content_priority", "potential", "allow", "enable",
    "on_activate", "visible", "ai_will_do", "name", "desc", "unique", "position",
    "requires_advance", "requires_all", "requires_any", "unlock", "cost",
}

# ---- Styling (mirrors analyze.py) ----
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(size=10, italic=True, color="555555")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
BASE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def load_script_values() -> dict:
    """Numeric named magnitude tokens from default_values.txt."""
    sv = GAME_DIR / "main_menu" / "common" / "script_values" / "default_values.txt"
    raw = parse_file(sv) if sv.exists() else {}
    return {k: v for k, v in raw.items() if isinstance(v, (int, float)) and not isinstance(v, bool)}


TOKENS = load_script_values()


LOC_DIRS = [
    GAME_DIR / "main_menu" / "localization" / "english",
    GAME_DIR / "in_game" / "localization" / "english",
]


def pretty(ident: str) -> str:
    """Turn a snake_case identifier into a Title Case label."""
    return str(ident).replace("_", " ").title()


def _parse_loc(relpath: str, prefix: str = "") -> dict:
    """Parse a <name>_l_english.yml into {key: value}, stripping an optional key prefix. First dir wins."""
    out = {}
    for d in LOC_DIRS:
        path = d / relpath
        if not path.exists():
            continue
        for m in re.finditer(r'^\s*([\w.]+):\d*\s*"(.*)"\s*$', path.read_text(encoding="utf-8-sig"), re.MULTILINE):
            key = m.group(1)
            if prefix:
                if not key.startswith(prefix):
                    continue
                key = key[len(prefix):]
            out.setdefault(key, m.group(2))
    return out


# Modifier display names (global_crown_estate_power -> "Crown Power"), and bare-id display names for
# aspects, papal bulls, laws, law options, advances, and religions.
MODIFIER_LOC = _parse_loc("modifier_types_l_english.yml", prefix="MODIFIER_TYPE_NAME_")
NAME_LOC = {}
for _f in ("religion_l_english.yml", "events/religion/church_aspects_l_english.yml",
           "advances_l_english.yml", "laws_and_policies_l_english.yml"):
    for _k, _v in _parse_loc(_f).items():
        NAME_LOC.setdefault(_k, _v)
# Papal authority states and reform-desire modifiers are static modifiers, named under their own prefix.
for _k, _v in _parse_loc("static_modifiers_l_english.yml", prefix="STATIC_MODIFIER_NAME_").items():
    NAME_LOC.setdefault(_k, _v)


def _resolve(text: str) -> str:
    """Resolve $key$ macro refs and [concept|e] data functions inside a loc value to plain text."""
    if not text or ("$" not in text and "[" not in text):
        return text

    def macro(m):
        k = m.group(1)
        if k in MODIFIER_LOC:
            return MODIFIER_LOC[k]
        if k in NAME_LOC:
            return NAME_LOC[k]
        return pretty(k[len("game_concept_"):] if k.startswith("game_concept_") else k)

    text = re.sub(r"\$([\w.]+)\$", macro, text)
    text = re.sub(r"\[([^\]|]+)(?:\|[^\]]*)?\]", lambda m: m.group(1), text)
    return text.strip()


def loc_mod(key: str) -> str:
    """Localized modifier-type name, falling back to a prettified key."""
    return _resolve(MODIFIER_LOC[key]) if key in MODIFIER_LOC else pretty(key)


def loc_name(ident: str) -> str:
    """Localized name for an aspect / bull / law / option / advance / religion id, else prettified."""
    return _resolve(NAME_LOC[ident]) if ident in NAME_LOC else pretty(ident)


def religion_label(rid: str) -> str:
    return loc_name(rid)


def fmt_num(v) -> str:
    """Compact number formatting: drop trailing zeros, keep sign for readability."""
    if isinstance(v, bool):
        return "yes" if v else "no"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f"{v:g}"
    return str(v)


def _load_modifier_formats() -> dict:
    """Per-modifier display format (percent / already_percent / boolean / decimals) from the game."""
    d = GAME_DIR / "main_menu" / "common" / "modifier_type_definitions"
    out = {}
    if d.exists():
        for f in sorted(d.glob("*.txt")):
            for k, v in parse_file(f).items():
                if isinstance(v, dict):
                    out[k] = {
                        "percent": v.get("percent") is True,
                        "already_percent": v.get("already_percent") is True,
                        "boolean": v.get("boolean") is True,
                        "decimals": v.get("decimals"),
                    }
    return out


MODIFIER_FMT = _load_modifier_formats()


def _clean_num(x) -> str:
    """Round off float noise and drop trailing zeros."""
    x = round(x, 6)
    if x == int(x):
        return str(int(x))
    return f"{x:.6f}".rstrip("0").rstrip(".")


def format_value(key: str, value) -> str:
    """Format a modifier value the way the game does: percent modifiers as %, flats as plain numbers."""
    if isinstance(value, bool):
        return "yes" if value else "no"
    fmt = MODIFIER_FMT.get(key, {})
    if fmt.get("boolean"):
        return "yes" if value else "no"
    if fmt.get("percent"):
        return _clean_num(value * 100) + "%"
    if fmt.get("already_percent"):
        return _clean_num(value) + "%"
    if fmt.get("decimals") == 0:
        return str(int(round(value)))
    return _clean_num(value)


def resolve_num(val):
    """Resolve a modifier value to a float: raw numbers as-is, named tokens via default_values.txt."""
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        return float(TOKENS[val]) if val in TOKENS else None
    if isinstance(val, dict) and "__value__" in val:
        return resolve_num(val["__value__"])
    if isinstance(val, list):
        total, found = 0.0, False
        for v in val:
            r = resolve_num(v)
            if r is not None:
                total, found = total + r, True
        return total if found else None
    return None


def classify_modifiers(block: dict, exclude: set = frozenset()):
    """Split a modifier block into (numeric/flag mods, societal pushes).

    Returns (mods, societal) where mods is a list of (key, display_string) and societal is a
    list of value-axis names the block pushes toward.
    """
    mods, societal = [], []
    if not isinstance(block, dict):
        return mods, societal
    for key, val in block.items():
        if key in exclude or key == "game_data":
            continue
        if key.startswith(SOCIETAL_PREFIX):
            societal.append(key[len(SOCIETAL_PREFIX):])
            continue
        raw = val["__value__"] if isinstance(val, dict) and "__value__" in val else val
        if isinstance(raw, bool):
            mods.append((key, "yes" if raw else "no"))
            continue
        num = resolve_num(raw)
        if num is not None:
            mods.append((key, format_value(key, num)))
        elif isinstance(raw, str):
            mods.append((key, raw))
    return mods, societal


def mods_to_text(mods) -> str:
    return "\n".join(f"{loc_mod(k)} = {v}" for k, v in mods)


def numeric_modifiers(block: dict, exclude: set = frozenset()) -> dict:
    """Modifier key -> float (numeric, incl. societal monthly_towards_*) or True (flag). Skips meta keys."""
    out = {}
    if not isinstance(block, dict):
        return out
    for key, val in block.items():
        if key in exclude or key == "game_data":
            continue
        raw = val["__value__"] if isinstance(val, dict) and "__value__" in val else val
        if isinstance(raw, bool):
            if raw:
                out[key] = True
            continue
        num = resolve_num(raw)
        if num is not None:
            prev = out.get(key)
            out[key] = (prev if isinstance(prev, float) else 0.0) + num
    return out


def merge_numeric(target: dict, src: dict):
    """Sum src's numeric values into target; True flags win."""
    for k, v in src.items():
        if v is True:
            target[k] = True
        elif not isinstance(target.get(k), bool):
            target[k] = (target.get(k) or 0.0) + v


def _collect_has_aspect(node, out: set):
    """Collect every has_religious_aspect target id in a block."""
    if isinstance(node, dict):
        for key, val in node.items():
            if key == "has_religious_aspect":
                for v in (val if isinstance(val, list) else [val]):
                    if isinstance(v, str):
                        out.add(v.split(":")[-1])
            elif isinstance(val, (dict, list)):
                _collect_has_aspect(val, out)
    elif isinstance(node, list):
        for item in node:
            _collect_has_aspect(item, out)


def _negated_aspects(node) -> set:
    """Aspect ids referenced under a NOT/NOR in an enabled block = mutually exclusive with this aspect."""
    out = set()

    def walk(n):
        if isinstance(n, dict):
            for key, val in n.items():
                if key.lower() in ("not", "nor"):
                    _collect_has_aspect(val, out)
                elif isinstance(val, (dict, list)):
                    walk(val)
        elif isinstance(n, list):
            for item in n:
                walk(item)

    walk(node)
    return out


# ============================ Scrapers ============================

def scrape_religions() -> dict:
    """Base stats for every religion in RELIGION_ORDER from religions/christian.txt."""
    parsed = parse_file(COMMON_DIR / "religions" / "christian.txt")
    out = {}
    for rid in RELIGION_ORDER:
        data = parsed.get(rid, {})
        flags = {}
        for key, _ in RELIGION_FLAG_KEYS:
            if key in data:
                flags[key] = data[key]
        mods, _ = classify_modifiers(data.get("definition_modifier", {}))
        goods = []
        for gkey in ("goods_demand_modifier", "clergy_goods_demand_modifier"):
            gblock = data.get(gkey, {})
            if isinstance(gblock, dict):
                for g, v in gblock.items():
                    tag = "clergy " if gkey.startswith("clergy") else ""
                    goods.append((f"{tag}{g}", fmt_num(v)))
        enable = data.get("enable")
        enable = None if isinstance(enable, str) and enable.startswith("9999") else enable
        out[rid] = {
            "flags": flags,
            "mods": dict(mods),
            "num": numeric_modifiers(data.get("definition_modifier", {})),
            "goods": goods,
            "enable": enable,
            "language": data.get("language", ""),
        }
    return out


def _religion_refs(node, out: set):
    """Collect every `religion = <id>` reference in a block, skipping NOT/NOR subtrees."""
    if isinstance(node, dict):
        for key, val in node.items():
            if key.lower() in ("not", "nor"):
                continue
            if key == "religion":
                for v in (val if isinstance(val, list) else [val]):
                    if isinstance(v, str):
                        out.add(v.split(":")[-1])
            elif isinstance(val, (dict, list)):
                _religion_refs(val, out)
    elif isinstance(node, list):
        for item in node:
            _religion_refs(item, out)


_NATION_GATE_KEYS = {"has_or_had_tag", "tag", "culture", "primary_culture",
                     "has_culture_group", "has_culture", "culture_group", "region", "area"}


def _is_nation_gated(node) -> bool:
    """True if a potential/allow block gates by a specific tag, culture, or region."""
    if isinstance(node, dict):
        for key, val in node.items():
            if key in _NATION_GATE_KEYS:
                return True
            if isinstance(val, (dict, list)) and _is_nation_gated(val):
                return True
    elif isinstance(node, list):
        return any(_is_nation_gated(item) for item in node)
    return False


def _has_trigger_yes(node, name: str) -> bool:
    """True if `<name> = yes` appears in a block, skipping NOT/NOR subtrees (e.g. is_protestant)."""
    if isinstance(node, dict):
        for key, val in node.items():
            if key.lower() in ("not", "nor"):
                continue
            if key == name and val is True:
                return True
            if isinstance(val, (dict, list)) and _has_trigger_yes(val, name):
                return True
    elif isinstance(node, list):
        return any(_has_trigger_yes(item, name) for item in node)
    return False


def scrape_aspects() -> list:
    """Every religious aspect at least one Christian religion in the set can take."""
    rows = []
    aspect_dir = COMMON_DIR / "religious_aspects"
    for f in sorted(aspect_dir.glob("*.txt")):
        if f.stem == "readme":
            continue
        for aid, data in parse_file(f).items():
            if not isinstance(data, dict):
                continue
            religions = set()
            for v in (data["religion"] if isinstance(data.get("religion"), list) else [data.get("religion")]):
                if isinstance(v, str):
                    religions.add(v.split(":")[-1])
            eligible = religions & ASPECT_RELIGIONS
            if not eligible:
                continue
            mods, societal = classify_modifiers(data.get("modifier", {}))
            opinions = data.get("opinions", {})
            likes = [k for k, v in opinions.items() if resolve_num(v) and resolve_num(v) > 0 and k != aid]
            dislikes = [k for k, v in opinions.items() if resolve_num(v) and resolve_num(v) < 0]
            notes = []
            if "enabled" in data or "visible" in data:
                notes.append("conditional (see game for unlock/exclusion rules)")
            rows.append({
                "id": aid,
                "source": f.stem,
                "eligible": eligible,
                "all_religions": religions,
                "mods": mods,
                "num": numeric_modifiers(data.get("modifier", {})),
                "excludes": _negated_aspects(data.get("enabled")),
                "societal": societal,
                "likes": likes,
                "dislikes": dislikes,
                "notes": "; ".join(notes),
            })
    rows.sort(key=lambda r: (-len(r["eligible"]), r["id"]))
    return rows


def _find_add_religion_modifier(node):
    """Depth-first search for an add_religion_modifier effect; returns (modifier, years) or (None, None)."""
    if isinstance(node, dict):
        if "add_religion_modifier" in node:
            arm = node["add_religion_modifier"]
            if isinstance(arm, dict):
                return arm.get("modifier"), arm.get("years")
        for val in node.values():
            r = _find_add_religion_modifier(val)
            if r[0]:
                return r
    elif isinstance(node, list):
        for item in node:
            r = _find_add_religion_modifier(item)
            if r[0]:
                return r
    return None, None


def scrape_bulls() -> dict:
    """Catholic static religion modifiers grouped by mechanic, with bull duration/cost from resolutions."""
    static = parse_file(GAME_DIR / "main_menu" / "common" / "static_modifiers" / "religion.txt")

    def mods_for(name):
        return classify_modifiers(static.get(name, {}))[0]

    # Bull duration + cost from the resolution files that grant each religion modifier.
    prices = parse_file(COMMON_DIR / "prices" / "00_hardcoded.txt")
    curia_cost = prices.get("propose_curia_action", {})
    cost_text = ", ".join(f"{fmt_num(v)} {pretty(k)}" for k, v in curia_cost.items()) if isinstance(curia_cost, dict) else ""
    bull_meta = {}  # modifier name -> {"years", "cost"}
    for f in sorted((COMMON_DIR / "resolutions").glob("*.txt")):
        data = parse_file(f)
        for res in data.values():
            if not isinstance(res, dict) or res.get("international_organization_type") != "catholic_church":
                continue
            mod, years = _find_add_religion_modifier(res.get("effect"))
            if mod:
                bull_meta[mod] = {"years": years, "cost": cost_text}

    def rows_for(names):
        rows = []
        for name in names:
            if name not in static:
                continue
            meta = bull_meta.get(name, {})
            rows.append({
                "id": name,
                "mods": mods_for(name),
                "num": numeric_modifiers(static.get(name, {})),
                "years": meta.get("years"),
                "cost": meta.get("cost", ""),
            })
        return rows

    return {
        "curia": rows_for(CURIA_BULLS),
        "authority": rows_for(PAPAL_AUTHORITY_MODS),
        "reform_desire": rows_for(REFORM_DESIRE_MODS),
    }


def scrape_laws() -> list:
    """Options within each religious law available to a religion in the set (Muslim-only laws skipped)."""
    parsed = parse_file(COMMON_DIR / "laws" / "00_religious.txt")
    meta_keys = {"law_category", "law_religion_group", "potential", "locked", "on_activate",
                 "allow", "unique", "name", "desc"}
    rel_set = set(RELIGION_ORDER)
    rows = []
    for law_id, law in parsed.items():
        if not isinstance(law, dict):
            continue
        grp = law.get("law_religion_group")
        if isinstance(grp, dict):
            law_rel = set(grp.get("__bare_values__", []))
        elif isinstance(grp, str):
            law_rel = {grp}
        else:
            law_rel = set()
        if law_rel:
            avail = law_rel & rel_set
            if not avail:
                continue
            available = ", ".join(religion_label(r) for r in RELIGION_ORDER if r in avail)
        else:
            avail = set(rel_set)
            available = "All religions"
        for opt_id, opt in law.items():
            if opt_id in meta_keys or not isinstance(opt, dict):
                continue
            mods, societal, conditional, num = [], [], False, {}
            cm = opt.get("country_modifier")
            for block in (cm if isinstance(cm, list) else [cm]):
                if not isinstance(block, dict):
                    continue
                if "potential_trigger" in block:
                    conditional = True
                m, s = classify_modifiers(block, exclude={"potential_trigger"})
                mods.extend(m)
                societal.extend(s)
                merge_numeric(num, numeric_modifiers(block, exclude={"potential_trigger"}))
            estates = opt.get("estate_preferences", {})
            estate_list = estates.get("__bare_values__", []) if isinstance(estates, dict) else []
            notes = []
            if opt.get("unique"):
                notes.append("unique / country-specific")
            if conditional:
                notes.append("some effects tag/culture-conditional")
            rows.append({
                "law": law_id,
                "option": opt_id,
                "available": available,
                "avail_religions": avail,
                "national": _is_nation_gated(opt.get("potential")) or _is_nation_gated(opt.get("allow")),
                "mods": mods,
                "num": num,
                "societal": societal,
                "estates": [pretty(e.replace("_estate", "")) for e in estate_list],
                "notes": "; ".join(notes),
            })
    return rows


def scrape_advances() -> list:
    """Advances gated to one of the Christian religions in the set, with their modifiers."""
    religions = set(RELIGION_ORDER)
    rows = []
    for f in sorted((COMMON_DIR / "advances").glob("*.txt")):
        for aid, data in parse_file(f).items():
            if not isinstance(data, dict):
                continue
            refs = set()
            _religion_refs(data.get("potential"), refs)
            gated = refs & religions
            if _has_trigger_yes(data.get("potential"), "is_protestant"):
                gated |= set(PROTESTANT)
            if not gated:
                continue
            mods, societal = classify_modifiers(data, exclude=ADVANCE_META_KEYS)
            if not mods and not societal:
                continue
            age = data.get("age", "")
            rows.append({
                "id": aid,
                "source": f.stem.replace("country_", "").replace("culture_", "").replace("religion_", ""),
                "age": age.replace("age_", "").replace("_", " ") if isinstance(age, str) else "",
                "gated": gated,
                "generic": not _is_nation_gated(data.get("potential")),
                "mods": mods,
                "num": numeric_modifiers(data, exclude=ADVANCE_META_KEYS),
                "societal": societal,
            })
    rows.sort(key=lambda r: (r["age"], r["id"]))
    return rows


# ============================ Sheet builders ============================

def _header(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _title(ws, title, subtitle):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT


def _autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    _title(ws, "EU5 Christian Religion Modifiers",
           "Every religion in group = christian, with Catholicism as the base for the papal mechanics. "
           "Generated from vanilla game data.")
    lines = [
        ("", ""),
        ("Religions covered", f"{len(RELIGION_ORDER)} faiths in group = christian"),
        ("  Base", "Catholicism (papal bulls, cardinals, reform desire)"),
        ("  Orthodox", "Orthodoxy"),
        ("  Reformation", "Lutheranism, Calvinism (Reformed), Anglicanism"),
        ("  Proto-Protestant", "Hussite, Lollardy"),
        ("  Western Heresy", "Catharism, Waldensianism"),
        ("  Eastern / Dualist", "Bogomilism, Bosnian Church (Krstjani), Paulicianism"),
        ("  Other Eastern", "Miaphysite, Nestorianism, Strigolniki"),
        ("", ""),
        ("Sheets", ""),
        ("  Base Stats", "Per-religion flags and definition_modifier (the modifiers you get just for being that faith)"),
        ("  Religious Aspects", "The pickable aspects (each aspect-holding faith fills 3 slots) and their modifiers"),
        ("  Papal Bulls", "Catholic-only: curia-action bulls, papal authority states, reform-desire modifiers"),
        ("  Religious Laws", "Religious law options available to these faiths (Catholic, Orthodox, and universal)"),
        ("  Religion Advances", "Advances (technology) gated to one of these religions"),
        ("  Max Achievable", "Modifiers down the left, faiths across the top: the highest value each reaches through some valid setup"),
        ("", ""),
        ("Notes", ""),
        ("  Aspect slots", "10 of the faiths have religious_aspects = 3; Catholicism, Orthodoxy, and a few Eastern faiths have none."),
        ("  Catholicism", "Has no aspect slots; its equivalent customization is papal bulls + doctrines (see Papal Bulls)."),
        ("  Societal push", "monthly_towards_X modifiers nudge a country's societal values; shown as a 'Societal Push' column, not a number."),
        ("  Values", "Named magnitudes (e.g. large_permanent_target_satisfaction) resolved to numbers via default_values.txt."),
        ("  Reform-era labels", "Editorial grouping, not a game field."),
    ]
    r = 4
    for a, b in lines:
        if a and not b and not a.startswith("  "):
            cell = ws.cell(row=r, column=1, value=a)
            cell.font = Font(bold=True, size=11)
        else:
            ws.cell(row=r, column=1, value=a).font = Font(bold=bool(a and not a.startswith("  ")))
            ws.cell(row=r, column=2, value=b).alignment = TOP
        r += 1
    _autosize(ws, {"A": 22, "B": 95})


def build_base_stats(wb, religions):
    ws = wb.create_sheet("Base Stats")
    _title(ws, "Base Religion Stats",
           "Flags and definition_modifier per faith. Catholicism (base) highlighted. Blank = modifier absent.")
    flag_headers = [label for _, label in RELIGION_FLAG_KEYS]
    mod_keys = []
    for rid in RELIGION_ORDER:
        for k in religions[rid]["mods"]:
            if k not in mod_keys:
                mod_keys.append(k)
    goods_present = any(religions[rid]["goods"] for rid in RELIGION_ORDER)
    headers = (["Religion", "Type", "Enable"] + flag_headers
               + [loc_mod(k) for k in mod_keys] + (["Goods Demand"] if goods_present else []))
    hrow = 4
    _header(ws, hrow, headers)
    for i, rid in enumerate(RELIGION_ORDER):
        r = hrow + 1 + i
        data = religions[rid]
        row_vals = [religion_label(rid), REFORM_TYPE.get(rid, "Christian"), data["enable"] if data["enable"] else ""]
        for key, _ in RELIGION_FLAG_KEYS:
            v = data["flags"].get(key)
            row_vals.append(fmt_num(v) if v is not None else "")
        for k in mod_keys:
            row_vals.append(data["mods"].get(k, ""))
        if goods_present:
            row_vals.append("\n".join(f"{g} = {v}" for g, v in data["goods"]))
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c > 1 else TOP
            if rid == BASE_RELIGION:
                cell.fill = BASE_FILL
    ws.freeze_panes = ws.cell(row=hrow + 1, column=2)
    widths = {"A": 15, "B": 20, "C": 9}
    for idx in range(len(flag_headers)):
        widths[get_column_letter(4 + idx)] = 11
    for idx in range(len(mod_keys)):
        widths[get_column_letter(4 + len(flag_headers) + idx)] = 15
    if goods_present:
        widths[get_column_letter(len(headers))] = 20
    _autosize(ws, widths)


def build_aspects(wb, aspects):
    ws = wb.create_sheet("Religious Aspects")
    _title(ws, "Religious Aspects",
           "Pickable aspects for the aspect-holding faiths (each fills 3 slots). "
           "X marks which faiths may take each aspect. Sorted by breadth of availability.")
    aspect_cols = [r for r in RELIGION_ORDER if r in ASPECT_RELIGIONS]
    short = {r: loc_name(r)[:4] for r in aspect_cols}
    headers = ["Aspect", "Source"] + [short[r] for r in aspect_cols] + \
              ["Modifiers", "Societal Push", "Synergy / Conflict", "Notes"]
    hrow = 4
    _header(ws, hrow, headers)
    for i, a in enumerate(aspects):
        r = hrow + 1 + i
        syn = []
        if a["likes"]:
            syn.append("+ " + ", ".join(loc_name(x) for x in a["likes"]))
        if a["dislikes"]:
            syn.append("- " + ", ".join(loc_name(x) for x in a["dislikes"]))
        row_vals = [loc_name(a["id"]), a["source"]]
        for rid in aspect_cols:
            row_vals.append("X" if rid in a["eligible"] else "")
        row_vals += [
            mods_to_text(a["mods"]),
            ", ".join(pretty(s) for s in a["societal"]),
            "\n".join(syn),
            a["notes"],
        ]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if 2 < c <= 2 + len(aspect_cols) else TOP
            if i % 2:
                cell.fill = ALT_FILL
    ws.freeze_panes = ws.cell(row=hrow + 1, column=2)
    widths = {"A": 22, "B": 12}
    for idx in range(len(aspect_cols)):
        widths[get_column_letter(3 + idx)] = 6
    base = 3 + len(aspect_cols)
    _autosize(ws, {**widths, get_column_letter(base): 40, get_column_letter(base + 1): 20,
                   get_column_letter(base + 2): 26, get_column_letter(base + 3): 24})


def _section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")


def build_bulls(wb, bulls):
    ws = wb.create_sheet("Papal Bulls")
    _title(ws, "Papal Bulls & Catholic Faith Mechanics (Base)",
           "Catholicism only. Curia-action bulls are temporary religion-wide modifiers the Pope enacts; "
           "papal authority swings the faith's whole standing; reform desire drives the Reformation.")
    headers = ["Modifier", "Effects", "Duration", "Cost"]
    ncols = len(headers)
    r = 4
    sections = [
        ("Curia Action Bulls (enacted via the Catholic Church, apply to the whole religion)", bulls["curia"]),
        ("Papal Authority States (applied automatically from the papal_authority level)", bulls["authority"]),
        ("Reform Desire Modifiers (Catholic faith drift toward the Reformation)", bulls["reform_desire"]),
    ]
    for title, rows in sections:
        _section(ws, r, title, ncols)
        r += 1
        _header(ws, r, headers)
        r += 1
        for a in rows:
            dur = f"{fmt_num(a['years'])} years" if a.get("years") else "while active"
            vals = [loc_name(a["id"]), mods_to_text(a["mods"]), dur, a.get("cost", "")]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = THIN_BORDER
                cell.alignment = TOP if c == 2 else CENTER
            r += 1
        r += 1
    _autosize(ws, {"A": 30, "B": 45, "C": 12, "D": 22})


def build_laws(wb, laws):
    ws = wb.create_sheet("Religious Laws")
    _title(ws, "Religious Laws",
           "Law options available to these faiths. Catholic-only laws are the base's tools; "
           "censorship and slave conversion are open to every religion.")
    headers = ["Law", "Option", "Available To", "Modifiers", "Societal Push", "Estate Pref.", "Notes"]
    hrow = 4
    _header(ws, hrow, headers)
    r = hrow + 1
    last_law = None
    for a in laws:
        law_disp = loc_name(a["law"]) if a["law"] != last_law else ""
        last_law = a["law"]
        vals = [
            law_disp, loc_name(a["option"]), a["available"],
            mods_to_text(a["mods"]),
            ", ".join(pretty(s) for s in a["societal"]),
            ", ".join(a["estates"]),
            a["notes"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = TOP
        r += 1
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    _autosize(ws, {"A": 18, "B": 24, "C": 15, "D": 42, "E": 18, "F": 16, "G": 28})


def build_advances(wb, advances):
    ws = wb.create_sheet("Religion Advances")
    _title(ws, "Religion-Gated Advances (Technology)",
           "Advances whose potential requires one of these faiths: the dedicated Protestant line (is_protestant), "
           "the Catholic line (religion:catholic), and national/cultural advances that key off religion. "
           "Advances open to every religion (most of the Age of Reformation) are omitted.")
    headers = ["Advance", "Source", "Age", "Religions", "Modifiers", "Societal Push"]
    hrow = 4
    _header(ws, hrow, headers)
    for i, a in enumerate(advances):
        r = hrow + 1 + i
        vals = [
            loc_name(a["id"]), a["source"], a["age"],
            ", ".join(religion_label(x) for x in sorted(a["gated"])),
            mods_to_text(a["mods"]),
            ", ".join(pretty(s) for s in a["societal"]),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = TOP
            if i % 2:
                cell.fill = ALT_FILL
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    _autosize(ws, {"A": 26, "B": 12, "C": 14, "D": 26, "E": 40, "F": 20})


def build_max_achievable(wb, religions, aspects, bulls, laws, advances):
    """One row per faith, one column per modifier, each cell the highest value reachable by any
    valid single-modifier configuration. Columns are independent: no build hits every max at once.

    Per source: base faith modifier is always on; the best <=3 aspects (respecting mutual exclusivity
    and the 3 slots) are stacked; each religious law contributes its best faith-wide option; Catholicism
    adds its single best papal bull plus high papal authority; every faith-wide advance is summed. Only
    positive optional contributions count (we are maximizing), so a cell can go below zero only when the
    base faith modifier is itself negative. Nation/culture-specific advances and law options are excluded,
    since no single nation of the faith can stack them; this keeps every cell reachable by one build.
    """
    from itertools import combinations

    high_auth = [b for b in bulls["authority"] if b["id"] == "high_papal_authority"]
    bull_sources = bulls["curia"] + high_auth

    # Column universe, split into numeric and flag modifiers, from the sources that feed a max.
    numeric_keys, flag_keys = set(), set()

    def register(num):
        for k, v in num.items():
            (flag_keys if v is True else numeric_keys).add(k)

    for rid in RELIGION_ORDER:
        register(religions[rid]["num"])
    for a in aspects:
        register(a["num"])
    for b in bull_sources:
        register(b["num"])
    for o in laws:
        if not o["national"]:
            register(o["num"])
    for a in advances:
        if a["generic"]:
            register(a["num"])
    flag_keys -= numeric_keys
    columns = sorted(numeric_keys | flag_keys)

    law_by_id = {}
    for o in laws:
        law_by_id.setdefault(o["law"], []).append(o)
    excl = {a["id"]: set(a["excludes"]) for a in aspects}
    # An institutionalized upgrade replaces its base aspect, so treat them as mutually exclusive.
    for a in aspects:
        if a["id"].startswith("instituionalized_"):
            base = a["id"][len("instituionalized_"):]
            excl[a["id"]].add(base)
            excl.setdefault(base, set()).add(a["id"])

    def conflict(i, j):
        return j in excl.get(i, ()) or i in excl.get(j, ())

    def aspect_max(rel, key):
        cands = sorted(
            ((a["num"][key], a["id"]) for a in aspects
             if rel in a["eligible"] and isinstance(a["num"].get(key), float) and a["num"][key] > 0),
            reverse=True)[:8]
        best = 0.0
        for k in range(1, 4):
            for combo in combinations(cands, k):
                ids = [c[1] for c in combo]
                if any(conflict(ids[i], ids[j]) for i in range(len(ids)) for j in range(i + 1, len(ids))):
                    continue
                best = max(best, sum(c[0] for c in combo))
        return best

    def law_max(rel, key):
        total = 0.0
        for opts in law_by_id.values():
            if rel not in opts[0]["avail_religions"]:
                continue
            best = max((o["num"][key] for o in opts
                        if not o["national"] and isinstance(o["num"].get(key), float)), default=0.0)
            total += max(0.0, best)
        return total

    def advance_max(rel, key):
        return sum(max(0.0, a["num"][key]) for a in advances
                   if rel in a["gated"] and a["generic"] and isinstance(a["num"].get(key), float))

    def bull_max(key):
        curia_best = max((b["num"][key] for b in bulls["curia"]
                          if isinstance(b["num"].get(key), float) and b["num"][key] > 0), default=0.0)
        auth = sum(b["num"][key] for b in high_auth
                   if isinstance(b["num"].get(key), float) and b["num"][key] > 0)
        return curia_best + auth

    def numeric_cell(rel, key):
        base = religions[rel]["num"].get(key)
        total = (base if isinstance(base, float) else 0.0)
        total += aspect_max(rel, key) + law_max(rel, key) + advance_max(rel, key)
        if rel == BASE_RELIGION:
            total += bull_max(key)
        return round(total, 6)

    def flag_cell(rel, key):
        if religions[rel]["num"].get(key) is True:
            return True
        if any(rel in a["eligible"] and a["num"].get(key) is True for a in aspects):
            return True
        for opts in law_by_id.values():
            if rel not in opts[0]["avail_religions"]:
                continue
            if any(o["num"].get(key) is True for o in opts if not o["national"]):
                return True
        if any(rel in a["gated"] and a["generic"] and a["num"].get(key) is True for a in advances):
            return True
        if rel == BASE_RELIGION and any(b["num"].get(key) is True for b in bull_sources):
            return True
        return False

    # Modifiers down the left (rows), faiths across the top (columns). Numeric modifiers first, then flags,
    # each sorted by localized name. All text horizontal.
    numeric_rows = sorted((k for k in columns if k not in flag_keys), key=lambda k: loc_mod(k).lower())
    flag_rows = sorted((k for k in columns if k in flag_keys), key=lambda k: loc_mod(k).lower())
    row_keys = numeric_rows + flag_rows

    ws = wb.create_sheet("Max Achievable")
    _title(ws, "Max Achievable Modifiers",
           "Modifiers down the left, faiths across the top. Each cell is the highest total that faith reaches "
           "through some valid setup (base + best 3 aspects + best faith-wide law per category + best single papal "
           "bull + high papal authority + all faith-wide advances). Cells are independent - no single build reaches "
           "every max at once. Nation/culture-specific advances and law options are excluded (see Advances/Laws "
           "sheets). Values are literal maxima; percentages shown as %; flags show 'yes'; blank = 0 / no access.")
    hrow, subrow, dstart = 4, 5, 6

    hcell = ws.cell(row=hrow, column=1, value="Modifier")
    hcell.font, hcell.fill, hcell.border = HEADER_FONT_WHITE, HEADER_FILL, THIN_BORDER
    hcell.alignment = Alignment(horizontal="left", vertical="center")
    tcell = ws.cell(row=subrow, column=1, value="Faith type")
    tcell.font, tcell.border = Font(size=8, italic=True, color="FFFFFF"), THIN_BORDER
    tcell.fill = HEADER_FILL
    tcell.alignment = Alignment(horizontal="left")
    for j, rid in enumerate(RELIGION_ORDER):
        col = 2 + j
        c1 = ws.cell(row=hrow, column=col, value=religion_label(rid))
        c1.font, c1.fill, c1.border, c1.alignment = HEADER_FONT_WHITE, HEADER_FILL, THIN_BORDER, CENTER
        c2 = ws.cell(row=subrow, column=col, value=REFORM_TYPE.get(rid, "Christian"))
        c2.font, c2.border, c2.alignment = Font(size=8, italic=True, color="555555"), THIN_BORDER, CENTER
        if rid == BASE_RELIGION:
            c2.fill = BASE_FILL

    for i, key in enumerate(row_keys):
        r = dstart + i
        name = ws.cell(row=r, column=1, value=loc_mod(key))
        name.border, name.alignment = THIN_BORDER, Alignment(horizontal="left", vertical="center", wrap_text=True)
        name.font = Font(bold=True)
        if i % 2:
            name.fill = ALT_FILL
        for j, rid in enumerate(RELIGION_ORDER):
            if key in flag_keys:
                val = "yes" if flag_cell(rid, key) else ""
            else:
                v = numeric_cell(rid, key)
                val = "" if v == 0 else format_value(key, v)
            cell = ws.cell(row=r, column=2 + j, value=val)
            cell.border, cell.alignment = THIN_BORDER, CENTER
            if rid == BASE_RELIGION:
                cell.fill = BASE_FILL
            elif i % 2:
                cell.fill = ALT_FILL
    ws.freeze_panes = ws.cell(row=dstart, column=2)
    ws.column_dimensions["A"].width = 34
    for j in range(len(RELIGION_ORDER)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13
    return len(row_keys)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Scraping religions...")
    religions = scrape_religions()
    print("Scraping aspects...")
    aspects = scrape_aspects()
    print("Scraping papal bulls...")
    bulls = scrape_bulls()
    print("Scraping laws...")
    laws = scrape_laws()
    print("Scraping advances...")
    advances = scrape_advances()
    print(f"  {len(religions)} religions, {len(aspects)} aspects, "
          f"{sum(len(v) for v in bulls.values())} bull/faith modifiers, "
          f"{len(laws)} law options, {len(advances)} religion-gated advances")

    wb = Workbook()
    build_overview(wb)
    build_base_stats(wb, religions)
    build_aspects(wb, aspects)
    build_bulls(wb, bulls)
    build_laws(wb, laws)
    build_advances(wb, advances)
    ncols = build_max_achievable(wb, religions, aspects, bulls, laws, advances)
    print(f"  Max Achievable sheet: {ncols} modifier columns")

    out = OUTPUT_DIR / "eu5_religion_analysis.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        if sys.platform == "win32":
            subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], capture_output=True, timeout=5)
            import time
            time.sleep(1)
        wb.save(out)
    print(f"\nSaved to: {out}")
    if sys.platform == "win32":
        os.startfile(out)


if __name__ == "__main__":
    main()
